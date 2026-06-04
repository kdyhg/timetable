from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import mimetypes
import os
import random
import re
import sys
import time
import uuid
from collections import Counter, defaultdict
from copy import deepcopy
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).parent.resolve()
WEB_DIR = ROOT / "web"
RUNTIME_DATA_DIR = os.environ.get("TIMETABLE_DATA_DIR")
if not RUNTIME_DATA_DIR and os.environ.get("VERCEL"):
    RUNTIME_DATA_DIR = "/tmp/timetable-data"
DATA_DIR = Path(RUNTIME_DATA_DIR).resolve() if RUNTIME_DATA_DIR else ROOT / "data"
IMPORT_DIR = DATA_DIR / "imports"
SOLVE_SESSION_DIR = DATA_DIR / "solve_sessions"
LAST_SCHEDULE_FILE = DATA_DIR / "last_schedule.json"
OPERATION_LOG_FILE = DATA_DIR / "operation_log.jsonl"

DEFAULT_DAYS = ["월", "화", "수", "목", "금"]
DEFAULT_MAX_PERIOD = 7
TEMPLATE_INPUT_ROWS = 200
TEMPLATE_CLASS_COLUMN_COUNT = 30
SAMPLE_CLASS_COLUMNS = ["1-1", "1-2", "1-3", "2-1", "2-2", "2수학A", "2기계1", "3-1"]
WEEKDAY_ORDER = ["월", "화", "수", "목", "금", "토", "일"]

SHEET_SPECS = [
    (
        "기본설정",
        ["항목", "값", "설명"],
        [
            ("학교명", "", "출력물과 NEIS 파일에 사용할 학교명"),
            ("학년도", "2026", "숫자 4자리"),
            ("수업요일", "월,화,수,목,금", "쉼표로 구분"),
            ("일일최대교시", "7", "가장 긴 수업일의 교시 수"),
            ("점심시간후교시", "4", "예: 4 입력 시 4교시 후 점심"),
            ("점심시간보호", "Y", "Y이면 점심 전후 교시를 같은 교사에게 동시에 배정하지 않음"),
            ("최대연강허용", "3", "교사별 같은 날 연속 수업 최대 허용 교시"),
            ("균등분배강도", "soft", "off/soft/hard 중 선택, 현재는 soft 권장"),
            ("AI우선사용", "Y", "AI 제안과 진단을 자동배정 흐름에 표시"),
            ("시간표제목", "AI 공동작성 시간표", "작업 이름"),
        ],
    ),
    (
        "교사",
        ["교사명", "비고", "식사후교시", "배정금지"],
        [],
    ),
    (
        "학급-계열",
        ["학급명", "학년", "계열", "담임교사명", "요일별시수", "가상학급여부"],
        [],
    ),
    (
        "고정 일과",
        ["대상유형", "대상명", "교시", "월", "화", "수", "목", "금", "유형", "임장교사명", "설명"],
        [],
    ),
    (
        "과목",
        ["과목명", "단축명", "NEIS과목명", "유사과목그룹"],
        [],
    ),
    (
        "교사별 시수표",
        ["교사명", "과목명", "연속패턴", "특별실명", "동시그룹", "복수교사그룹"],
        [],
    ),
    (
        "특별실",
        ["특별실명", "배정금지"],
        [],
    ),
    (
        "동시-합반-분반",
        ["그룹명", "유형", "학급명목록", "과목명", "설명"],
        [],
    ),
    (
        "연속수업",
        ["과목명", "학급명", "연속시수", "횟수"],
        [],
    ),
    (
        "복수교사",
        ["그룹명", "과목명", "학급명", "교사명목록"],
        [],
    ),
    (
        "배정금지-희망조건",
        ["대상유형", "대상명", "조건유형", "요일", "교시", "강도", "우선순위", "설명"],
        [],
    ),
    (
        "NEIS 코드",
        ["과목명", "NEIS과목명", "NEIS과목코드", "교사명", "NEIS교사명"],
        [],
    ),
]

SPECS_BY_NAME = {name: headers for name, headers, _ in SHEET_SPECS}

REQUIRED_FIELDS = {
    "기본설정": {"항목", "값"},
    "교사": {"교사명"},
    "학급-계열": {"학급명", "학년"},
    "고정 일과": {"대상유형", "교시"},
    "과목": {"과목명", "단축명"},
    "교사별 시수표": {"교사명", "과목명"},
    "특별실": {"특별실명"},
    "동시-합반-분반": {"그룹명", "유형", "학급명목록"},
    "연속수업": {"과목명", "학급명", "연속시수"},
    "복수교사": {"그룹명", "과목명", "학급명", "교사명목록"},
    "배정금지-희망조건": {"대상유형", "대상명", "조건유형"},
    "NEIS 코드": {"과목명", "NEIS과목명"},
}

SHEET_GUIDANCE = {
    "기본설정": "왼쪽 입력 영역에서 B열 '값'만 수정하세요. A열 항목명과 C열 설명은 유지합니다.",
    "교사": "시간표에 참여하는 모든 교사를 한 행에 한 명씩 이름으로 입력합니다. 내부 교사코드는 업로드 때 자동 생성됩니다.",
    "학급-계열": "여기서 학급은 시간표 결과가 아니라 학교에 이미 편성된 반(예: 1-1, 2기계1)입니다. 요일별시수는 기본설정의 수업요일 순서대로 입력합니다.",
    "고정 일과": "시수배당표에는 없지만 시간표에는 있어야 하는 HR, 자습, 창체, 행사 등을 학년/계열/학급별로 고정합니다.",
    "과목": "정식 과목명, 단축명, NEIS 과목명을 관리합니다. 유사과목그룹은 같은 날 배정 회피에 사용됩니다.",
    "교사별 시수표": "교사명과 과목명을 한 번 쓰고, 오른쪽 학급명 열에 주당 시수 숫자만 입력합니다. 학급명 열은 학급-계열 시트의 반 이름입니다.",
    "특별실": "컴퓨터실, 체육관, 실험실처럼 동시에 하나의 수업만 가능한 공간을 입력합니다.",
    "동시-합반-분반": "수준별 이동수업, 합반수업, 분반수업처럼 여러 학급이 함께 움직이는 구조를 입력합니다.",
    "연속수업": "미술, 실습처럼 2시간 이상 붙여 배정해야 하는 수업을 입력합니다.",
    "복수교사": "원어민 협력수업처럼 한 학급 한 과목에 2명 이상 교사가 동시에 들어가는 경우를 입력합니다.",
    "배정금지-희망조건": "교사, 학급, 과목, 특별실의 배정금지/이동금지/희망 조건을 입력합니다.",
    "NEIS 코드": "NEIS 일괄파일 출력을 위해 과목명/과목코드와 교사명을 매칭합니다.",
}

EXAMPLE_ROWS = {
    "기본설정": [
        ("정규 시간표", ["수업요일", "월,화,수,목,금", "월~금 정규수업"], "정규 주간시간표는 요일명을 쉼표로 입력합니다."),
        ("보충수업", ["수업요일", "1/3,1/4,1/5,1/7", "방학 중 실제 수업일"], "요일 대신 날짜를 넣을 수 있습니다."),
        ("0교시 표기", ["일일최대교시", "8", "0교시 포함 8칸"], "출력 교시명은 별도 기능에서 확장할 수 있습니다."),
    ],
    "교사": [
        ("일반 교사", ["김하늘", "국어과", "", ""], "식사시간과 금지시간이 없으면 뒤 칸은 비워둡니다."),
        ("점심시간 다름", ["박민수", "순회", "3", ""], "기본 점심시간과 다른 교사는 식사후교시를 입력합니다."),
        ("요일 제한", ["이서연", "겸임", "", "금 6-7"], "간단 메모용입니다. 엄격 조건은 배정금지-희망조건 시트에 입력합니다."),
    ],
    "학급-계열": [
        ("일반학급", ["1-1", "1", "공통", "김하늘", "7,7,6,7,7", "N"], "수업요일이 월~금이면 수요일만 6교시, 나머지는 7교시라는 뜻입니다."),
        ("특성화 계열", ["2기계1", "2", "기계", "정다은", "7,7,6,7,6", "N"], "계열명을 넣으면 학과별 출력과 검토가 쉬워집니다."),
        ("가상/추가학급", ["2수학A", "2", "수준별", "", "3", "Y"], "이동수업/분반용 임시 반은 가상학급여부를 Y로 표시합니다."),
    ],
    "고정 일과": [
        ("학년 HR", ["학년", "1", "5", "", "", "HR", "", "", "HR", "", "1학년 전체 수요일 5교시 HR"], "사진 예시처럼 특정 학년의 특정 교시에 HR을 고정합니다."),
        ("연속 HR", ["학년", "1", "6", "", "", "HR", "", "", "HR", "", "수요일 5~6교시 HR처럼 두 행으로 입력합니다."], "같은 표시명을 여러 교시에 넣으면 연속 고정 시간이 됩니다."),
        ("자습", ["계열", "공통", "7", "", "", "", "자습", "", "자습", "", "공통계열 목요일 7교시 자습"], "교사는 배정하지 않아도 시간표 칸에는 자습으로 표시됩니다."),
        ("창체", ["학급", "1-1", "1", "창체", "", "", "", "", "창체", "김하늘", "1-1 월요일 1교시 창체, 임장교사 표시"], "특정 학급만 고정하거나 임장교사를 메모할 수 있습니다."),
        ("전체 행사", ["전체", "", "4", "", "행사", "", "", "", "행사", "", "전체 학급 화요일 4교시 행사"], "대상유형이 전체이면 대상명은 비워둡니다."),
    ],
    "과목": [
        ("일반 과목", ["국어", "국", "국어", ""], "단축명은 시간표 칸에 짧게 표시됩니다."),
        ("유사과목", ["사회", "사", "사회", "인문"], "국사/사회처럼 같은 날 몰림을 줄이고 싶으면 같은 그룹을 씁니다."),
        ("실습 과목", ["프로그래밍", "프", "프로그래밍", "실습"], "특별실 또는 연속수업과 함께 사용할 수 있습니다."),
    ],
    "교사별 시수표": [
        ("일반 배정", ["김하늘", "국어", "", "", "", "", "4", "4", "4", "", "", "", "", ""], "김하늘 교사가 국어를 1-1, 1-2, 1-3에 각각 4시간 수업합니다."),
        ("연속수업", ["정다은", "프로그래밍", "2,2", "컴퓨터실1", "", "", "", "", "", "", "", "", "4", ""], "2기계1에 4시간을 넣고, 2시간 연속 2회가 필요하면 연속패턴을 2,2로 씁니다."),
        ("특별실 사용", ["박민수", "체육", "", "체육관", "", "", "", "2", "", "", "", "", "", ""], "특별실명은 특별실 시트에서 선택하고, 해당 학급 칸에 시수만 입력합니다."),
        ("동시수업", ["최도윤", "수학", "", "", "수학수준A", "", "", "", "", "3", "3", "3", "", ""], "동시그룹이 같은 행들은 같은 시간대 후보로 묶입니다."),
        ("복수교사", ["이서연", "영어", "", "", "", "영어협력1", "", "", "1", "", "", "", "", ""], "복수교사그룹은 복수교사 시트와 맞춥니다."),
    ],
    "특별실": [
        ("컴퓨터실", ["컴퓨터실1", ""], "실습 과목의 특별실명으로 사용합니다."),
        ("체육관", ["체육관", "월 1,금 7"], "행사나 회의로 못 쓰는 시간은 배정금지에 메모합니다."),
        ("운동장", ["운동장", "우천시 확인"], "엄격한 금지는 배정금지-희망조건 시트에도 입력합니다."),
    ],
    "동시-합반-분반": [
        ("수준별 동시", ["수학수준A", "동시", "2-1,2-2,2수학A", "수학", "2개반+가상학급 수준별 이동수업"], "동시에 움직일 학급명을 쉼표로 나열합니다."),
        ("합반수업", ["창체합반1", "합반", "1-1,1-2", "창체", "창체 합반"], "여러 학급이 한 교사/공간으로 합쳐지는 경우입니다."),
        ("분반수업", ["실험분반1", "분반", "3-1,3실험A,3실험B", "프로그래밍", "실험 분반"], "실제 학급을 여러 가상학급으로 나누는 경우입니다."),
    ],
    "연속수업": [
        ("미술 2연속", ["미술", "1-1", "2", "1"], "주 2시간을 한 번에 붙여 배정합니다."),
        ("실습 3연속", ["프로그래밍", "2기계1", "3", "1"], "실습교과처럼 긴 블록이 필요한 경우입니다."),
        ("2+2 블록", ["프로그래밍", "2-2", "2", "2"], "2시간 연속수업을 주 2회 배정합니다."),
    ],
    "복수교사": [
        ("원어민 협력", ["영어협력1", "영어", "1-3", "이서연,원어민"], "두 교사가 같은 시간에 같은 학급에 들어갑니다."),
        ("보건 협력", ["보건협력1", "보건", "2-1", "보건교사,담임"], "교사명은 교사 시트에 모두 존재해야 합니다."),
    ],
    "배정금지-희망조건": [
        ("교사 금지", ["교사", "김하늘", "배정금지", "금", "6-7", "hard", "10", "출장으로 금요일 6~7교시 불가"], "강도 hard는 반드시 지키고, 우선순위 숫자가 클수록 중요합니다."),
        ("학급 이동금지", ["학급", "1-1", "이동금지", "월", "1", "hard", "8", "학급자치 고정"], "이미 배정한 시간을 움직이지 않게 할 때 사용합니다."),
        ("과목 희망", ["과목", "미술", "희망", "화,목", "5-6", "soft", "4", "미술은 오후 선호"], "soft는 가능하면 지키는 조건입니다."),
        ("특별실 금지", ["특별실", "컴퓨터실1", "배정금지", "수", "3", "hard", "9", "컴퓨터실 점검"], "특별실 충돌 방지에 사용합니다."),
    ],
    "NEIS 코드": [
        ("과목 매칭", ["국어", "국어", "KOR101", "김하늘", "김하늘"], "NEIS에 등록된 과목명/교사명과 철자를 맞춥니다."),
        ("교사명 확인", ["수학", "수학", "MAT201", "최도윤", "최도윤"], "대상란이 비는 문제를 줄이려면 NEIS 명칭을 그대로 입력합니다."),
    ],
}


def ensure_dirs() -> None:
    IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    SOLVE_SESSION_DIR.mkdir(parents=True, exist_ok=True)


SENSITIVE_LOG_KEYS = {"apiKey", "api_key", "authorization", "Authorization", "x-goog-api-key", "token", "secret"}


def safe_log(text: str) -> None:
    try:
        if sys.stdout:
            sys.stdout.write(text)
            if not text.endswith("\n"):
                sys.stdout.write("\n")
            sys.stdout.flush()
    except Exception:
        pass


def sanitized_for_log(value):
    if isinstance(value, dict):
        return {
            key: ("[redacted]" if key in SENSITIVE_LOG_KEYS else sanitized_for_log(item))
            for key, item in value.items()
        }
    if isinstance(value, list):
        return [sanitized_for_log(item) for item in value]
    if isinstance(value, str):
        if re.search(r"(sk-|AIza|key-|token)", value, flags=re.IGNORECASE):
            return "[redacted]"
        return value[:500]
    return value


def append_operation_log(event: str, payload: dict | None = None) -> None:
    try:
        ensure_dirs()
        entry = {
            "time": now_iso(),
            "event": event,
            "payload": sanitized_for_log(payload or {}),
        }
        with OPERATION_LOG_FILE.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(entry, ensure_ascii=False) + "\n")
    except Exception:
        pass


def read_operation_logs(limit: int = 80) -> list[dict]:
    if not OPERATION_LOG_FILE.exists():
        return []
    try:
        lines = OPERATION_LOG_FILE.read_text(encoding="utf-8").splitlines()
        output = []
        for line in lines[-max(1, min(limit, 500)):]:
            try:
                output.append(json.loads(line))
            except json.JSONDecodeError:
                output.append({"time": "", "event": "raw", "payload": {"message": line[:500]}})
        return output
    except OSError:
        return []


def operation_logs_text(limit: int = 500) -> bytes:
    lines = []
    for item in read_operation_logs(limit):
        lines.append(json.dumps(item, ensure_ascii=False))
    return ("\n".join(lines) + ("\n" if lines else "")).encode("utf-8")


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def as_text(value) -> str:
    value = clean(value)
    if value is None:
        return ""
    return str(value).strip()


def is_blank_row(values) -> bool:
    return all(as_text(v) == "" for v in values)


def parse_int(value, default=None):
    if value is None or as_text(value) == "":
        return default
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default


def parse_positive_int(value):
    number = parse_int(value)
    if number is None or number <= 0:
        return None
    return number


def parse_nonnegative_int(value):
    number = parse_int(value)
    if number is None or number < 0:
        return None
    return number


def parse_csv_tokens(value) -> list[str]:
    text = as_text(value)
    if not text:
        return []
    text = text.replace("，", ",").replace("、", ",")
    return [part.strip() for part in text.split(",") if part.strip()]


def parse_days(value) -> list[str]:
    text = as_text(value)
    if not text:
        return DEFAULT_DAYS
    range_match = re.fullmatch(r"\s*([월화수목금토일])\s*(?:~|-|부터)\s*([월화수목금토일])\s*(?:까지)?\s*", text)
    if range_match:
        start, end = range_match.groups()
        start_index = WEEKDAY_ORDER.index(start)
        end_index = WEEKDAY_ORDER.index(end)
        if start_index <= end_index:
            return WEEKDAY_ORDER[start_index : end_index + 1]
        return WEEKDAY_ORDER[start_index:] + WEEKDAY_ORDER[: end_index + 1]

    normalized = text.replace("，", ",").replace("、", ",")
    if re.fullmatch(r"[월화수목금토일\s,;/]+", normalized):
        days = []
        for char in normalized:
            if char in WEEKDAY_ORDER and char not in days:
                days.append(char)
        return days or DEFAULT_DAYS

    days = parse_csv_tokens(value)
    return days or DEFAULT_DAYS


def split_day_hour_tokens(value) -> list[str]:
    text = as_text(value)
    if not text:
        return []
    text = text.replace("，", ",").replace("、", ",").replace(";", ",")
    tokens = [part.strip() for part in text.split(",") if part.strip()]
    if len(tokens) == 1 and re.fullmatch(r"\d+(?:\s+\d+)+", tokens[0]):
        return [part.strip() for part in tokens[0].split() if part.strip()]
    return tokens


def parse_day_hour_limits(value, days: list[str], max_period: int):
    tokens = split_day_hour_tokens(value)
    default_limits = {day: max_period for day in days}
    if not tokens:
        return default_limits, None

    named_limits = {}
    positional_limits = []
    for token in tokens:
        named_match = re.fullmatch(r"([월화수목금토일])\s*(?:요일)?\s*[:=]?\s*(\d+)\s*(?:교시|시간)?", token)
        if named_match:
            day, raw_number = named_match.groups()
            if day not in days:
                return default_limits, f"{day}요일은 기본설정의 수업요일에 없습니다."
            number = parse_nonnegative_int(raw_number)
            if number is None or number > max_period:
                return default_limits, f"{token} 값은 0부터 일일최대교시({max_period}) 사이여야 합니다."
            named_limits[day] = number
            continue

        number = parse_nonnegative_int(token.replace("교시", "").replace("시간", ""))
        if number is None or number > max_period:
            return default_limits, f"'{token}'은 0부터 일일최대교시({max_period}) 사이의 숫자여야 합니다."
        positional_limits.append(number)

    if named_limits and positional_limits:
        return default_limits, "요일별시수는 위치순 숫자 방식과 요일명 방식 중 하나만 사용하세요."
    if named_limits:
        limits = default_limits.copy()
        limits.update(named_limits)
        return limits, None
    if len(positional_limits) == 1:
        return {day: positional_limits[0] for day in days}, None
    if len(positional_limits) == len(days):
        return {day: positional_limits[index] for index, day in enumerate(days)}, None
    return default_limits, f"숫자를 {len(days)}개 입력하거나 한 숫자만 입력하세요. 예: 7,7,6,7,7"


def parse_period_tokens(value, max_period: int) -> list[int]:
    text = as_text(value)
    if not text:
        return list(range(1, max_period + 1))
    periods: set[int] = set()
    for token in re.split(r"[,/ ]+", text):
        token = token.strip()
        if not token:
            continue
        if "-" in token:
            start, end = token.split("-", 1)
            a = parse_int(start)
            b = parse_int(end)
            if a is not None and b is not None:
                for period in range(min(a, b), max(a, b) + 1):
                    periods.add(period)
        else:
            number = parse_int(token)
            if number is not None:
                periods.add(number)
    return sorted(p for p in periods if 1 <= p <= max_period)


def parse_block_pattern(pattern, total_hours: int) -> list[int]:
    blocks = []
    used = 0
    for token in parse_csv_tokens(pattern):
        size = parse_positive_int(token)
        if size:
            blocks.append(size)
            used += size
    while used < total_hours:
        blocks.append(1)
        used += 1
    if used > total_hours:
        trimmed = []
        remaining = total_hours
        for size in blocks:
            if remaining <= 0:
                break
            trimmed.append(min(size, remaining))
            remaining -= size
        return trimmed
    return blocks


def column_index(headers: list[str], header: str) -> int:
    try:
        return headers.index(header) + 1
    except ValueError:
        return 1


def issue(severity: str, sheet: str, row: int, column: str, message: str, fix: str) -> dict:
    return {
        "severity": severity,
        "sheet": sheet,
        "cell": f"{column}{row}" if column else "",
        "message": message,
        "fix": fix,
    }


def add_issue(issues, severity, sheet, row, column, message, fix):
    issues.append(issue(severity, sheet, row, column, message, fix))


ENTITY_PREFIXES = {
    "teachers": "T",
    "classes": "C",
    "subjects": "S",
    "rooms": "R",
}

ENTITY_CODE_FIELDS = {
    "teachers": "교사코드",
    "classes": "학급코드",
    "subjects": "과목코드",
    "rooms": "특별실코드",
}

ENTITY_COLLECTION_BY_TYPE = {
    "교사": "teachers",
    "학급": "classes",
    "과목": "subjects",
    "특별실": "rooms",
}

ENTITY_NAME_FIELDS = {
    "teachers": "교사명",
    "classes": "학급명",
    "subjects": "과목명",
    "rooms": "특별실명",
}


def generated_code(target: str, index: int) -> str:
    return f"{ENTITY_PREFIXES[target]}{index:03d}"


def add_lookup(records: dict, target: str, name: str, code: str) -> None:
    records.setdefault("_lookups", {}).setdefault(target, {})[name] = code
    records.setdefault("codeMapping", {}).setdefault(target, {})[code] = name


def resolve_name(records, issues, sheet, row, column, target, label, value, required=True):
    name = as_text(value)
    if not name:
        if required:
            add_issue(issues, "error", sheet, row, column, f"{label}이(가) 비어 있습니다.", f"{label} 시트에 입력한 이름을 선택하거나 입력하세요.")
        return ""
    code = records.get("_lookups", {}).get(target, {}).get(name)
    if not code:
        add_issue(issues, "error", sheet, row, column, f"{label} '{name}'을(를) 찾을 수 없습니다.", f"{label} 시트에 먼저 '{name}'을(를) 입력한 뒤 드롭다운에서 선택하세요.")
        return ""
    return code


def resolve_name_list(records, issues, sheet, row, column, target, label, value, required=True):
    names = parse_csv_tokens(value)
    if not names and required:
        add_issue(issues, "error", sheet, row, column, f"{label} 목록이 비어 있습니다.", f"쉼표로 구분해 {label} 이름을 입력하세요.")
        return []
    codes = []
    for name in names:
        code = resolve_name(records, issues, sheet, row, column, target, label, name, required=True)
        if code:
            codes.append(code)
    return codes


def display_name(records: dict, target: str, code: str) -> str:
    collection = ENTITY_COLLECTION_BY_TYPE.get(target, target)
    item = records.get(collection, {}).get(code, {})
    return item.get(ENTITY_NAME_FIELDS.get(collection, "displayName")) or item.get("displayName") or code


def display_names(records: dict, target: str, codes) -> list[str]:
    return [display_name(records, target, code) for code in codes]


def find_entity_in_text(records: dict, text: str) -> tuple[str, str, str] | None:
    candidates = []
    for target_type, collection in ENTITY_COLLECTION_BY_TYPE.items():
        for code, item in records.get(collection, {}).items():
            name = item.get(ENTITY_NAME_FIELDS[collection]) or item.get("displayName") or code
            for token in {code, name}:
                token = as_text(token)
                if token and token in text:
                    candidates.append((len(token), target_type, code, name))
    if not candidates:
        return None
    _, target_type, code, name = max(candidates, key=lambda item: item[0])
    return target_type, code, name


def find_entities_in_text(records: dict, text: str) -> list[tuple[str, str, str]]:
    candidates = []
    seen = set()
    for target_type, collection in ENTITY_COLLECTION_BY_TYPE.items():
        for code, item in records.get(collection, {}).items():
            name = item.get(ENTITY_NAME_FIELDS[collection]) or item.get("displayName") or code
            for token in {code, name}:
                token = as_text(token)
                if token and token in text:
                    key = (target_type, code)
                    if key not in seen:
                        seen.add(key)
                        candidates.append((len(token), target_type, code, name))
    return [(target_type, code, name) for _, target_type, code, name in sorted(candidates, key=lambda item: item[0], reverse=True)]


def make_workbook_bytes(workbook: Workbook) -> bytes:
    stream = io.BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def quoted_sheet(sheet_name: str) -> str:
    return "'" + sheet_name.replace("'", "''") + "'"


def add_dropdown(wb: Workbook, sheet_name: str, header: str, *, source_sheet: str | None = None, source_header: str | None = None, options: list[str] | None = None, prompt: str = "") -> None:
    ws = wb[sheet_name]
    headers = SPECS_BY_NAME[sheet_name]
    target_col = get_column_letter(column_index(headers, header))
    if source_sheet:
        source_headers = SPECS_BY_NAME[source_sheet]
        source_col = get_column_letter(column_index(source_headers, source_header or source_headers[0]))
        formula = f"{quoted_sheet(source_sheet)}!${source_col}$2:${source_col}${TEMPLATE_INPUT_ROWS}"
    else:
        formula = '"' + ",".join(options or []) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.errorTitle = "목록에서 선택"
    validation.error = "이미 입력된 이름을 선택하거나, 원본 시트에 먼저 값을 추가하세요."
    validation.promptTitle = "입력 도움"
    validation.prompt = prompt or "셀 오른쪽의 목록 버튼으로 선택할 수 있습니다."
    ws.add_data_validation(validation)
    validation.add(f"{target_col}2:{target_col}{TEMPLATE_INPUT_ROWS}")


def add_dropdown_range(wb: Workbook, sheet_name: str, cell_range: str, *, source_sheet: str | None = None, source_header: str | None = None, options: list[str] | None = None, prompt: str = "") -> None:
    ws = wb[sheet_name]
    if source_sheet:
        source_headers = SPECS_BY_NAME[source_sheet]
        source_col = get_column_letter(column_index(source_headers, source_header or source_headers[0]))
        formula = f"{quoted_sheet(source_sheet)}!${source_col}$2:${source_col}${TEMPLATE_INPUT_ROWS}"
    else:
        formula = '"' + ",".join(options or []) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.errorTitle = "목록에서 선택"
    validation.error = "이미 입력된 이름을 선택하거나, 원본 시트에 먼저 값을 추가하세요."
    validation.promptTitle = "입력 도움"
    validation.prompt = prompt or "셀 오른쪽의 목록 버튼으로 선택할 수 있습니다."
    ws.add_data_validation(validation)
    validation.add(cell_range)


def template_headers_for(sheet_name: str, headers: list[str]) -> list[str]:
    if sheet_name == "교사별 시수표":
        return headers + SAMPLE_CLASS_COLUMNS + [""] * (TEMPLATE_CLASS_COLUMN_COUNT - len(SAMPLE_CLASS_COLUMNS))
    return headers


def add_template_dropdowns(wb: Workbook) -> None:
    add_dropdown(wb, "학급-계열", "담임교사명", source_sheet="교사", source_header="교사명", prompt="교사 시트에 입력한 교사명을 선택합니다.")
    add_dropdown(wb, "학급-계열", "가상학급여부", options=["N", "Y"], prompt="실제 반은 N, 이동수업/분반용 가상반은 Y를 선택합니다.")

    add_dropdown(wb, "고정 일과", "대상유형", options=["전체", "학년", "계열", "학급"], prompt="고정 시간을 적용할 범위를 선택합니다.")
    add_dropdown_range(wb, "고정 일과", "D2:H200", options=["HR", "자습", "창체", "행사", "공강", "X"], prompt="시간표 칸에 표시할 이름을 선택하거나 직접 입력합니다.")
    add_dropdown(wb, "고정 일과", "유형", options=["HR", "자습", "창체", "행사", "공강", "기타"], prompt="고정 시간의 종류를 선택합니다.")
    add_dropdown(wb, "고정 일과", "임장교사명", source_sheet="교사", source_header="교사명", prompt="선택 사항입니다. 임장교사를 표시만 하고 시수로 계산하지 않습니다.")

    add_dropdown(wb, "교사별 시수표", "교사명", source_sheet="교사", source_header="교사명", prompt="교사 시트에 입력한 교사명을 선택합니다.")
    add_dropdown(wb, "교사별 시수표", "과목명", source_sheet="과목", source_header="과목명", prompt="과목 시트에 입력한 과목명을 선택합니다.")
    add_dropdown(wb, "교사별 시수표", "특별실명", source_sheet="특별실", source_header="특별실명", prompt="특별실을 쓰는 수업만 선택합니다.")
    class_start = len(SPECS_BY_NAME["교사별 시수표"]) + 1
    class_end = class_start + TEMPLATE_CLASS_COLUMN_COUNT - 1
    add_dropdown_range(
        wb,
        "교사별 시수표",
        f"{get_column_letter(class_start)}1:{get_column_letter(class_end)}1",
        source_sheet="학급-계열",
        source_header="학급명",
        prompt="학급-계열 시트에 입력한 학급명을 선택해 학급별 시수 열을 만듭니다.",
    )

    add_dropdown(wb, "동시-합반-분반", "유형", options=["동시", "합반", "분반"], prompt="수업 구조를 선택합니다.")
    add_dropdown(wb, "동시-합반-분반", "학급명목록", source_sheet="학급-계열", source_header="학급명", prompt="여러 학급은 쉼표로 구분합니다. 목록에서 선택한 뒤 필요한 이름을 추가로 입력할 수 있습니다.")
    add_dropdown(wb, "동시-합반-분반", "과목명", source_sheet="과목", source_header="과목명", prompt="과목 시트에 입력한 과목명을 선택합니다.")

    add_dropdown(wb, "연속수업", "과목명", source_sheet="과목", source_header="과목명")
    add_dropdown(wb, "연속수업", "학급명", source_sheet="학급-계열", source_header="학급명")

    add_dropdown(wb, "복수교사", "과목명", source_sheet="과목", source_header="과목명")
    add_dropdown(wb, "복수교사", "학급명", source_sheet="학급-계열", source_header="학급명")
    add_dropdown(wb, "복수교사", "교사명목록", source_sheet="교사", source_header="교사명", prompt="여러 교사는 쉼표로 구분합니다.")

    add_dropdown(wb, "배정금지-희망조건", "대상유형", options=["교사", "학급", "과목", "특별실"])
    add_dropdown(wb, "배정금지-희망조건", "조건유형", options=["배정금지", "이동금지", "임시금지", "희망", "비선호"])
    add_dropdown(wb, "배정금지-희망조건", "요일", options=["월", "화", "수", "목", "금"])
    add_dropdown(wb, "배정금지-희망조건", "강도", options=["hard", "soft"])
    add_dropdown(wb, "배정금지-희망조건", "우선순위", options=["1", "2", "3", "4", "5", "6", "7", "8", "9", "10"])

    add_dropdown(wb, "NEIS 코드", "과목명", source_sheet="과목", source_header="과목명")
    add_dropdown(wb, "NEIS 코드", "교사명", source_sheet="교사", source_header="교사명")


def create_template_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="173F3F")
    optional_header_fill = PatternFill("solid", fgColor="48656A")
    header_font = Font(color="FFFFFF", bold=True)
    input_fill = PatternFill("solid", fgColor="FFF7D6")
    locked_fill = PatternFill("solid", fgColor="EEF1F1")
    example_title_fill = PatternFill("solid", fgColor="1D4E89")
    example_header_fill = PatternFill("solid", fgColor="DCEBFA")
    example_row_fill = PatternFill("solid", fgColor="F6FAFE")
    note_fill = PatternFill("solid", fgColor="EAF6F2")
    class_header_fill = PatternFill("solid", fgColor="7A5C1E")
    thin_side = Side(style="thin", color="C9D1D1")
    input_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for name, headers, examples in SHEET_SPECS:
        ws = wb.create_sheet(name)
        display_headers = template_headers_for(name, headers)
        ws.append(display_headers)
        required = REQUIRED_FIELDS.get(name, set())

        for index, cell in enumerate(ws[1], start=1):
            header = display_headers[index - 1]
            is_class_header = name == "교사별 시수표" and index > len(headers)
            cell.fill = class_header_fill if is_class_header else (header_fill if header in required else optional_header_fill)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = input_border
            if is_class_header:
                cell.comment = Comment("학급-계열 시트에 입력한 학급명을 선택하거나 직접 입력합니다. 아래 행에는 주당 시수 숫자만 넣으세요.", "Codex")
            elif name == "학급-계열" and header == "요일별시수":
                cell.comment = Comment("기본설정의 수업요일 순서대로 입력합니다. 예: 월~금이면 7,7,6,7,7은 수요일만 6교시입니다. 한 숫자만 쓰면 모든 요일에 동일하게 적용됩니다.", "Codex")
            else:
                required_text = "필수 입력" if header in required else "선택 입력"
                cell.comment = Comment(f"{required_text}: {header} 값을 입력하세요. 왼쪽 노란색 영역만 업로드 검증 대상입니다.", "Codex")
        for row in examples:
            ws.append(row)

        input_rows = max(TEMPLATE_INPUT_ROWS, ws.max_row + 10)
        for row_index in range(2, input_rows + 1):
            for column_index in range(1, len(display_headers) + 1):
                cell = ws.cell(row=row_index, column=column_index)
                if name == "기본설정" and column_index != 2:
                    cell.fill = locked_fill
                else:
                    cell.fill = input_fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = input_border

        example_start = len(display_headers) + 2
        example_headers = ["상황"] + display_headers + ["작성팁"]
        ws.cell(row=1, column=example_start).value = "작성 예시 영역 - 참고용(업로드 검증에서 무시됨)"
        ws.merge_cells(
            start_row=1,
            start_column=example_start,
            end_row=1,
            end_column=example_start + len(example_headers) - 1,
        )
        example_title = ws.cell(row=1, column=example_start)
        example_title.fill = example_title_fill
        example_title.font = header_font
        example_title.alignment = Alignment(horizontal="center", vertical="center")
        example_title.border = input_border

        for offset, header in enumerate(example_headers):
            cell = ws.cell(row=2, column=example_start + offset)
            cell.value = header
            cell.fill = example_header_fill
            cell.font = Font(bold=True, color="12324A")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = input_border

        for row_offset, (case_name, values, tip) in enumerate(EXAMPLE_ROWS.get(name, []), start=3):
            padded_values = values + [""] * max(0, len(display_headers) - len(values))
            row_values = [case_name] + padded_values[: len(display_headers)] + [tip]
            for offset, value in enumerate(row_values):
                cell = ws.cell(row=row_offset, column=example_start + offset)
                cell.value = value
                cell.fill = example_row_fill
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = input_border

        ws.cell(row=1, column=1).comment = Comment(
            f"실제 입력 영역입니다. {SHEET_GUIDANCE.get(name, '')} 오른쪽 파란 영역은 예시이므로 그대로 두어도 업로드 데이터로 읽지 않습니다.",
            "Codex",
        )
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(display_headers))}1"
        for index, header in enumerate(display_headers, start=1):
            ws.column_dimensions[get_column_letter(index)].width = max(14, min(28, len(header) * 3))
        if name == "교사별 시수표":
            for index in range(len(headers) + 1, len(display_headers) + 1):
                ws.column_dimensions[get_column_letter(index)].width = 10
        ws.column_dimensions[get_column_letter(len(display_headers) + 1)].width = 3
        ws.column_dimensions[get_column_letter(example_start)].width = 18
        for index in range(example_start + 1, example_start + len(example_headers)):
            ws.column_dimensions[get_column_letter(index)].width = 18
        ws.column_dimensions[get_column_letter(example_start + len(example_headers) - 1)].width = 44
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 30
        ws.sheet_properties.tabColor = "173F3F"
        if name == "교사별 시수표":
            ws["A1"].comment = Comment("교사명/과목명을 한 번 쓰고, 오른쪽 학급명 열 아래에 시수만 입력하세요. 학급명 헤더는 학급-계열 시트에서 선택할 수 있습니다.", "Codex")
        if name == "배정금지-희망조건":
            ws["A1"].comment = Comment("대상유형 예: 교사, 학급, 과목, 특별실. 대상명은 해당 시트에 입력한 이름을 사용합니다.", "Codex")

    add_template_dropdowns(wb)

    guide = wb.create_sheet("도움말", 0)
    guide.append(["구분", "내용"])
    guide.append(["색상 구분", "노란색은 사용자가 작성하는 실제 입력 영역입니다. 회색은 설명/고정 영역입니다. 파란색 오른쪽 영역은 참고 예시이며 업로드 검증에서 무시됩니다."])
    guide.append(["입력 원칙", "필요한 값만 입력하세요. 비어 있는 행은 검증에서 무시됩니다. 기본설정 시트는 B열 값만 수정합니다."])
    guide.append(["코드 규칙", "사용자는 이름으로 입력합니다. 교사코드, 학급코드, 과목코드, 특별실코드는 업로드 후 서버가 자동 생성하고 AI에는 코드화된 데이터만 전달합니다."])
    guide.append(["요일별시수", "학급-계열 시트의 요일별시수는 기본설정 수업요일 순서대로 해석합니다. 월~금 기준 7,7,6,7,7은 수요일만 6교시이고 금요일은 7교시입니다. 한 숫자만 입력하면 모든 요일에 동일하게 적용됩니다."])
    guide.append(["고정 일과", "교시가 있지만 시수배당표에 들어가지 않는 HR, 자습, 창체, 행사 등은 고정 일과 시트에 입력합니다. 대상유형/대상명/교시를 쓰고, 월~금 해당 칸에 시간표에 보일 표시명을 적습니다. 임장교사명은 선택 사항이며 시수로 계산하지 않습니다."])
    guide.append(["교사별 시수표", "교사명과 과목명은 한 행에 한 번만 입력합니다. 오른쪽 1행의 학급명 열 아래에 주당 시수 숫자만 입력하세요. 예: 국어 행에서 1-1, 1-2, 1-3 칸에 각각 4를 입력하면 세 학급 수업으로 자동 변환됩니다."])
    guide.append(["학급명의 의미", "교사별 시수표의 학급명 열은 시간표 결과가 아니라 수업 대상 반입니다. 예를 들어 1-1, 1-2, 2기계1처럼 학교에 이미 편성된 반을 적습니다. 이동수업/분반용 임시 반은 학급-계열 시트에 가상학급으로 먼저 추가합니다."])
    guide.append(["클릭 입력", "교사별 시수표의 교사명/과목명/특별실명과 학급명 헤더, 연속수업, 복수교사, NEIS 코드 등 참조가 필요한 칸은 드롭다운으로 이미 입력된 이름을 선택할 수 있습니다."])
    guide.append(["필수/선택", "진한 초록 헤더는 필수 입력, 청록 회색 헤더는 선택 입력입니다. 선택 입력도 조건에 따라 필요할 수 있습니다."])
    guide.append(["예시 활용", "각 시트 오른쪽의 작성 예시를 참고해 왼쪽 노란색 입력 영역에 직접 작성하세요. 예시 영역을 그대로 두어도 실제 데이터로 읽지 않습니다."])
    guide.append(["AI 개인정보", "AI에는 실제 이름 대신 코드화된 데이터만 전달하도록 설계되어 있습니다."])
    for cell in guide[1]:
        cell.fill = note_fill
        cell.font = Font(bold=True)
        cell.border = input_border
    for row in guide.iter_rows(min_row=2, max_row=guide.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = input_border
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["B"].width = 90
    return wb


def sheet_rows(wb: Workbook, sheet_name: str, headers: list[str]):
    ws = wb[sheet_name]
    for row_index in range(2, ws.max_row + 1):
        values = [ws.cell(row=row_index, column=col).value for col in range(1, len(headers) + 1)]
        if is_blank_row(values):
            continue
        yield row_index, {header: clean(values[i]) for i, header in enumerate(headers)}


def validate_headers(wb: Workbook) -> list[dict]:
    issues = []
    for sheet_name, expected, _ in SHEET_SPECS:
        if sheet_name not in wb.sheetnames:
            add_issue(issues, "error", sheet_name, 1, "A", f"필수 시트 '{sheet_name}'가 없습니다.", "통합 입력 엑셀 양식을 다시 내려받아 해당 시트를 추가하세요.")
            continue
        ws = wb[sheet_name]
        actual = [as_text(ws.cell(row=1, column=i).value) for i in range(1, len(expected) + 1)]
        for index, header in enumerate(expected, start=1):
            if actual[index - 1] != header:
                col = get_column_letter(index)
                add_issue(issues, "error", sheet_name, 1, col, f"헤더가 '{header}'여야 하지만 '{actual[index - 1]}'입니다.", "1행 헤더는 양식과 정확히 일치해야 합니다.")
    return issues


def validate_workbook(wb: Workbook) -> dict:
    issues = validate_headers(wb)
    records = {
        "config": {},
        "teachers": {},
        "classes": {},
        "subjects": {},
        "rooms": {},
        "loads": [],
        "fixedPeriods": [],
        "constraints": [],
        "syncGroups": [],
        "syncBundles": [],
        "continuous": [],
        "coTeachers": [],
        "neis": [],
        "_lookups": {"teachers": {}, "classes": {}, "subjects": {}, "rooms": {}},
        "codeMapping": {"teachers": {}, "classes": {}, "subjects": {}, "rooms": {}},
    }

    if issues:
        return {"ok": False, "issues": issues, "records": records, "stats": summarize_records(records)}

    for row, item in sheet_rows(wb, "기본설정", SPECS_BY_NAME["기본설정"]):
        key = as_text(item["항목"])
        if key:
            records["config"][key] = as_text(item["값"])

    duplicate_check(
        wb,
        records,
        issues,
        sheet="교사",
        name_field="교사명",
        target="teachers",
        label="교사",
    )
    duplicate_check(
        wb,
        records,
        issues,
        sheet="학급-계열",
        name_field="학급명",
        target="classes",
        label="학급",
    )
    duplicate_check(
        wb,
        records,
        issues,
        sheet="과목",
        name_field="과목명",
        target="subjects",
        label="과목",
    )
    duplicate_check(
        wb,
        records,
        issues,
        sheet="특별실",
        name_field="특별실명",
        target="rooms",
        label="특별실",
    )

    validate_class_day_hours(records, issues)
    parse_fixed_periods(wb, records, issues)
    parse_loads(wb, records, issues)
    parse_constraints(wb, records, issues)
    parse_sync_groups(wb, records, issues)
    build_sync_bundles(records, issues)
    parse_continuous(wb, records, issues)
    parse_co_teachers(wb, records, issues)
    parse_neis(wb, records, issues)

    stats = summarize_records(records)
    warnings = []
    if stats["loadCount"] == 0:
        warnings.append(issue("warning", "교사별 시수표", 2, "A", "교사별 시수표가 비어 있습니다.", "실제 자동배정을 하려면 교사별 시수표를 입력하세요."))
    issues.extend(warnings)
    return {
        "ok": not any(item["severity"] == "error" for item in issues),
        "issues": issues,
        "records": records,
        "stats": stats,
    }


def duplicate_check(wb, records, issues, sheet, name_field, target, label):
    headers = SPECS_BY_NAME[sheet]
    seen = {}
    code_index = 0
    for row, item in sheet_rows(wb, sheet, headers):
        name = as_text(item.get(name_field))
        if not name:
            col = get_column_letter(column_index(headers, name_field))
            add_issue(issues, "error", sheet, row, col, f"{label}명이 비어 있습니다.", f"{label}명을 입력하세요. 내부 코드는 앱이 자동 생성합니다.")
            continue
        if name in seen:
            add_issue(issues, "error", sheet, row, "A", f"{label}명 '{name}'이(가) 중복되었습니다.", f"{seen[name]}행과 {row}행 중 하나의 이름을 구분되게 수정하세요.")
            continue
        seen[name] = row
        code_index += 1
        code = generated_code(target, code_index)
        record = {key: as_text(value) for key, value in item.items()}
        record["row"] = row
        record[ENTITY_CODE_FIELDS[target]] = code
        record["displayName"] = name
        records[target][code] = record
        add_lookup(records, target, name, code)


def validate_class_day_hours(records, issues):
    days, max_period = schedule_dimensions(records)
    headers = SPECS_BY_NAME["학급-계열"]
    column = get_column_letter(column_index(headers, "요일별시수"))
    for class_code, item in records.get("classes", {}).items():
        limits, message = parse_day_hour_limits(item.get("요일별시수"), days, max_period)
        item["_dayLimits"] = limits
        if message:
            add_issue(
                issues,
                "error",
                "학급-계열",
                item.get("row", 2),
                column,
                f"{item.get('학급명', class_code)}의 요일별시수 형식이 올바르지 않습니다.",
                f"{message} 월~금 순서라면 예: 7,7,6,7,7",
            )


def classes_for_fixed_target(records, issues, sheet, row, target_type, target_name):
    classes = records.get("classes", {})
    if target_type == "전체":
        if not classes:
            add_issue(issues, "error", sheet, row, "A", "전체 고정 일과를 적용할 학급이 없습니다.", "학급-계열 시트에 학급을 먼저 입력하세요.")
        return list(classes.keys())
    if target_type == "학급":
        code = resolve_name(records, issues, sheet, row, "B", "classes", "학급", target_name)
        return [code] if code else []
    if target_type == "학년":
        if not target_name:
            add_issue(issues, "error", sheet, row, "B", "학년 대상명은 비워둘 수 없습니다.", "예: 1, 2, 3처럼 학년을 입력하세요.")
            return []
        matched = [code for code, item in classes.items() if as_text(item.get("학년")) == target_name]
        if not matched:
            add_issue(issues, "error", sheet, row, "B", f"학년 '{target_name}'에 해당하는 학급이 없습니다.", "학급-계열 시트의 학년 값을 확인하세요.")
        return matched
    if target_type == "계열":
        if not target_name:
            add_issue(issues, "error", sheet, row, "B", "계열 대상명은 비워둘 수 없습니다.", "예: 공통, 기계, 수준별처럼 계열명을 입력하세요.")
            return []
        matched = [code for code, item in classes.items() if as_text(item.get("계열")) == target_name]
        if not matched:
            add_issue(issues, "error", sheet, row, "B", f"계열 '{target_name}'에 해당하는 학급이 없습니다.", "학급-계열 시트의 계열 값을 확인하세요.")
        return matched
    add_issue(issues, "error", sheet, row, "A", "대상유형은 전체, 학년, 계열, 학급 중 하나여야 합니다.", "대상유형을 목록에서 선택하세요.")
    return []


def parse_fixed_periods(wb, records, issues):
    sheet = "고정 일과"
    headers = SPECS_BY_NAME[sheet]
    days = [day for day in DEFAULT_DAYS if day in headers]
    schedule_days, max_period = schedule_dimensions(records)
    seen = {}
    for row, item in sheet_rows(wb, sheet, headers):
        target_type = as_text(item["대상유형"])
        target_name = as_text(item["대상명"])
        period = parse_positive_int(item["교시"])
        if period is None:
            add_issue(issues, "error", sheet, row, "C", "교시는 1 이상의 숫자여야 합니다.", "예: 5")
            continue
        if period > max_period:
            add_issue(issues, "error", sheet, row, "C", f"{period}교시는 기본설정의 일일최대교시({max_period})보다 큽니다.", "기본설정 일일최대교시를 늘리거나 교시 값을 수정하세요.")
            continue
        labels_by_day = {day: as_text(item[day]) for day in days}
        if not any(labels_by_day.values()):
            add_issue(issues, "warning", sheet, row, "D", "요일 칸에 표시명이 입력되지 않았습니다.", "월~금 중 해당 요일 칸에 HR, 자습, 창체처럼 시간표에 보일 이름을 입력하세요.")
        class_codes = classes_for_fixed_target(records, issues, sheet, row, target_type, target_name)
        supervisor_name = as_text(item["임장교사명"])
        supervisor_code = resolve_name(records, issues, sheet, row, "J", "teachers", "교사", supervisor_name, required=False) if supervisor_name else ""
        kind = as_text(item["유형"]) or "비수업"
        for day in days:
            label = labels_by_day[day]
            if not label:
                continue
            if day not in schedule_days:
                day_col = get_column_letter(column_index(headers, day))
                add_issue(issues, "error", sheet, row, day_col, f"{day}요일은 기본설정의 수업요일에 없습니다.", "기본설정 수업요일에 해당 요일을 추가하거나 이 칸을 비우세요.")
                continue
            for class_code in class_codes:
                day_limit = records["classes"][class_code].get("_dayLimits", {}).get(day, max_period)
                if period > day_limit:
                    day_col = get_column_letter(column_index(headers, day))
                    add_issue(issues, "error", sheet, row, day_col, f"{records['classes'][class_code].get('학급명', class_code)}은 {day}요일 {day_limit}교시까지만 운영됩니다.", "학급-계열 시트의 요일별시수를 늘리거나 고정 일과 교시를 수정하세요.")
                    continue
                key = (class_code, day, period)
                if key in seen:
                    prev_row = seen[key]
                    day_col = get_column_letter(column_index(headers, day))
                    add_issue(issues, "error", sheet, row, day_col, f"{records['classes'][class_code].get('학급명', class_code)} {day} {period}교시 고정 일과가 중복됩니다.", f"{prev_row}행과 {row}행 중 하나만 남기세요.")
                    continue
                seen[key] = row
                records["fixedPeriods"].append({
                    "row": row,
                    "targetType": target_type,
                    "targetName": target_name,
                    "classCode": class_code,
                    "className": records["classes"][class_code].get("학급명", class_code),
                    "day": day,
                    "period": period,
                    "label": label,
                    "kind": kind,
                    "supervisorCode": supervisor_code,
                    "supervisorName": supervisor_name,
                    "description": as_text(item["설명"]),
                })


def parse_loads(wb, records, issues):
    sheet = "교사별 시수표"
    headers = SPECS_BY_NAME[sheet]
    ws = wb[sheet]
    fixed_count = len(headers)
    class_start_col = fixed_count + 1
    class_end_col = fixed_count + TEMPLATE_CLASS_COLUMN_COUNT
    class_columns = []
    for column_index in range(class_start_col, class_end_col + 1):
        class_name = as_text(ws.cell(row=1, column=column_index).value)
        class_columns.append((column_index, class_name))

    for row in range(2, ws.max_row + 1):
        fixed_values = [ws.cell(row=row, column=col).value for col in range(1, fixed_count + 1)]
        hour_values = [ws.cell(row=row, column=col).value for col, _ in class_columns]
        if is_blank_row(fixed_values + hour_values):
            continue
        item = {header: clean(fixed_values[i]) for i, header in enumerate(headers)}
        teacher_name = as_text(item["교사명"])
        subject_name = as_text(item["과목명"])
        teacher = resolve_name(records, issues, sheet, row, "A", "teachers", "교사", teacher_name)
        subject = resolve_name(records, issues, sheet, row, "B", "subjects", "과목", subject_name)
        room_name = as_text(item["특별실명"])
        room = resolve_name(records, issues, sheet, row, "D", "rooms", "특별실", room_name, required=False) if room_name else ""
        created_count = 0
        for column_index, class_name in class_columns:
            raw_hours = ws.cell(row=row, column=column_index).value
            if as_text(raw_hours) in {"", "0"}:
                continue
            column_letter = get_column_letter(column_index)
            if not class_name:
                add_issue(issues, "error", sheet, row, column_letter, "학급명 헤더가 비어 있는데 시수가 입력되어 있습니다.", "1행의 해당 열에 학급-계열 시트에 등록한 학급명을 선택하세요.")
                continue
            klass = resolve_name(records, issues, sheet, row, column_letter, "classes", "학급", class_name)
            hours = parse_positive_int(raw_hours)
            if hours is None:
                add_issue(issues, "error", sheet, row, column_letter, "학급별 시수는 1 이상의 숫자여야 합니다.", "해당 학급 칸에는 주당 시수 숫자만 입력하세요.")
                continue
            if not (teacher and subject and klass):
                continue
            records["loads"].append({
                "row": row,
                "teacherCode": teacher,
                "teacherName": teacher_name,
                "subjectCode": subject,
                "subjectName": subject_name,
                "classCode": klass,
                "className": class_name,
                "weeklyHours": hours,
                "continuousPattern": as_text(item["연속패턴"]),
                "roomCode": room,
                "roomName": room_name,
                "syncGroup": as_text(item["동시그룹"]),
                "coTeacherGroup": as_text(item["복수교사그룹"]),
            })
            created_count += 1
        if teacher_name and subject_name and created_count == 0 and not any(as_text(v) not in {"", "0"} for v in hour_values):
            add_issue(issues, "warning", sheet, row, get_column_letter(class_start_col), "이 행에는 학급별 시수가 입력되지 않았습니다.", "오른쪽 학급명 열 아래에 주당 시수를 숫자로 입력하세요.")


def parse_constraints(wb, records, issues):
    sheet = "배정금지-희망조건"
    headers = SPECS_BY_NAME[sheet]
    allowed_targets = {"교사": "teachers", "학급": "classes", "과목": "subjects", "특별실": "rooms"}
    allowed_types = {"배정금지", "이동금지", "임시금지", "희망", "비선호"}
    for row, item in sheet_rows(wb, sheet, headers):
        target_type = as_text(item["대상유형"])
        target_name = as_text(item["대상명"])
        condition_type = as_text(item["조건유형"]) or "배정금지"
        priority = parse_positive_int(item.get("우선순위")) or (10 if as_text(item.get("강도")) == "hard" else 5)
        if target_type not in allowed_targets:
            add_issue(issues, "error", sheet, row, "A", "대상유형은 교사, 학급, 과목, 특별실 중 하나여야 합니다.", "대상유형을 허용된 값으로 수정하세요.")
            continue
        target_code = resolve_name(records, issues, sheet, row, "B", allowed_targets[target_type], target_type, target_name)
        if condition_type not in allowed_types:
            add_issue(issues, "warning", sheet, row, "C", f"조건유형 '{condition_type}'은 아직 표준 유형이 아닙니다.", "배정금지, 이동금지, 임시금지, 희망, 비선호 중 하나를 권장합니다.")
        records["constraints"].append({
            "row": row,
            "targetType": target_type,
            "targetCode": target_code,
            "targetName": target_name,
            "conditionType": condition_type,
            "days": parse_csv_tokens(item["요일"]),
            "periodsText": as_text(item["교시"]),
            "strength": as_text(item["강도"]) or "hard",
            "priority": priority,
            "description": as_text(item["설명"]),
        })


def parse_sync_groups(wb, records, issues):
    sheet = "동시-합반-분반"
    headers = SPECS_BY_NAME[sheet]
    for row, item in sheet_rows(wb, sheet, headers):
        code = as_text(item["그룹명"])
        group_type = as_text(item["유형"])
        subject_name = as_text(item["과목명"])
        subject = resolve_name(records, issues, sheet, row, "D", "subjects", "과목", subject_name, required=False) if subject_name else ""
        class_names = parse_csv_tokens(item["학급명목록"])
        class_codes = resolve_name_list(records, issues, sheet, row, "C", "classes", "학급", item["학급명목록"])
        if not code:
            add_issue(issues, "error", sheet, row, "A", "그룹명이 필요합니다.", "동시-합반-분반 그룹의 이름을 입력하세요.")
        records["syncGroups"].append({
            "row": row,
            "groupCode": code,
            "groupName": code,
            "type": group_type,
            "classCodes": class_codes,
            "classNames": class_names,
            "subjectCode": subject,
            "subjectName": subject_name,
            "description": as_text(item["설명"]),
        })


def sync_lane_key_for_load(load: dict) -> str:
    return f"{as_text(load.get('syncGroup'))}::{as_text(load.get('subjectCode'))}"


def build_sync_bundles(records: dict, issues: list[dict] | None = None) -> list[dict]:
    """Build simultaneous-class bundles from 교사별 시수표 동시그룹 values."""
    grouped = defaultdict(list)
    for load in records.get("loads", []):
        group = as_text(load.get("syncGroup"))
        if group:
            grouped[group].append(load)

    bundles = []
    occurrence_sizes = {}
    sheet_name = SHEET_SPECS[5][0]
    headers = SPECS_BY_NAME[sheet_name]
    sync_column = get_column_letter(column_index(headers, "동시그룹")) if "동시그룹" in headers else "E"

    for group_code, loads in sorted(grouped.items(), key=lambda item: item[0]):
        lanes = defaultdict(list)
        for load in loads:
            lanes[as_text(load.get("subjectCode"))].append(load)

        lane_hours = {
            subject_code: sum(parse_positive_int(load.get("weeklyHours")) or 0 for load in lane_loads)
            for subject_code, lane_loads in lanes.items()
        }
        nonzero_totals = {hours for hours in lane_hours.values() if hours > 0}
        if len(nonzero_totals) > 1 and issues is not None:
            detail = ", ".join(
                f"{display_name(records, '과목', subject_code)} {hours}시간"
                for subject_code, hours in sorted(lane_hours.items(), key=lambda item: display_name(records, "과목", item[0]))
            )
            for load in loads:
                add_issue(
                    issues,
                    "error",
                    sheet_name,
                    load.get("row", 2),
                    sync_column,
                    f"동시그룹 '{group_code}'의 과목별 합산 시수가 같아야 합니다.",
                    f"현재 {detail}입니다. 같은 동시그룹의 과목 lane 시수를 동일하게 맞춰주세요.",
                )
            continue

        total_hours = next(iter(nonzero_totals), 0)
        lane_units = {}
        for subject_code, lane_loads in lanes.items():
            units = []
            for load in sorted(lane_loads, key=lambda item: (item.get("className", ""), item.get("teacherName", ""), item.get("row", 0))):
                for _ in range(parse_positive_int(load.get("weeklyHours")) or 0):
                    units.append(load)
            lane_units[subject_code] = units

        occurrences = []
        for index in range(total_hours):
            occurrence_id = f"{group_code}:{index + 1}"
            units = []
            for subject_code, units_for_lane in sorted(lane_units.items(), key=lambda item: display_name(records, "과목", item[0])):
                if index >= len(units_for_lane):
                    continue
                load = units_for_lane[index]
                lane_key = f"{group_code}::{subject_code}"
                units.append({
                    "laneKey": lane_key,
                    "subjectCode": subject_code,
                    "load": load,
                })
            if units:
                occurrences.append({
                    "syncGroup": group_code,
                    "syncOccurrenceId": occurrence_id,
                    "units": units,
                })
                occurrence_sizes[occurrence_id] = len(units)

        bundles.append({
            "syncGroup": group_code,
            "laneHours": lane_hours,
            "occurrences": occurrences,
        })

    records["syncBundles"] = bundles
    records["_syncOccurrenceSize"] = occurrence_sizes
    return bundles


def parse_continuous(wb, records, issues):
    sheet = "연속수업"
    headers = SPECS_BY_NAME[sheet]
    for row, item in sheet_rows(wb, sheet, headers):
        subject_name = as_text(item["과목명"])
        class_name = as_text(item["학급명"])
        subject = resolve_name(records, issues, sheet, row, "A", "subjects", "과목", subject_name)
        klass = resolve_name(records, issues, sheet, row, "B", "classes", "학급", class_name)
        block = parse_positive_int(item["연속시수"])
        count = parse_positive_int(item["횟수"]) or 1
        if block is None:
            add_issue(issues, "error", sheet, row, "C", "연속시수는 1 이상의 숫자여야 합니다.", "예: 2")
        records["continuous"].append({"row": row, "subjectCode": subject, "subjectName": subject_name, "classCode": klass, "className": class_name, "blockSize": block or 1, "count": count})


def parse_co_teachers(wb, records, issues):
    sheet = "복수교사"
    headers = SPECS_BY_NAME[sheet]
    for row, item in sheet_rows(wb, sheet, headers):
        subject_name = as_text(item["과목명"])
        class_name = as_text(item["학급명"])
        teacher_names = parse_csv_tokens(item["교사명목록"])
        subject = resolve_name(records, issues, sheet, row, "B", "subjects", "과목", subject_name)
        klass = resolve_name(records, issues, sheet, row, "C", "classes", "학급", class_name)
        teachers = resolve_name_list(records, issues, sheet, row, "D", "teachers", "교사", item["교사명목록"])
        records["coTeachers"].append({
            "row": row,
            "groupCode": as_text(item["그룹명"]),
            "groupName": as_text(item["그룹명"]),
            "subjectCode": subject,
            "subjectName": subject_name,
            "classCode": klass,
            "className": class_name,
            "teacherCodes": teachers,
            "teacherNames": teacher_names,
        })


def parse_neis(wb, records, issues):
    sheet = "NEIS 코드"
    headers = SPECS_BY_NAME[sheet]
    for row, item in sheet_rows(wb, sheet, headers):
        subject_name = as_text(item["과목명"])
        teacher_name = as_text(item["교사명"])
        subject = resolve_name(records, issues, sheet, row, "A", "subjects", "과목", subject_name, required=False) if subject_name else ""
        teacher = resolve_name(records, issues, sheet, row, "D", "teachers", "교사", teacher_name, required=False) if teacher_name else ""
        if subject and not as_text(item["NEIS과목명"]):
            add_issue(issues, "warning", sheet, row, "B", "NEIS과목명이 비어 있습니다.", "NEIS 일괄파일을 만들려면 NEIS 과목명을 입력하세요.")
        record = {key: as_text(value) for key, value in item.items()}
        record["과목코드"] = subject
        record["교사코드"] = teacher
        records["neis"].append(record)


def summarize_records(records) -> dict:
    return {
        "teacherCount": len(records.get("teachers", {})),
        "classCount": len(records.get("classes", {})),
        "subjectCount": len(records.get("subjects", {})),
        "roomCount": len(records.get("rooms", {})),
        "loadCount": len(records.get("loads", [])),
        "fixedPeriodCount": len(records.get("fixedPeriods", [])),
        "constraintCount": len(records.get("constraints", [])),
        "syncGroupCount": len(records.get("syncGroups", [])),
        "syncBundleCount": len(records.get("syncBundles", [])),
        "continuousCount": len(records.get("continuous", [])),
        "coTeacherCount": len(records.get("coTeachers", [])),
    }


def create_report_workbook(validation: dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "검증결과"
    ws.append(["심각도", "시트", "셀", "문제", "수정 가이드"])
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="173F3F")
        cell.font = Font(color="FFFFFF", bold=True)
    if not validation["issues"]:
        ws.append(["success", "", "", "검증 오류가 없습니다.", "자동배정으로 진행할 수 있습니다."])
    else:
        for item in validation["issues"]:
            ws.append([item["severity"], item["sheet"], item["cell"], item["message"], item["fix"]])
            if item["severity"] == "error":
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill("solid", fgColor="FCE7E2")
            elif item["severity"] == "warning":
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill("solid", fgColor="FFF3D6")
    widths = [12, 18, 10, 60, 70]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    summary = wb.create_sheet("요약")
    summary.append(["항목", "값"])
    for key, value in validation["stats"].items():
        summary.append([key, value])
    return wb


POSTGRES_ENV_NAMES = ("POSTGRES_URL", "DATABASE_URL", "POSTGRES_URL_NON_POOLING", "POSTGRES_PRISMA_URL")
REDIS_URL_ENV_NAMES = ("KV_REST_API_URL", "UPSTASH_REDIS_REST_URL")
REDIS_TOKEN_ENV_NAMES = ("KV_REST_API_TOKEN", "UPSTASH_REDIS_REST_TOKEN")
_POSTGRES_SCHEMA_READY = False


def first_env(names: tuple[str, ...]) -> str:
    for name in names:
        value = as_text(os.environ.get(name))
        if value:
            return value
    return ""


def postgres_url() -> str:
    return first_env(POSTGRES_ENV_NAMES)


def redis_config() -> tuple[str, str]:
    return first_env(REDIS_URL_ENV_NAMES), first_env(REDIS_TOKEN_ENV_NAMES)


def blob_enabled() -> bool:
    return bool(as_text(os.environ.get("BLOB_READ_WRITE_TOKEN")))


def storage_mode() -> str:
    parts = []
    if postgres_url():
        parts.append("postgres")
    url, token = redis_config()
    if url and token:
        parts.append("redis")
    if blob_enabled():
        parts.append("blob")
    return "+".join(parts) if parts else "local"


def public_import_metadata(metadata: dict) -> dict:
    return {key: metadata[key] for key in ["id", "createdAt", "fileName", "ok", "stats", "issues"] if key in metadata}


def blob_put(pathname: str, payload: bytes, content_type: str) -> dict:
    from vercel.blob import BlobClient

    client = BlobClient()
    uploaded = client.put(
        pathname,
        payload,
        access=os.environ.get("BLOB_ACCESS", "private"),
        content_type=content_type,
        overwrite=True,
    )
    if isinstance(uploaded, dict):
        return uploaded
    return {key: getattr(uploaded, key) for key in ("url", "download_url", "pathname") if hasattr(uploaded, key)}


def blob_get(pathname: str) -> bytes:
    from vercel.blob import BlobClient

    client = BlobClient()
    result = client.get(pathname, access=os.environ.get("BLOB_ACCESS", "private"))
    status_code = getattr(result, "status_code", getattr(result, "statusCode", 200)) if result else None
    if result is None or status_code != 200:
        raise FileNotFoundError(pathname)
    stream = getattr(result, "stream", None)
    if stream is None:
        return b""
    if isinstance(stream, bytes):
        return stream
    if hasattr(stream, "read"):
        return stream.read()
    return b"".join(bytes(chunk) for chunk in stream)


def postgres_connect():
    try:
        import psycopg
    except ImportError as exc:
        raise RuntimeError("Postgres 저장소를 사용하려면 psycopg 패키지가 필요합니다.") from exc
    return psycopg.connect(postgres_url(), autocommit=True)


def ensure_postgres_schema() -> None:
    global _POSTGRES_SCHEMA_READY
    if _POSTGRES_SCHEMA_READY:
        return
    with postgres_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS timetable_imports (
                    id TEXT PRIMARY KEY,
                    created_at TEXT NOT NULL,
                    file_name TEXT NOT NULL,
                    ok BOOLEAN NOT NULL,
                    issues_json TEXT NOT NULL,
                    stats_json TEXT NOT NULL,
                    records_json TEXT NOT NULL,
                    input_blob_path TEXT,
                    report_blob_path TEXT,
                    input_bytes BYTEA,
                    report_bytes BYTEA
                )
                """
            )
            cur.execute(
                """
                CREATE TABLE IF NOT EXISTS timetable_state (
                    key TEXT PRIMARY KEY,
                    payload_json TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
                """
            )
    _POSTGRES_SCHEMA_READY = True


def save_import_postgres(metadata: dict, original_bytes: bytes, report_bytes: bytes) -> dict:
    ensure_postgres_schema()
    input_blob_path = ""
    report_blob_path = ""
    input_bytes = original_bytes
    stored_report_bytes = report_bytes
    if blob_enabled():
        input_blob_path = f"imports/{metadata['id']}/input.xlsx"
        report_blob_path = f"imports/{metadata['id']}/report.xlsx"
        blob_put(input_blob_path, original_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        blob_put(report_blob_path, report_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        input_bytes = None
        stored_report_bytes = None
    with postgres_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO timetable_imports (
                    id, created_at, file_name, ok, issues_json, stats_json, records_json,
                    input_blob_path, report_blob_path, input_bytes, report_bytes
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (id) DO UPDATE SET
                    created_at = EXCLUDED.created_at,
                    file_name = EXCLUDED.file_name,
                    ok = EXCLUDED.ok,
                    issues_json = EXCLUDED.issues_json,
                    stats_json = EXCLUDED.stats_json,
                    records_json = EXCLUDED.records_json,
                    input_blob_path = EXCLUDED.input_blob_path,
                    report_blob_path = EXCLUDED.report_blob_path,
                    input_bytes = EXCLUDED.input_bytes,
                    report_bytes = EXCLUDED.report_bytes
                """,
                (
                    metadata["id"],
                    metadata["createdAt"],
                    metadata["fileName"],
                    metadata["ok"],
                    json.dumps(metadata["issues"], ensure_ascii=False),
                    json.dumps(metadata["stats"], ensure_ascii=False),
                    json.dumps(metadata["records"], ensure_ascii=False),
                    input_blob_path,
                    report_blob_path,
                    input_bytes,
                    stored_report_bytes,
                ),
            )
    metadata["_inputBlobPath"] = input_blob_path
    metadata["_reportBlobPath"] = report_blob_path
    return metadata


def list_imports_postgres() -> list[dict]:
    ensure_postgres_schema()
    with postgres_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, created_at, file_name, ok, issues_json, stats_json
                FROM timetable_imports
                ORDER BY created_at DESC
                LIMIT 100
                """
            )
            rows = cur.fetchall()
    return [
        {
            "id": row[0],
            "createdAt": row[1],
            "fileName": row[2],
            "ok": row[3],
            "issues": json.loads(row[4]),
            "stats": json.loads(row[5]),
        }
        for row in rows
    ]


def load_import_postgres(import_id: str) -> dict | None:
    ensure_postgres_schema()
    with postgres_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id, created_at, file_name, ok, issues_json, stats_json, records_json, input_blob_path, report_blob_path
                FROM timetable_imports
                WHERE id = %s
                """,
                (import_id,),
            )
            row = cur.fetchone()
    if not row:
        return None
    return {
        "id": row[0],
        "createdAt": row[1],
        "fileName": row[2],
        "ok": row[3],
        "issues": json.loads(row[4]),
        "stats": json.loads(row[5]),
        "records": json.loads(row[6]),
        "_inputBlobPath": row[7] or "",
        "_reportBlobPath": row[8] or "",
    }


def load_report_postgres(import_id: str) -> bytes | None:
    ensure_postgres_schema()
    with postgres_connect() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT report_blob_path, report_bytes FROM timetable_imports WHERE id = %s", (import_id,))
            row = cur.fetchone()
    if not row:
        return None
    report_blob_path, report_bytes = row
    if report_blob_path:
        return blob_get(report_blob_path)
    return bytes(report_bytes) if report_bytes is not None else None


def save_state_postgres(key: str, payload: dict) -> None:
    ensure_postgres_schema()
    with postgres_connect() as conn:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO timetable_state (key, payload_json, updated_at)
                VALUES (%s, %s, %s)
                ON CONFLICT (key) DO UPDATE SET
                    payload_json = EXCLUDED.payload_json,
                    updated_at = EXCLUDED.updated_at
                """,
                (key, json.dumps(payload, ensure_ascii=False), now_iso()),
            )


def load_state_postgres(key: str) -> dict | None:
    ensure_postgres_schema()
    with postgres_connect() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT payload_json FROM timetable_state WHERE key = %s", (key,))
            row = cur.fetchone()
    return json.loads(row[0]) if row else None


def redis_command(command: list) -> object:
    url, token = redis_config()
    if not url or not token:
        raise RuntimeError("Redis REST URL/TOKEN 환경변수가 필요합니다.")
    request = Request(
        url.rstrip("/"),
        data=json.dumps(command, ensure_ascii=False).encode("utf-8"),
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        method="POST",
    )
    with urlopen(request, timeout=8) as response:
        payload = json.loads(response.read().decode("utf-8"))
    if payload.get("error"):
        raise RuntimeError(payload["error"])
    return payload.get("result")


def save_import_redis(metadata: dict, original_bytes: bytes, report_bytes: bytes) -> dict:
    if blob_enabled():
        input_blob_path = f"imports/{metadata['id']}/input.xlsx"
        report_blob_path = f"imports/{metadata['id']}/report.xlsx"
        blob_put(input_blob_path, original_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        blob_put(report_blob_path, report_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        metadata["_inputBlobPath"] = input_blob_path
        metadata["_reportBlobPath"] = report_blob_path
    else:
        metadata["_reportBase64"] = base64.b64encode(report_bytes).decode("ascii")
    redis_command(["SET", f"timetable:import:{metadata['id']}", json.dumps(metadata, ensure_ascii=False)])
    redis_command(["LPUSH", "timetable:imports", metadata["id"]])
    return metadata


def list_imports_redis() -> list[dict]:
    ids = redis_command(["LRANGE", "timetable:imports", "0", "99"]) or []
    imports = []
    seen = set()
    for import_id in ids:
        if import_id in seen:
            continue
        seen.add(import_id)
        raw = redis_command(["GET", f"timetable:import:{import_id}"])
        if raw:
            imports.append(public_import_metadata(json.loads(raw)))
    return imports


def load_import_redis(import_id: str) -> dict | None:
    raw = redis_command(["GET", f"timetable:import:{import_id}"])
    return json.loads(raw) if raw else None


def load_report_redis(import_id: str) -> bytes | None:
    metadata = load_import_redis(import_id)
    if not metadata:
        return None
    if metadata.get("_reportBlobPath"):
        return blob_get(metadata["_reportBlobPath"])
    if metadata.get("_reportBase64"):
        return base64.b64decode(metadata["_reportBase64"])
    return None


def save_state_redis(key: str, payload: dict) -> None:
    redis_command(["SET", f"timetable:state:{key}", json.dumps(payload, ensure_ascii=False)])


def load_state_redis(key: str) -> dict | None:
    raw = redis_command(["GET", f"timetable:state:{key}"])
    return json.loads(raw) if raw else None


def save_import(validation: dict, original_name: str, original_bytes: bytes) -> dict:
    import_id = uuid.uuid4().hex[:12]
    report_bytes = make_workbook_bytes(create_report_workbook(validation))
    metadata = {
        "id": import_id,
        "createdAt": now_iso(),
        "fileName": original_name,
        "ok": validation["ok"],
        "issues": validation["issues"],
        "stats": validation["stats"],
        "records": validation["records"],
    }
    if postgres_url():
        return save_import_postgres(metadata, original_bytes, report_bytes)
    redis_url, redis_token = redis_config()
    if redis_url and redis_token:
        return save_import_redis(metadata, original_bytes, report_bytes)
    ensure_dirs()
    folder = IMPORT_DIR / import_id
    folder.mkdir(parents=True, exist_ok=True)
    (folder / "input.xlsx").write_bytes(original_bytes)
    (folder / "report.xlsx").write_bytes(report_bytes)
    (folder / "metadata.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    return metadata


def list_imports() -> list[dict]:
    if postgres_url():
        return list_imports_postgres()
    redis_url, redis_token = redis_config()
    if redis_url and redis_token:
        return list_imports_redis()
    ensure_dirs()
    imports = []
    for metadata_path in IMPORT_DIR.glob("*/metadata.json"):
        try:
            metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
            imports.append(public_import_metadata(metadata))
        except (OSError, json.JSONDecodeError):
            continue
    return sorted(imports, key=lambda item: item.get("createdAt", ""), reverse=True)


def load_import(import_id: str) -> dict | None:
    if postgres_url():
        return load_import_postgres(import_id)
    redis_url, redis_token = redis_config()
    if redis_url and redis_token:
        return load_import_redis(import_id)
    path = IMPORT_DIR / import_id / "metadata.json"
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def load_report_bytes(import_id: str) -> bytes | None:
    if postgres_url():
        return load_report_postgres(import_id)
    redis_url, redis_token = redis_config()
    if redis_url and redis_token:
        return load_report_redis(import_id)
    report_path = IMPORT_DIR / import_id / "report.xlsx"
    return report_path.read_bytes() if report_path.exists() else None


def read_json_body(handler: BaseHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0"))
    raw = handler.rfile.read(length) if length else b"{}"
    if not raw:
        return {}
    return json.loads(raw.decode("utf-8"))


def parse_multipart_file(handler: BaseHTTPRequestHandler):
    content_type = handler.headers.get("Content-Type", "")
    match = re.search(r"boundary=([^;]+)", content_type)
    if not match:
        raise ValueError("multipart/form-data boundary를 찾을 수 없습니다.")
    boundary = match.group(1).strip().strip('"').encode()
    length = int(handler.headers.get("Content-Length", "0"))
    body = handler.rfile.read(length)
    for part in body.split(b"--" + boundary):
        if b"filename=" not in part:
            continue
        header, _, payload = part.partition(b"\r\n\r\n")
        if not payload:
            continue
        header_text = header.decode("utf-8", errors="replace")
        filename_match = re.search(r'filename="([^"]*)"', header_text)
        filename = filename_match.group(1) if filename_match else "upload.xlsx"
        payload = payload.rstrip(b"\r\n")
        return filename, payload
    raise ValueError("업로드된 파일을 찾을 수 없습니다.")


def get_records_from_body(body: dict) -> dict | None:
    import_id = body.get("importId")
    if import_id:
        metadata = load_import(str(import_id))
        if metadata:
            records = metadata["records"]
        else:
            records = latest_import_records() if body.get("fallbackLatestImport") else None
    else:
        records = body.get("records")
        if records is None and body.get("fallbackLatestImport"):
            records = latest_import_records()
    if records is not None and body.get("effectiveConfig"):
        records = records_with_config(records, body.get("effectiveConfig"))
    if records is not None and body.get("chatConstraints"):
        records = records_with_chat_constraints(records, body.get("chatConstraints"))
    if records is not None and "syncBundles" not in records:
        build_sync_bundles(records, None)
    return records


def get_schedule_from_body(body: dict) -> dict | None:
    schedule = body.get("schedule")
    if schedule:
        return schedule
    if body.get("fallbackLastSchedule"):
        result = load_last_schedule() or {}
        selected = result.get("selected", {})
        return selected.get("schedule")
    return None


def latest_import_records() -> dict | None:
    for item in list_imports():
        metadata = load_import(item.get("id", ""))
        if metadata and metadata.get("records"):
            return metadata["records"]
    return None


def records_with_config(records: dict, updates: dict | None) -> dict:
    if not updates:
        return records
    copied = deepcopy(records)
    copied.setdefault("config", {}).update({key: as_text(value) for key, value in updates.items()})
    return copied


def normalized_chat_constraint(records: dict, constraint: dict, index: int) -> dict | None:
    if constraint.get("engineSupported") is False:
        return None
    target_type = as_text(constraint.get("targetType"))
    target_code = as_text(constraint.get("targetCode"))
    target_name = as_text(constraint.get("targetName"))
    collection = ENTITY_COLLECTION_BY_TYPE.get(target_type)
    if not collection:
        return None
    if not target_code and target_name:
        target_code = records.get("_lookups", {}).get(collection, {}).get(target_name, "")
    if not target_code or target_code not in records.get(collection, {}):
        return None
    condition_type = as_text(constraint.get("conditionType")) or "배정금지"
    if condition_type not in {"배정금지", "이동금지", "임시금지", "희망", "비선호"}:
        condition_type = "배정금지"
    strength = as_text(constraint.get("strength")) or ("soft" if condition_type in {"희망", "비선호"} else "hard")
    return {
        "row": f"chat-{index}",
        "source": "ai-chat",
        "targetType": target_type,
        "targetCode": target_code,
        "targetName": target_name or display_name(records, target_type, target_code),
        "conditionType": condition_type,
        "days": [day for day in constraint.get("days", []) if day in DEFAULT_DAYS],
        "periodsText": as_text(constraint.get("periodsText")) or ",".join(str(item) for item in constraint.get("periods", []) if parse_positive_int(item)),
        "strength": "soft" if strength.lower() == "soft" else "hard",
        "priority": parse_positive_int(constraint.get("priority")) or (5 if strength.lower() == "soft" else 10),
        "description": as_text(constraint.get("description")) or "AI 대화로 추가한 제약조건",
    }


def records_with_chat_constraints(records: dict, constraints) -> dict:
    if not isinstance(constraints, list) or not constraints:
        return records
    copied = deepcopy(records)
    copied.setdefault("constraints", [])
    existing = {
        (
            item.get("targetType"),
            item.get("targetCode"),
            item.get("conditionType"),
            tuple(item.get("days") or []),
            item.get("periodsText"),
            item.get("strength"),
        )
        for item in copied["constraints"]
    }
    for index, constraint in enumerate(constraints, start=1):
        normalized = normalized_chat_constraint(copied, constraint, index)
        if not normalized:
            continue
        key = (
            normalized.get("targetType"),
            normalized.get("targetCode"),
            normalized.get("conditionType"),
            tuple(normalized.get("days") or []),
            normalized.get("periodsText"),
            normalized.get("strength"),
        )
        if key in existing:
            continue
        existing.add(key)
        copied["constraints"].append(normalized)
    return copied


def schedule_dimensions(records: dict):
    config = records.get("config", {})
    days = parse_days(config.get("수업요일"))
    max_period = parse_positive_int(config.get("일일최대교시")) or DEFAULT_MAX_PERIOD
    return days, max_period


def truthy_config(value) -> bool:
    return as_text(value).upper() in {"Y", "YES", "TRUE", "1", "ON", "사용", "예"}


def constraint_settings(records: dict) -> dict:
    config = records.get("config", {})
    days, _ = schedule_dimensions(records)
    day_max = {}
    for day in days:
        value = parse_positive_int(config.get(f"교사{day}최대시수"))
        if value:
            day_max[day] = value
    return {
        "days": days,
        "lunchAfter": parse_positive_int(config.get("점심시간후교시")) or 0,
        "protectLunch": truthy_config(config.get("점심시간보호", "Y")),
        "maxConsecutive": parse_positive_int(config.get("최대연강허용")) or 0,
        "balanceStrength": as_text(config.get("균등분배강도")) or "soft",
        "aiPreferred": truthy_config(config.get("AI우선사용", "Y")),
        "assignmentMethod": as_text(config.get("배정방법")) or "fixed-first",
        "preferenceOrder": as_text(config.get("선호도순서")) or "안배>연강>식사시간",
        "metaIterations": min(parse_positive_int(config.get("배정횟수")) or 60, 240),
        "allowRelaxForUnassigned": truthy_config(config.get("미배정방지조건완화", "Y")),
        "teacherDayMaxEnabled": truthy_config(config.get("교사요일최대적용", "N")),
        "teacherDayMax": day_max,
    }


def apply_solve_options(records: dict, options: dict | None) -> dict:
    if not options:
        return records
    updates = {}
    mapping = {
        "assignmentMethod": "배정방법",
        "preferenceOrder": "선호도순서",
        "iterations": "배정횟수",
        "allowRelaxForUnassigned": "미배정방지조건완화",
        "teacherDayMaxEnabled": "교사요일최대적용",
        "maxConsecutive": "최대연강허용",
        "balanceStrength": "균등분배강도",
        "protectLunch": "점심시간보호",
    }
    for source, target in mapping.items():
        if source in options:
            updates[target] = options[source]
    day_max = options.get("teacherDayMax") if isinstance(options.get("teacherDayMax"), dict) else {}
    for day, value in day_max.items():
        if as_text(value):
            updates[f"교사{day}최대시수"] = value
    return records_with_config(records, updates)


def normalize_search_strength(value: str | None) -> str:
    normalized = as_text(value).lower()
    if normalized in {"fast", "quick", "빠른형"}:
        return "fast"
    if normalized in {"balanced", "balance", "균형형"}:
        return "balanced"
    return "strong"


def normalize_variation_mode(value: str | None) -> str:
    normalized = as_text(value).lower()
    if normalized in {"random", "full-random", "완전랜덤"}:
        return "random"
    if normalized in {"replay", "reproducible", "재현가능"}:
        return "reproducible"
    return "quality-first"


def search_iteration_budget(settings: dict, search_strength: str) -> int:
    base = parse_positive_int(settings.get("metaIterations")) or 60
    if search_strength == "fast":
        return max(24, min(base, 60))
    if search_strength == "balanced":
        return max(60, min(max(base, 84), 160))
    return max(96, min(max(base, 132), 240))


def search_generation_budget(iterations: int, search_strength: str) -> int:
    if search_strength == "fast":
        return max(3, min(5, iterations // 16))
    if search_strength == "balanced":
        return max(5, min(8, iterations // 14))
    return max(7, min(12, iterations // 12))


def solve_run_seed(solve_options: dict | None) -> int:
    explicit = parse_positive_int((solve_options or {}).get("seed"))
    if explicit:
        return explicit
    return random.SystemRandom().randint(1, 2_147_483_647)


def preference_weights(settings: dict) -> dict:
    order_text = settings.get("preferenceOrder") or ""
    tokens = [part.strip() for part in re.split(r"[>,/]+", order_text) if part.strip()]
    weights = {"balance": 1.8, "consecutive": 1.0, "lunch": 1.0, "sameSubject": 1.0, "dayMax": 1.0}
    aliases = {
        "안배": "balance",
        "균등": "balance",
        "평균": "balance",
        "연강": "consecutive",
        "3시간연속": "consecutive",
        "식사시간": "lunch",
        "점심": "lunch",
        "유사과목": "sameSubject",
        "같은날": "sameSubject",
        "최대": "dayMax",
    }
    for index, token in enumerate(tokens):
        multiplier = max(0.7, 1.9 - index * 0.25)
        for alias, key in aliases.items():
            if alias in token:
                weights[key] = multiplier
                break
    if settings.get("balanceStrength") == "hard":
        weights["balance"] *= 1.4
    elif settings.get("balanceStrength") == "off":
        weights["balance"] = 0.0
    return weights


def build_forbidden_index(records: dict):
    days, max_period = schedule_dimensions(records)
    index = defaultdict(set)
    for constraint in records.get("constraints", []):
        if constraint.get("conditionType") not in {"배정금지", "이동금지", "임시금지"}:
            continue
        if constraint.get("strength") == "soft":
            continue
        constraint_days = constraint.get("days") or days
        periods = parse_period_tokens(constraint.get("periodsText"), max_period)
        target_key = f"{constraint.get('targetType')}:{constraint.get('targetCode')}"
        for day in constraint_days:
            for period in periods:
                index[target_key].add((day, period))
    return index


def constraint_matches_entry(constraint: dict, entry: dict, day: str, period: int, max_period: int) -> bool:
    target_type = constraint.get("targetType")
    target_code = constraint.get("targetCode")
    target_map = {
        "교사": entry.get("teacherCode"),
        "학급": entry.get("classCode"),
        "과목": entry.get("subjectCode"),
        "특별실": entry.get("roomCode"),
    }
    if target_map.get(target_type) != target_code:
        return False
    days = constraint.get("days")
    if days and day not in days:
        return False
    return period in parse_period_tokens(constraint.get("periodsText"), max_period)


def soft_constraint_score(records: dict, entry: dict, day: str, period: int, max_period: int) -> float:
    score = 0.0
    for constraint in records.get("constraints", []):
        if constraint.get("strength") != "soft" and constraint.get("conditionType") not in {"희망", "비선호"}:
            continue
        if not constraint_matches_entry(constraint, entry, day, period, max_period):
            continue
        priority = parse_positive_int(constraint.get("priority")) or 5
        if constraint.get("conditionType") == "희망":
            score -= priority * 0.35
        else:
            score += priority * 0.7
    return score


def entry_for_load(load: dict, records: dict, block_size: int) -> dict:
    teacher = records["teachers"].get(load["teacherCode"], {})
    subject = records["subjects"].get(load["subjectCode"], {})
    klass = records["classes"].get(load["classCode"], {})
    room = records["rooms"].get(load.get("roomCode", ""), {})
    return {
        "teacherCode": load["teacherCode"],
        "teacherName": teacher.get("교사명", load["teacherCode"]),
        "subjectCode": load["subjectCode"],
        "subjectName": subject.get("과목명", load["subjectCode"]),
        "classCode": load["classCode"],
        "className": klass.get("학급명", load["classCode"]),
        "roomCode": load.get("roomCode", ""),
        "roomName": room.get("특별실명", ""),
        "blockSize": block_size,
        "source": "auto",
    }


def sync_entry_for_unit(unit: dict, records: dict) -> dict:
    load = unit["load"]
    entry = entry_for_load(load, records, 1)
    entry["syncGroup"] = as_text(load.get("syncGroup"))
    entry["syncOccurrenceId"] = unit.get("syncOccurrenceId", "")
    entry["syncLaneKey"] = unit.get("laneKey", sync_lane_key_for_load(load))
    return entry


def fixed_entry(item: dict) -> dict:
    return {
        "teacherCode": item.get("supervisorCode", ""),
        "teacherName": item.get("supervisorName", ""),
        "subjectCode": "__FIXED__",
        "subjectName": item.get("label") or "고정",
        "classCode": item["classCode"],
        "className": item.get("className", item["classCode"]),
        "roomCode": "",
        "roomName": "",
        "blockSize": 1,
        "source": "fixed",
        "fixed": True,
        "kind": item.get("kind", "비수업"),
        "description": item.get("description", ""),
    }


def empty_schedule(records: dict) -> dict:
    days, max_period = schedule_dimensions(records)
    classes = {}
    for class_code, item in records.get("classes", {}).items():
        day_limits = item.get("_dayLimits")
        if not day_limits:
            day_limits, _ = parse_day_hour_limits(item.get("요일별시수"), days, max_period)
        classes[class_code] = {
            "name": item.get("학급명", class_code),
            "dayLimits": {day: day_limits.get(day, max_period) for day in days},
            "grid": {day: {str(period): None for period in range(1, max_period + 1)} for day in days},
        }
    return {
        "days": days,
        "periods": list(range(1, max_period + 1)),
        "classes": classes,
    }


def class_period_available(schedule, class_code, day, period) -> bool:
    class_data = schedule.get("classes", {}).get(class_code, {})
    period = parse_int(period)
    if not class_data or not day or period is None or period < 1:
        return False
    grid = class_data.get("grid", {})
    if day not in grid or str(period) not in grid.get(day, {}):
        return False
    day_limit = parse_int(class_data.get("dayLimits", {}).get(day), len(schedule.get("periods", [])))
    return period <= day_limit


def apply_fixed_periods(records: dict, schedule: dict, teacher_busy=None) -> None:
    available_days = set(schedule.get("days", []))
    available_periods = set(schedule.get("periods", []))
    for item in records.get("fixedPeriods", []):
        class_code = item.get("classCode")
        day = item.get("day")
        period = item.get("period")
        if class_code not in schedule.get("classes", {}):
            continue
        if day not in available_days or period not in available_periods:
            continue
        if not class_period_available(schedule, class_code, day, period):
            continue
        cell = fixed_entry(item)
        schedule["classes"][class_code]["grid"][day][str(period)] = cell
        if teacher_busy is not None and cell.get("teacherCode"):
            teacher_busy[cell["teacherCode"]].add((day, period))


def hard_forbidden(index, entry, day, period) -> bool:
    keys = [
        f"교사:{entry['teacherCode']}",
        f"학급:{entry['classCode']}",
        f"과목:{entry['subjectCode']}",
    ]
    if entry.get("roomCode"):
        keys.append(f"특별실:{entry['roomCode']}")
    return any((day, period) in index.get(key, set()) for key in keys)


def slot_free(schedule, class_code, day, start_period, size, teacher_busy, room_busy, entry, max_period, forbidden):
    for offset in range(size):
        period = start_period + offset
        if period > max_period:
            return False
        if not class_period_available(schedule, class_code, day, period):
            return False
        if schedule["classes"][class_code]["grid"][day][str(period)]:
            return False
        if (day, period) in teacher_busy[entry["teacherCode"]]:
            return False
        if entry.get("roomCode") and (day, period) in room_busy[entry["roomCode"]]:
            return False
        if hard_forbidden(forbidden, entry, day, period):
            return False
    return True


def place_entry(schedule, class_code, day, start_period, size, teacher_busy, room_busy, entry):
    for offset in range(size):
        period = start_period + offset
        cell = deepcopy(entry)
        cell["blockIndex"] = offset + 1
        schedule["classes"][class_code]["grid"][day][str(period)] = cell
        teacher_busy[entry["teacherCode"]].add((day, period))
        if entry.get("roomCode"):
            room_busy[entry["roomCode"]].add((day, period))


def sync_occurrence_entries(occurrence: dict, records: dict) -> list[dict]:
    entries = []
    occurrence_id = occurrence.get("syncOccurrenceId", "")
    for unit in occurrence.get("units", []):
        prepared = dict(unit)
        prepared["syncOccurrenceId"] = occurrence_id
        entries.append(sync_entry_for_unit(prepared, records))
    return entries


def sync_slot_free(schedule, occurrence: dict, records: dict, day: str, period: int, teacher_busy, room_busy, forbidden, settings: dict) -> bool:
    seen_classes = set()
    seen_teachers = set()
    seen_rooms = set()
    for entry in sync_occurrence_entries(occurrence, records):
        class_code = entry.get("classCode")
        teacher_code = entry.get("teacherCode")
        room_code = entry.get("roomCode")
        if class_code in seen_classes:
            return False
        if teacher_code and teacher_code in seen_teachers:
            return False
        if room_code and room_code in seen_rooms:
            return False
        seen_classes.add(class_code)
        if teacher_code:
            seen_teachers.add(teacher_code)
        if room_code:
            seen_rooms.add(room_code)
        if not class_period_available(schedule, class_code, day, period):
            return False
        if schedule["classes"][class_code]["grid"][day][str(period)]:
            return False
        if teacher_code and (day, period) in teacher_busy[teacher_code]:
            return False
        if room_code and (day, period) in room_busy[room_code]:
            return False
        if hard_forbidden(forbidden, entry, day, period):
            return False
        if teacher_code and violates_teacher_flow(teacher_busy, teacher_code, day, [period], settings):
            return False
        if teacher_code and teacher_day_max_penalty(teacher_busy, teacher_code, day, 1, settings) is None:
            return False
    return True


def place_sync_occurrence(schedule, occurrence: dict, records: dict, day: str, period: int, teacher_busy, room_busy) -> None:
    for entry in sync_occurrence_entries(occurrence, records):
        class_code = entry["classCode"]
        cell = deepcopy(entry)
        cell["blockIndex"] = 1
        schedule["classes"][class_code]["grid"][day][str(period)] = cell
        if cell.get("teacherCode"):
            teacher_busy[cell["teacherCode"]].add((day, period))
        if cell.get("roomCode"):
            room_busy[cell["roomCode"]].add((day, period))


def day_subject_count(schedule, class_code, day, subject_code) -> int:
    return sum(1 for cell in schedule["classes"][class_code]["grid"][day].values() if cell and cell.get("subjectCode") == subject_code)


def teacher_day_count(teacher_busy, teacher_code, day) -> int:
    return sum(1 for busy_day, _ in teacher_busy[teacher_code] if busy_day == day)


def max_consecutive_count(periods: set[int]) -> int:
    longest = 0
    current = 0
    previous = None
    for period in sorted(periods):
        if previous is not None and period == previous + 1:
            current += 1
        else:
            current = 1
        longest = max(longest, current)
        previous = period
    return longest


def teacher_periods_for_day(teacher_busy, teacher_code, day) -> set[int]:
    return {period for busy_day, period in teacher_busy[teacher_code] if busy_day == day}


def violates_teacher_flow(teacher_busy, teacher_code, day, periods: list[int], settings: dict) -> bool:
    if not teacher_code:
        return False
    prospective = teacher_periods_for_day(teacher_busy, teacher_code, day) | set(periods)
    max_consecutive = settings.get("maxConsecutive") or 0
    if max_consecutive and max_consecutive_count(prospective) > max_consecutive:
        return True
    lunch_after = settings.get("lunchAfter") or 0
    if settings.get("protectLunch") and lunch_after and {lunch_after, lunch_after + 1}.issubset(prospective):
        return True
    return False


def period_bucket(period: int) -> str:
    if period <= 2:
        return "1-2"
    if period <= 4:
        return "3-4"
    if period <= 6:
        return "5-6"
    return "7+"


def teacher_balance_penalty(teacher_busy, teacher_code, day, period, settings: dict) -> float:
    if settings.get("balanceStrength") == "off":
        return 0.0
    weights = preference_weights(settings)
    lunch_after = settings.get("lunchAfter") or 0
    day_loads = Counter(busy_day for busy_day, _ in teacher_busy[teacher_code])
    projected_day_load = day_loads[day] + 1
    schedule_days = settings.get("days") or DEFAULT_DAYS
    average_day_load = (sum(day_loads.values()) + 1) / max(1, len(schedule_days))
    day_penalty = abs(projected_day_load - average_day_load) * 1.6 * weights["balance"]
    if len(day_loads) < len(schedule_days) and day_loads[day] > 0:
        day_penalty += 1.2 * weights["balance"]
    day_periods = teacher_periods_for_day(teacher_busy, teacher_code, day) | {period}
    if lunch_after:
        morning = sum(1 for item in day_periods if item <= lunch_after)
        afternoon = sum(1 for item in day_periods if item > lunch_after)
        segment_penalty = abs(morning - afternoon) * 1.1 * weights["balance"]
    else:
        midpoint = max(1, (max(day_periods) if day_periods else period) // 2)
        early = sum(1 for item in day_periods if item <= midpoint)
        late = sum(1 for item in day_periods if item > midpoint)
        segment_penalty = abs(early - late) * 0.8 * weights["balance"]
    bucket = period_bucket(period)
    existing_bucket_count = sum(
        1
        for _, existing_period in teacher_busy[teacher_code]
        if period_bucket(existing_period) == bucket
    )
    bucket_penalty = existing_bucket_count * 0.45 * weights["balance"]
    return day_penalty + segment_penalty + bucket_penalty


def teacher_day_max_penalty(teacher_busy, teacher_code, day, block_size: int, settings: dict) -> float | None:
    if not settings.get("teacherDayMaxEnabled"):
        return 0.0
    limit = settings.get("teacherDayMax", {}).get(day)
    if not limit:
        return 0.0
    projected = teacher_day_count(teacher_busy, teacher_code, day) + block_size
    if projected <= limit:
        return 0.0
    if not settings.get("allowRelaxForUnassigned"):
        return None
    return (projected - limit) * 4.0 * preference_weights(settings)["dayMax"]


def teacher_distribution_metrics(schedule: dict) -> dict:
    by_teacher = defaultdict(lambda: {"days": Counter(), "buckets": Counter(), "total": 0})
    schedule_days = schedule.get("days", [])
    available_buckets = sorted({period_bucket(parse_int(period) or 0) for period in schedule.get("periods", [])})
    for class_data in schedule.get("classes", {}).values():
        for day, periods in class_data.get("grid", {}).items():
            for period_text, cell in periods.items():
                if not cell or not cell.get("teacherCode") or cell.get("source") == "fixed":
                    continue
                period = parse_int(period_text)
                teacher = cell["teacherCode"]
                by_teacher[teacher]["days"][day] += 1
                by_teacher[teacher]["buckets"][period_bucket(period or 0)] += 1
                by_teacher[teacher]["total"] += 1
    imbalance = 0.0
    empty_weekdays = 0
    for data in by_teacher.values():
        if not data["total"]:
            continue
        day_counts = [data["days"].get(day, 0) for day in schedule_days]
        if day_counts:
            imbalance += (max(day_counts) - min(day_counts)) * 2.0
            expected = data["total"] / max(1, len(day_counts))
            imbalance += sum(abs(count - expected) for count in day_counts) * 0.8
        empty_weekdays += sum(1 for count in day_counts if count == 0)
        bucket_counts = [data["buckets"].get(bucket, 0) for bucket in available_buckets]
        if bucket_counts:
            imbalance += (max(bucket_counts) - min(bucket_counts)) * 1.5
            bucket_expected = data["total"] / max(1, len(bucket_counts))
            imbalance += sum(abs(count - bucket_expected) for count in bucket_counts) * 0.5
    return {"imbalance": imbalance, "emptyWeekdays": empty_weekdays}


def teacher_issue_summary(records: dict, schedule: dict, validation: dict | None = None) -> list[dict]:
    settings = constraint_settings(records)
    days = schedule.get("days", [])
    lunch_after = settings.get("lunchAfter") or 0
    by_teacher = defaultdict(lambda: {"name": "", "days": defaultdict(set), "total": 0})
    for class_data in schedule.get("classes", {}).values():
        for day, periods in class_data.get("grid", {}).items():
            for period_text, cell in periods.items():
                if not cell or not cell.get("teacherCode"):
                    continue
                period = parse_int(period_text)
                if period is None:
                    continue
                teacher = cell["teacherCode"]
                by_teacher[teacher]["name"] = cell.get("teacherName") or teacher
                by_teacher[teacher]["days"][day].add(period)
                if cell.get("source") != "fixed":
                    by_teacher[teacher]["total"] += 1

    conflict_teachers = set()
    for violation in (validation or {}).get("violations", []):
        if violation.get("type") == "teacher_conflict":
            match = re.search(r"교사\s+(\S+)", violation.get("message", ""))
            if match:
                conflict_teachers.add(match.group(1))

    issues = []
    for teacher, data in by_teacher.items():
        tags = []
        details = []
        day_counts = [len(data["days"].get(day, set())) for day in days]
        if data["total"] >= max(3, len(days)):
            empty_days = sum(1 for count in day_counts if count == 0)
            if day_counts and (max(day_counts) - min(day_counts) >= 3 or empty_days >= 2):
                tags.append("안배")
                details.append(f"요일편차 {max(day_counts)}:{min(day_counts)}")
        consecutive_days = []
        lunch_days = []
        for day in days:
            periods = data["days"].get(day, set())
            longest = max_consecutive_count(periods)
            if longest >= 3:
                consecutive_days.append(f"{day}{longest}")
            if lunch_after and {lunch_after, lunch_after + 1}.issubset(periods):
                lunch_days.append(day)
        if consecutive_days:
            tags.append("3연강")
            details.append(" ".join(consecutive_days[:3]))
        if lunch_days:
            tags.append("식사")
            details.append("점심전후 " + ",".join(lunch_days[:3]))
        if teacher in conflict_teachers:
            tags.append("중복")
            details.append("시간중복")
        if tags:
            issues.append({
                "teacherCode": teacher,
                "teacherName": data["name"] or teacher,
                "totalHours": data["total"],
                "issues": tags,
                "details": details,
                "severity": "error" if "중복" in tags else "warning",
            })
    return sorted(issues, key=lambda item: item["teacherName"])


def timetable_quality_score(candidate: dict) -> float:
    validation = candidate.get("validation", {})
    errors = len([item for item in validation.get("violations", []) if item.get("severity") == "error"])
    unassigned = len(candidate.get("unassigned", []))
    metrics = teacher_distribution_metrics(candidate.get("schedule", {}))
    return max(0.0, 1000.0 - unassigned * 180.0 - errors * 140.0 - metrics["imbalance"] * 7.0 - metrics["emptyWeekdays"] * 12.0)


def solve_greedy(records: dict, strategy: str, gene: dict | None = None) -> dict:
    gene = gene or {}
    rng = random.Random(gene.get("seed", 0)) if gene.get("seed") is not None else random.Random()
    randomness = float(gene.get("randomness", 0.0) or 0.0)
    schedule = empty_schedule(records)
    days, max_period = schedule_dimensions(records)
    settings = constraint_settings(records)
    weights = preference_weights(settings)
    forbidden = build_forbidden_index(records)
    teacher_busy = defaultdict(set)
    room_busy = defaultdict(set)
    unassigned = []
    apply_fixed_periods(records, schedule, teacher_busy)

    sync_occurrences = [
        occurrence
        for bundle in records.get("syncBundles", [])
        for occurrence in bundle.get("occurrences", [])
    ]
    if randomness:
        sync_occurrences = sync_occurrences[:]
        rng.shuffle(sync_occurrences)
    else:
        sync_occurrences.sort(key=lambda item: (item.get("syncGroup", ""), item.get("syncOccurrenceId", "")))

    for occurrence in sync_occurrences:
        candidates = []
        day_order = list(days)
        period_order = list(range(1, max_period + 1))
        if randomness:
            rng.shuffle(day_order)
            rng.shuffle(period_order)
        entries = sync_occurrence_entries(occurrence, records)
        for day_index, day in enumerate(day_order):
            for period in period_order:
                if not sync_slot_free(schedule, occurrence, records, day, period, teacher_busy, room_busy, forbidden, settings):
                    continue
                same_day = sum(day_subject_count(schedule, entry["classCode"], day, entry["subjectCode"]) for entry in entries)
                teacher_load = sum(teacher_day_count(teacher_busy, entry["teacherCode"], day) for entry in entries if entry.get("teacherCode"))
                balance_penalty = sum(
                    teacher_balance_penalty(teacher_busy, entry["teacherCode"], day, period, settings)
                    for entry in entries
                    if entry.get("teacherCode")
                )
                constraint_penalty = sum(soft_constraint_score(records, entry, day, period, max_period) for entry in entries)
                day_max_penalty = sum(
                    teacher_day_max_penalty(teacher_busy, entry["teacherCode"], day, 1, settings) or 0
                    for entry in entries
                    if entry.get("teacherCode")
                )
                score = teacher_load * 1.6 + same_day * 2.4 * weights["sameSubject"] + balance_penalty * 1.25 + constraint_penalty + day_max_penalty + day_index * 0.05
                if randomness:
                    score += rng.uniform(-0.4, 0.4) * randomness
                candidates.append((score, day, period))
        if candidates:
            _, day, period = min(candidates, key=lambda item: item[0])
            place_sync_occurrence(schedule, occurrence, records, day, period, teacher_busy, room_busy)
            continue
        for entry in entries:
            unassigned.append({
                "teacherCode": entry["teacherCode"],
                "teacherName": entry.get("teacherName") or display_name(records, "교사", entry["teacherCode"]),
                "subjectCode": entry["subjectCode"],
                "subjectName": entry.get("subjectName") or display_name(records, "과목", entry["subjectCode"]),
                "classCode": entry["classCode"],
                "className": entry.get("className") or display_name(records, "학급", entry["classCode"]),
                "hours": 1,
                "syncGroup": entry.get("syncGroup", ""),
                "syncOccurrenceId": entry.get("syncOccurrenceId", ""),
                "syncLaneKey": entry.get("syncLaneKey", ""),
                "reason": f"동시그룹 {entry.get('syncGroup', '')} 전체가 들어갈 공통 교시를 찾지 못했습니다.",
            })

    blocks = []
    for load in records.get("loads", []):
        if as_text(load.get("syncGroup")):
            continue
        for block_size in parse_block_pattern(load.get("continuousPattern"), load["weeklyHours"]):
            blocks.append((load, block_size))
    if strategy in {"spread-days", "spread-periods", "genetic-balanced"}:
        blocks.sort(key=lambda item: (-item[1], item[0]["teacherCode"], item[0]["classCode"], item[0]["subjectCode"]))
    elif strategy == "unassigned-first":
        blocks.sort(key=lambda item: (-item[0]["weeklyHours"], -item[1], item[0]["teacherCode"]))
    if strategy == "special-room-first":
        blocks.sort(key=lambda item: (0 if item[0].get("roomCode") else 1, -item[1], item[0]["teacherCode"]))
    elif strategy == "balanced":
        blocks.sort(key=lambda item: (-item[1], item[0]["classCode"], item[0]["subjectCode"]))
    elif strategy == "gap-light":
        blocks.sort(key=lambda item: (item[0]["teacherCode"], -item[1], item[0]["classCode"]))
    if randomness:
        blocks = [item for _, item in sorted(enumerate(blocks), key=lambda pair: pair[0] + rng.random() * randomness)]

    for load, block_size in blocks:
        entry = entry_for_load(load, records, block_size)
        candidates = []
        day_order = list(days)
        period_order = list(range(1, max_period + 1))
        if randomness:
            rng.shuffle(day_order)
            rng.shuffle(period_order)
        for day_index, day in enumerate(day_order):
            for period in period_order:
                if not slot_free(schedule, load["classCode"], day, period, block_size, teacher_busy, room_busy, entry, max_period, forbidden):
                    continue
                periods = [period + offset for offset in range(block_size)]
                if violates_teacher_flow(teacher_busy, load["teacherCode"], day, periods, settings):
                    continue
                day_max_penalty = teacher_day_max_penalty(teacher_busy, load["teacherCode"], day, block_size, settings)
                if day_max_penalty is None:
                    continue
                same_day = day_subject_count(schedule, load["classCode"], day, load["subjectCode"])
                teacher_load = teacher_day_count(teacher_busy, load["teacherCode"], day)
                late_penalty = abs(period - ((max_period + 1) / 2)) * 0.08
                balance_penalty = sum(teacher_balance_penalty(teacher_busy, load["teacherCode"], day, item, settings) for item in periods)
                constraint_penalty = sum(soft_constraint_score(records, entry, day, item, max_period) for item in periods)
                consecutive_penalty = max_consecutive_count(teacher_periods_for_day(teacher_busy, load["teacherCode"], day) | set(periods)) * 0.2 * weights["consecutive"]
                lunch_penalty = 0.0
                lunch_after = settings.get("lunchAfter") or 0
                if lunch_after and {lunch_after, lunch_after + 1}.intersection(periods):
                    lunch_penalty = 0.35 * weights["lunch"]
                if strategy == "gap-light":
                    score = teacher_load * 1.6 + same_day * 2.2 * weights["sameSubject"] + late_penalty + balance_penalty + constraint_penalty + consecutive_penalty + lunch_penalty + (day_max_penalty or 0)
                elif strategy == "special-room-first":
                    score = (0 if load.get("roomCode") else 0.5) + same_day * 1.8 * weights["sameSubject"] + day_index * 0.05 + balance_penalty + constraint_penalty + consecutive_penalty + lunch_penalty + (day_max_penalty or 0)
                elif strategy in {"spread-days", "spread-periods", "genetic-balanced"}:
                    score = teacher_load * 2.0 + same_day * 2.6 * weights["sameSubject"] + balance_penalty * 1.5 + constraint_penalty + consecutive_penalty + lunch_penalty + (day_max_penalty or 0)
                else:
                    score = teacher_load * 1.2 + same_day * 2.2 * weights["sameSubject"] + day_index * 0.05 + late_penalty + balance_penalty + constraint_penalty + consecutive_penalty + lunch_penalty + (day_max_penalty or 0)
                if randomness:
                    score += rng.uniform(-0.4, 0.4) * randomness
                candidates.append((score, day, period))
        if not candidates:
            unassigned.append({
                "teacherCode": load["teacherCode"],
                "teacherName": load.get("teacherName") or display_name(records, "교사", load["teacherCode"]),
                "subjectCode": load["subjectCode"],
                "subjectName": load.get("subjectName") or display_name(records, "과목", load["subjectCode"]),
                "classCode": load["classCode"],
                "className": load.get("className") or display_name(records, "학급", load["classCode"]),
                "hours": block_size,
                "reason": "배정 가능한 빈 교시를 찾지 못했습니다. 배정금지, 최대연강, 점심시간보호, 학급 요일별시수 조건을 함께 확인하세요.",
            })
            continue
        _, day, period = min(candidates, key=lambda item: item[0])
        place_entry(schedule, load["classCode"], day, period, block_size, teacher_busy, room_busy, entry)

    repair_notes = []
    if unassigned:
        unassigned, repair_notes = repair_unassigned_blocks(records, schedule, unassigned, teacher_busy, room_busy, forbidden, settings, max_period, gene)
    validation = validate_schedule(records, schedule, unassigned)
    diagnostics = diagnose_schedule(records, schedule, validation, unassigned)
    for note in repair_notes:
        diagnostics.insert(0, {
            "type": "repair",
            "severity": "success",
            "title": "미배정 자동 보정 기록",
            "reason": note,
            "suggestion": "검증 결과를 확인하세요.",
        })
    result = {
        "strategy": strategy,
        "schedule": schedule,
        "unassigned": unassigned,
        "validation": validation,
        "diagnostics": diagnostics,
        "algorithm": "greedy-seed",
        "gene": {key: value for key, value in gene.items() if key != "apiKey"},
        "quality": teacher_distribution_metrics(schedule),
        "teacherIssues": teacher_issue_summary(records, schedule, validation),
        "repairNotes": repair_notes,
        "score": 0,
    }
    result["score"] = timetable_quality_score(result)
    return result


def load_for_unassigned(records: dict, item: dict) -> dict | None:
    for load in records.get("loads", []):
        if (
            load.get("teacherCode") == item.get("teacherCode")
            and load.get("subjectCode") == item.get("subjectCode")
            and load.get("classCode") == item.get("classCode")
        ):
            return load
    if item.get("teacherCode") and item.get("subjectCode") and item.get("classCode"):
        return {
            "teacherCode": item.get("teacherCode"),
            "subjectCode": item.get("subjectCode"),
            "classCode": item.get("classCode"),
            "weeklyHours": item.get("hours", 1),
            "roomCode": item.get("roomCode", ""),
        }
    return None


def remove_cell_from_busy(cell: dict, day: str, period: int, teacher_busy, room_busy) -> None:
    if cell.get("teacherCode"):
        teacher_busy[cell["teacherCode"]].discard((day, period))
    if cell.get("roomCode"):
        room_busy[cell["roomCode"]].discard((day, period))


def restore_single_cell(schedule: dict, class_code: str, day: str, period: int, cell: dict, teacher_busy, room_busy) -> None:
    schedule["classes"][class_code]["grid"][day][str(period)] = cell
    if cell.get("teacherCode"):
        teacher_busy[cell["teacherCode"]].add((day, period))
    if cell.get("roomCode"):
        room_busy[cell["roomCode"]].add((day, period))


def can_place_repair_entry(records: dict, schedule: dict, load: dict, entry: dict, day: str, period: int, block_size: int, teacher_busy, room_busy, forbidden, settings: dict, max_period: int) -> bool:
    if not slot_free(schedule, load["classCode"], day, period, block_size, teacher_busy, room_busy, entry, max_period, forbidden):
        return False
    periods = [period + offset for offset in range(block_size)]
    if violates_teacher_flow(teacher_busy, load["teacherCode"], day, periods, settings):
        return False
    return teacher_day_max_penalty(teacher_busy, load["teacherCode"], day, block_size, settings) is not None


def repair_unassigned_blocks(records: dict, schedule: dict, unassigned: list[dict], teacher_busy, room_busy, forbidden, settings: dict, max_period: int, gene: dict | None = None) -> tuple[list[dict], list[str]]:
    if not unassigned:
        return [], []
    rng = random.Random((gene or {}).get("seed", 0) + 911)
    remaining = []
    notes = []
    days = list(schedule.get("days", []))
    periods = list(schedule.get("periods", []))

    for item in unassigned:
        if item.get("syncOccurrenceId"):
            remaining.append(item)
            continue
        load = load_for_unassigned(records, item)
        block_size = parse_positive_int(item.get("hours")) or 1
        if not load:
            remaining.append(item)
            continue
        entry = entry_for_load(load, records, block_size)
        day_order = days[:]
        period_order = periods[:]
        rng.shuffle(day_order)
        rng.shuffle(period_order)
        placed = False
        for day in day_order:
            for period in period_order:
                if can_place_repair_entry(records, schedule, load, entry, day, period, block_size, teacher_busy, room_busy, forbidden, settings, max_period):
                    place_entry(schedule, load["classCode"], day, period, block_size, teacher_busy, room_busy, entry)
                    notes.append(f"{describe_unassigned_item(records, item)} 직접 배치")
                    placed = True
                    break
            if placed:
                break
        if placed:
            continue

        if block_size != 1:
            remaining.append(item)
            continue
        class_code = load["classCode"]
        class_grid = schedule.get("classes", {}).get(class_code, {}).get("grid", {})
        for day in day_order:
            for period in period_order:
                target_cell = class_grid.get(day, {}).get(str(period))
                if not target_cell or target_cell.get("source") == "fixed" or target_cell.get("blockSize", 1) != 1:
                    continue
                remove_cell_from_busy(target_cell, day, period, teacher_busy, room_busy)
                class_grid[day][str(period)] = None
                if not can_place_repair_entry(records, schedule, load, entry, day, period, 1, teacher_busy, room_busy, forbidden, settings, max_period):
                    restore_single_cell(schedule, class_code, day, period, target_cell, teacher_busy, room_busy)
                    continue
                moved_target = False
                target_load = {
                    "teacherCode": target_cell.get("teacherCode", ""),
                    "subjectCode": target_cell.get("subjectCode", ""),
                    "classCode": class_code,
                    "roomCode": target_cell.get("roomCode", ""),
                    "weeklyHours": 1,
                }
                for move_day in day_order:
                    for move_period in period_order:
                        if move_day == day and move_period == period:
                            continue
                        if can_place_repair_entry(records, schedule, target_load, target_cell, move_day, move_period, 1, teacher_busy, room_busy, forbidden, settings, max_period):
                            place_entry(schedule, class_code, move_day, move_period, 1, teacher_busy, room_busy, target_cell)
                            moved_target = True
                            break
                    if moved_target:
                        break
                if moved_target:
                    place_entry(schedule, class_code, day, period, 1, teacher_busy, room_busy, entry)
                    notes.append(f"{describe_unassigned_item(records, item)} 이동 후 배치")
                    placed = True
                    break
                restore_single_cell(schedule, class_code, day, period, target_cell, teacher_busy, room_busy)
            if placed:
                break
        if not placed:
            remaining.append(item)
    return remaining, notes


def candidate_error_count(candidate: dict) -> int:
    return len([item for item in candidate.get("validation", {}).get("violations", []) if item.get("severity") == "error"])


def candidate_unassigned_count(candidate: dict) -> int:
    return len(candidate.get("unassigned", []))


def needs_repair_candidates(candidates: list[dict]) -> bool:
    return any(candidate_error_count(item) or candidate_unassigned_count(item) for item in candidates)


def compact_relaxation_text(text: str) -> str:
    value = as_text(text)
    if "최대연강" in value or value.startswith("연강"):
        numbers = re.findall(r"\d+", value)
        if len(numbers) >= 2:
            return f"연강 {numbers[0]}→{numbers[1]}"
        return "연강 완화"
    if "점심" in value or "식사" in value:
        return "점심보호 해제"
    if "균등" in value or "안배" in value:
        return "안배 완화"
    if "최대시수" in value or "요일최대" in value:
        return "요일최대 해제"
    return value.replace("미배정 방지를 위해 ", "").replace("했습니다.", "")


def compact_relaxations(relaxations: list[str]) -> list[str]:
    compacted = []
    for item in relaxations or []:
        text = compact_relaxation_text(item)
        if text and text not in compacted:
            compacted.append(text)
    return compacted


def annotate_repair_candidate(candidate: dict, strategy: str, effective_config: dict, relaxations: list[str]) -> dict:
    candidate["strategy"] = strategy
    candidate["aiGenerated"] = True
    candidate["effectiveConfig"] = effective_config
    candidate["relaxations"] = compact_relaxations(relaxations)
    candidate["score"] = max(0, candidate.get("score", 0) - len(relaxations) * 6)
    if candidate["relaxations"]:
        candidate.setdefault("diagnostics", []).insert(0, {
            "type": "relaxation",
            "severity": "warning",
            "title": "AI 개선 후보",
            "reason": " · ".join(candidate["relaxations"]),
            "suggestion": "완화 조건을 확인하세요.",
        })
    return candidate


def generate_ai_repair_candidates(records: dict, strict_candidates: list[dict]) -> list[dict]:
    if not needs_repair_candidates(strict_candidates):
        return []
    days, max_period = schedule_dimensions(records)
    settings = constraint_settings(records)
    base_config = records.get("config", {})
    profiles = []
    if settings.get("balanceStrength") != "off":
        profiles.append((
            "ai-balance-off",
            {"균등분배강도": "off"},
            ["균등분배강도를 off로 낮춰 미배정 감소를 우선했습니다."],
        ))
    if settings.get("protectLunch"):
        profiles.append((
            "ai-lunch-relaxed",
            {"점심시간보호": "N"},
            ["점심시간보호를 N으로 완화했습니다."],
        ))
    if settings.get("maxConsecutive") and settings["maxConsecutive"] < max_period:
        next_limit = min(max_period, settings["maxConsecutive"] + 1)
        profiles.append((
            "ai-consecutive-plus1",
            {"최대연강허용": str(next_limit)},
            [f"최대연강허용을 {settings['maxConsecutive']}에서 {next_limit}로 완화했습니다."],
        ))
    combined = {}
    combined_relaxations = []
    if settings.get("protectLunch"):
        combined["점심시간보호"] = "N"
        combined_relaxations.append("점심시간보호를 N으로 완화했습니다.")
    if settings.get("maxConsecutive") and settings["maxConsecutive"] < max_period:
        next_limit = min(max_period, settings["maxConsecutive"] + 1)
        combined["최대연강허용"] = str(next_limit)
        combined_relaxations.append(f"최대연강허용을 {settings['maxConsecutive']}에서 {next_limit}로 완화했습니다.")
    if combined:
        profiles.append(("ai-relax-combined", combined, combined_relaxations))

    repair_candidates = []
    seen_configs = set()
    for strategy, updates, relaxations in profiles:
        effective_config = {**base_config, **updates}
        key = tuple(sorted(effective_config.items()))
        if key in seen_configs:
            continue
        seen_configs.add(key)
        effective_records = records_with_config(records, updates)
        candidate = solve_greedy(effective_records, "balanced")
        repair_candidates.append(annotate_repair_candidate(candidate, strategy, effective_config, relaxations))
    return repair_candidates


def relaxation_profiles(records: dict, include_relaxations: bool = True) -> list[tuple[str, dict, list[str]]]:
    _, max_period = schedule_dimensions(records)
    settings = constraint_settings(records)
    profiles = [("strict", {}, [])]
    if not include_relaxations or not settings.get("allowRelaxForUnassigned"):
        return profiles
    if settings.get("maxConsecutive") and settings["maxConsecutive"] < max_period:
        for next_limit in range(settings["maxConsecutive"] + 1, min(max_period, settings["maxConsecutive"] + 2) + 1):
            profiles.append((
                f"relax-consecutive-{next_limit}",
                {"최대연강허용": str(next_limit)},
                [f"연강 {settings['maxConsecutive']}→{next_limit}"],
            ))
    if settings.get("protectLunch"):
        profiles.append((
            "relax-lunch",
            {"점심시간보호": "N"},
            ["점심보호 해제"],
        ))
    if settings.get("balanceStrength") != "off":
        profiles.append((
            "relax-balance",
            {"균등분배강도": "off"},
            ["안배 완화"],
        ))
    if settings.get("teacherDayMaxEnabled"):
        profiles.append((
            "relax-day-max",
            {"교사요일최대적용": "N"},
            ["요일최대 해제"],
        ))
    combined = {}
    combined_notes = []
    for _, updates, notes in profiles[1:]:
        combined.update(updates)
        combined_notes.extend(notes)
    if combined:
        profiles.append(("relax-combined", combined, combined_notes))
    return profiles


def initial_genes(records: dict, rng: random.Random | None = None, seed_base: int | None = None, search_strength: str = "strong", iterations: int | None = None) -> list[dict]:
    rng = rng or random.Random(seed_base or solve_run_seed(None))
    settings = constraint_settings(records)
    iterations = iterations or settings.get("metaIterations") or 60
    method = settings.get("assignmentMethod", "fixed-first")
    strategies = ["genetic-balanced", "spread-days", "spread-periods", "balanced", "gap-light", "special-room-first", "unassigned-first"]
    if method == "unassigned-only":
        strategies = ["unassigned-first", "spread-days", "genetic-balanced", "balanced"]
    elif method == "from-start":
        strategies = ["spread-days", "spread-periods", "genetic-balanced", "balanced"]
    genes = []
    seed_base = seed_base or rng.randint(1, 1_000_000)
    for index in range(max(12, iterations)):
        strategy = strategies[index % len(strategies)]
        if search_strength == "strong" and index >= len(strategies):
            strategy = rng.choice(strategies)
        genes.append({
            "seed": seed_base + index * 9973 + rng.randint(0, 7919),
            "strategy": strategy,
            "randomness": min(1.6, 0.25 + (index % 7) * 0.11 + (0.18 if search_strength == "strong" else 0.0)),
        })
    return genes


def solve_gene(records: dict, gene: dict, profile: tuple[str, dict, list[str]]) -> dict:
    profile_name, updates, relaxations = profile
    effective_records = records_with_config(records, updates)
    strategy = gene.get("strategy", "genetic-balanced")
    candidate = solve_greedy(effective_records, strategy, gene=gene)
    candidate["strategy"] = f"ga-{profile_name}-{strategy}-{gene.get('seed')}"
    candidate["algorithm"] = "metaheuristic-genetic"
    candidate["effectiveConfig"] = {**records.get("config", {}), **updates}
    candidate["relaxations"] = compact_relaxations(relaxations)
    candidate["aiGenerated"] = bool(relaxations)
    if candidate["relaxations"]:
        candidate.setdefault("diagnostics", []).insert(0, {
            "type": "relaxation",
            "severity": "warning",
            "title": "미배정 방지 완화",
            "reason": " · ".join(candidate["relaxations"]),
            "suggestion": "완화 조건을 확인하세요.",
        })
    return candidate


def crossover_gene(parent_a: dict, parent_b: dict, rng: random.Random, generation: int) -> dict:
    return {
        "seed": rng.randint(1, 1_000_000) + generation,
        "strategy": rng.choice([parent_a.get("strategy", "genetic-balanced"), parent_b.get("strategy", "spread-days")]),
        "randomness": max(0.05, min(1.4, (float(parent_a.get("randomness", 0.4)) + float(parent_b.get("randomness", 0.4))) / 2 + rng.uniform(-0.18, 0.18))),
    }


def solve_metaheuristic(records: dict, include_relaxations: bool = True, seed: int | None = None, search_strength: str = "strong", return_stats: bool = False):
    settings = constraint_settings(records)
    iterations = search_iteration_budget(settings, normalize_search_strength(search_strength))
    seed = seed or solve_run_seed(None)
    rng = random.Random(seed)
    profiles = relaxation_profiles(records, include_relaxations=include_relaxations)
    genes = initial_genes(records, rng=rng, seed_base=seed, search_strength=normalize_search_strength(search_strength), iterations=iterations)
    population = []
    attempt_count = 0
    timed_out = False
    for gene in genes[: max(10, min(len(genes), iterations))]:
        for profile in profiles:
            population.append(solve_gene(records, gene, profile))
            attempt_count += 1
    if not population:
        population = [
            solve_gene(records, genes[0] if genes else {"seed": seed, "strategy": "balanced", "randomness": 0.0}, profiles[0])
        ]
        attempt_count += 1
    generations = search_generation_budget(iterations, normalize_search_strength(search_strength))
    for generation in range(generations):
        elites = sorted(population, key=candidate_rank, reverse=True)[: max(4, min(10, len(population)))]
        children = []
        for index in range(max(6, iterations // 8)):
            parent_a = elites[index % len(elites)].get("gene", {})
            parent_b = elites[(index * 3 + 1) % len(elites)].get("gene", {})
            child_gene = crossover_gene(parent_a, parent_b, rng, generation)
            if normalize_search_strength(search_strength) == "strong" and rng.random() < 0.35:
                child_gene["strategy"] = rng.choice(["genetic-balanced", "spread-days", "spread-periods", "balanced", "gap-light", "special-room-first", "unassigned-first"])
                child_gene["randomness"] = min(1.7, float(child_gene.get("randomness", 0.4)) + rng.uniform(0.05, 0.28))
            profile = profiles[index % len(profiles)]
            children.append(solve_gene(records, child_gene, profile))
            attempt_count += 1
        population = elites + children
        if any(candidate_unassigned_count(candidate) == 0 and candidate_error_count(candidate) == 0 for candidate in elites):
            if generation >= (2 if normalize_search_strength(search_strength) != "strong" else 4):
                break
    unique = {}
    for candidate in population:
        key = candidate_signature(candidate)
        if key not in unique or candidate_rank(candidate) > candidate_rank(unique[key]):
            unique[key] = candidate
    candidates = sorted(unique.values(), key=candidate_rank, reverse=True)[:8]
    stats = {
        "attemptCount": attempt_count,
        "generationCount": generations,
        "populationCount": len(population),
        "uniqueCandidateCount": len(unique),
        "profileCount": len(profiles),
        "searchStrength": normalize_search_strength(search_strength),
        "timedOut": timed_out,
    }
    return (candidates, stats) if return_stats else candidates


def candidate_rank(candidate: dict):
    relaxation_count = len(candidate.get("relaxations", []))
    return (
        -candidate_unassigned_count(candidate),
        -candidate_error_count(candidate),
        -relaxation_count,
        candidate.get("score", 0),
    )


def schedule_signature(schedule: dict) -> str:
    parts = []
    for class_code, class_data in sorted((schedule.get("classes") or {}).items()):
        grid = class_data.get("grid", {})
        for day in schedule.get("days", []):
            for period in schedule.get("periods", []):
                cell = grid.get(day, {}).get(str(period))
                if not cell:
                    continue
                parts.append("|".join([
                    class_code,
                    as_text(day),
                    str(period),
                    as_text(cell.get("teacherCode")),
                    as_text(cell.get("subjectCode")),
                    as_text(cell.get("roomCode")),
                    as_text(cell.get("source")),
                ]))
    return hashlib.sha256("\n".join(parts).encode("utf-8")).hexdigest()[:20]


def records_signature(records: dict) -> str:
    loads = []
    for item in records.get("loads", []):
        loads.append({
            "teacherCode": item.get("teacherCode"),
            "subjectCode": item.get("subjectCode"),
            "classCode": item.get("classCode"),
            "weeklyHours": item.get("weeklyHours"),
            "blockPattern": item.get("blockPattern"),
            "roomCode": item.get("roomCode"),
            "syncGroup": item.get("syncGroup"),
            "coTeacherGroup": item.get("coTeacherGroup"),
        })
    classes = {
        code: {
            "dayLimits": item.get("_dayLimits") or {},
            "grade": item.get("학년"),
            "series": item.get("계열"),
            "virtual": item.get("가상학급여부"),
        }
        for code, item in sorted((records.get("classes") or {}).items())
    }
    payload = {
        "config": records.get("config", {}),
        "teachers": sorted((records.get("teachers") or {}).keys()),
        "classes": classes,
        "subjects": sorted((records.get("subjects") or {}).keys()),
        "rooms": sorted((records.get("rooms") or {}).keys()),
        "loads": sorted(loads, key=lambda item: json.dumps(item, ensure_ascii=False, sort_keys=True)),
        "constraints": records.get("constraints", []),
        "fixedPeriods": records.get("fixedPeriods", []),
        "syncGroups": records.get("syncGroups", []),
        "continuous": records.get("continuous", []),
        "coTeachers": records.get("coTeachers", []),
    }
    serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(serialized.encode("utf-8")).hexdigest()[:20]


def candidate_signature(candidate: dict) -> str:
    signature = candidate.get("signature")
    if signature:
        return signature
    return schedule_signature(candidate.get("schedule", {}))


def annotate_candidate_signatures(candidates: list[dict]) -> list[dict]:
    for candidate in candidates:
        candidate["signature"] = candidate_signature(candidate)
    return candidates


def candidate_compatible_with_records(records: dict, candidate: dict | None) -> bool:
    if not candidate:
        return False
    schedule = candidate.get("schedule") or {}
    return set((schedule.get("classes") or {}).keys()) == set((records.get("classes") or {}).keys())


def select_quality_first_candidate(records: dict, candidates: list[dict], variation_mode: str, previous_result: dict | None) -> tuple[dict, bool, str]:
    ranked = sorted(candidates, key=candidate_rank, reverse=True)
    if not ranked:
        raise ValueError("선택할 후보 시간표가 없습니다.")
    best = ranked[0]
    if variation_mode == "random":
        pool = ranked[: min(4, len(ranked))]
        selected = random.choice(pool)
        return selected, True, "random"
    previous = deepcopy((previous_result or {}).get("selected") or {})
    previous_records_signature = (previous_result or {}).get("recordSignature")
    current_records_signature = records_signature(records)
    if (
        variation_mode != "quality-first"
        or previous_records_signature != current_records_signature
        or not candidate_compatible_with_records(records, previous)
    ):
        return best, True, "new-best"

    previous["signature"] = candidate_signature(previous)
    previous_rank = candidate_rank(previous)
    previous_signature = previous["signature"]
    eligible = [candidate for candidate in ranked if candidate_rank(candidate) >= previous_rank]
    diverse = [candidate for candidate in eligible if candidate_signature(candidate) != previous_signature]
    if diverse:
        return diverse[0], True, "new-diverse"
    if eligible:
        selected = eligible[0]
        return selected, candidate_signature(selected) != previous_signature, "new-equivalent"

    previous["retainedPrevious"] = True
    previous.setdefault("diagnostics", []).insert(0, {
        "type": "quality-first-retained",
        "severity": "warning",
        "title": "이전 후보 유지",
        "reason": "새 탐색 후보가 이전 시간표보다 미배정/검증 기준에서 나빠 이전 결과를 유지했습니다.",
        "suggestion": "선호도 팝업에서 강도를 높이거나 조건 완화를 켠 뒤 다시 탐색하세요.",
    })
    return previous, False, "previous-retained"


def selected_candidate_list(selected: dict, candidates: list[dict], limit: int = 4) -> list[dict]:
    selected_signature = candidate_signature(selected)
    output = [selected]
    for candidate in sorted(candidates, key=candidate_rank, reverse=True):
        if candidate_signature(candidate) == selected_signature:
            continue
        output.append(candidate)
        if len(output) >= limit:
            break
    return output


def diagnose_schedule(records: dict, schedule: dict, validation: dict, unassigned=None) -> list[dict]:
    unassigned = unassigned or []
    diagnostics = []
    for item in unassigned:
        diagnostics.append({
            "type": "unassigned",
            "severity": "error",
            "title": f"{describe_unassigned_item(records, item)} 미배정",
            "reason": item.get("reason", "배정 가능한 칸을 찾지 못했습니다."),
            "suggestion": "해당 교사/학급의 배정금지, 최대연강, 점심시간보호, 특별실 조건 중 우선순위가 낮은 조건을 완화해 보세요.",
        })
    for violation in validation.get("violations", []):
        diagnostics.append({
            "type": violation.get("type", "validation"),
            "severity": violation.get("severity", "warning"),
            "title": "검증 오류",
            "reason": violation.get("message", ""),
            "suggestion": "수동 이동이나 조건 우선순위를 조정한 뒤 다시 검증하세요.",
        })
    if not diagnostics:
        diagnostics.append({
            "type": "ok",
            "severity": "success",
            "title": "검증 통과",
            "reason": "미배정과 hard 검증 오류가 없습니다.",
            "suggestion": "후보 시간표를 비교하면서 교사별 오전/오후 균형을 확인하세요.",
        })
    return diagnostics


def summarize_candidate_for_ai(candidate: dict) -> dict:
    diagnostics = candidate.get("diagnostics", [])
    return {
        "strategy": candidate.get("strategy"),
        "score": candidate.get("score"),
        "unassignedCount": len(candidate.get("unassigned", [])),
        "errorCount": len([item for item in candidate.get("validation", {}).get("violations", []) if item.get("severity") == "error"]),
        "topDiagnostics": diagnostics[:5],
        "relaxations": candidate.get("relaxations", []),
        "aiGenerated": candidate.get("aiGenerated", False),
    }


AI_ADVICE_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "summary": {"type": "string"},
        "suggestions": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "type": {"type": "string"},
                    "title": {"type": "string"},
                    "explanation": {"type": "string"},
                    "steps": {"type": "array", "items": {"type": "string"}},
                },
                "required": ["type", "title", "explanation", "steps"],
            },
        },
    },
    "required": ["summary", "suggestions"],
}

AI_PROVIDER_LABELS = {
    "openai": "OpenAI",
    "gemini": "Gemini",
    "custom": "Custom",
}

AI_DEFAULT_MODELS = {
    "openai": os.environ.get("OPENAI_MODEL", "gpt-5.5"),
    "gemini": os.environ.get("GEMINI_MODEL", "gemini-3.5-flash"),
    "custom": "",
}

AI_SYSTEM_INSTRUCTIONS = (
    "You are an assistant helping Korean school staff build timetables. "
    "Only use the masked codes and aggregate data provided. Never infer or request real names. "
    "Return practical, verifiable suggestions that can reduce unassigned classes or validation errors. "
    "Do not claim a change is applied unless the optimization engine has validated it."
)


def normalize_provider(provider: str) -> str:
    value = as_text(provider).lower()
    if value in {"google", "google-gemini", "google_ai", "google-ai"}:
        return "gemini"
    if value in {"openai-compatible", "openai_compatible", "custom-openai", "custom"}:
        return "custom"
    if value not in AI_PROVIDER_LABELS:
        return "openai"
    return value


def normalize_ai_config(config=None, *, api_key: str = "", api_validated: bool = False) -> dict:
    raw = config if isinstance(config, dict) else {}
    provider = normalize_provider(raw.get("provider", "openai"))
    model = as_text(raw.get("model")) or AI_DEFAULT_MODELS.get(provider, "")
    key = as_text(raw.get("apiKey")) or as_text(api_key)
    return {
        "provider": provider,
        "providerLabel": AI_PROVIDER_LABELS.get(provider, provider),
        "apiKey": key,
        "model": model,
        "baseUrl": as_text(raw.get("baseUrl")),
        "validated": bool(raw.get("validated") or api_validated),
    }


def ai_config_from_body(body: dict, *, require_validated: bool = False) -> dict:
    raw = body.get("aiConfig") if isinstance(body, dict) else {}
    config = normalize_ai_config(
        raw,
        api_key=body.get("apiKey", "") if isinstance(body, dict) else "",
        api_validated=bool(body.get("apiValidated")) if isinstance(body, dict) else False,
    )
    if require_validated and not config.get("validated"):
        config["apiKey"] = ""
    return config


def public_ai_config(config: dict) -> dict:
    return {
        "provider": config.get("provider", "openai"),
        "providerLabel": config.get("providerLabel", AI_PROVIDER_LABELS.get(config.get("provider", "openai"), "OpenAI")),
        "model": config.get("model", ""),
        "baseUrl": config.get("baseUrl", ""),
        "validated": bool(config.get("validated")),
    }


def ai_prompt(task: str, context: dict) -> str:
    return json.dumps({"task": task, "context": context}, ensure_ascii=False)


def extract_json_text(text: str) -> str:
    value = as_text(text).strip()
    fence = re.search(r"```(?:json)?\s*(.*?)```", value, flags=re.IGNORECASE | re.DOTALL)
    if fence:
        return fence.group(1).strip()
    start = value.find("{")
    end = value.rfind("}")
    if start != -1 and end != -1 and end > start:
        return value[start:end + 1].strip()
    return value


def parse_advice_json(text: str, provider_label: str) -> dict:
    cleaned = extract_json_text(text)
    try:
        advice = json.loads(cleaned)
    except json.JSONDecodeError:
        return {
            "ok": False,
            "status": "parse_error",
            "provider": provider_label,
            "message": f"{provider_label} 응답을 구조화된 제안으로 해석하지 못했습니다.",
            "rawText": as_text(text)[:1000],
        }
    if not isinstance(advice, dict):
        return {
            "ok": False,
            "status": "parse_error",
            "provider": provider_label,
            "message": f"{provider_label} 응답 형식이 올바르지 않습니다.",
            "rawText": as_text(text)[:1000],
        }
    advice.setdefault("summary", "")
    advice.setdefault("suggestions", [])
    return {"ok": True, "advice": advice}


def schema_without_unsupported_fields(schema):
    if isinstance(schema, dict):
        return {
            key: schema_without_unsupported_fields(value)
            for key, value in schema.items()
            if key not in {"additionalProperties"}
        }
    if isinstance(schema, list):
        return [schema_without_unsupported_fields(item) for item in schema]
    return schema


def extract_response_text(payload: dict) -> str:
    if isinstance(payload.get("output_text"), str):
        return payload["output_text"]
    parts = []
    for item in payload.get("output", []):
        for content in item.get("content", []):
            if content.get("type") == "output_text" and "text" in content:
                parts.append(content["text"])
    return "\n".join(parts).strip()


def extract_chat_completion_text(payload: dict) -> str:
    choices = payload.get("choices", [])
    if not choices:
        return ""
    message = choices[0].get("message", {})
    content = message.get("content", "")
    if isinstance(content, list):
        return "\n".join(as_text(item.get("text", "")) for item in content if isinstance(item, dict)).strip()
    return as_text(content)


def extract_gemini_text(payload: dict) -> str:
    parts = []
    for candidate in payload.get("candidates", []):
        content = candidate.get("content", {})
        for part in content.get("parts", []):
            if "text" in part:
                parts.append(as_text(part["text"]))
    return "\n".join(part for part in parts if part).strip()


def call_openai_advisor(api_key: str, task: str, context: dict, model: str | None = None) -> dict:
    api_key = as_text(api_key)
    if not api_key:
        return {"ok": False, "status": "missing", "provider": "OpenAI", "message": "OpenAI API 키가 없어 로컬 진단만 사용했습니다."}
    model = as_text(model) or AI_DEFAULT_MODELS["openai"]
    request_body = {
        "model": model,
        "instructions": AI_SYSTEM_INSTRUCTIONS,
        "input": ai_prompt(task, context),
        "text": {
            "format": {
                "type": "json_schema",
                "name": "timetable_ai_advice",
                "description": "Masked timetable advice for reducing errors and improving schedule quality.",
                "schema": AI_ADVICE_SCHEMA,
                "strict": False,
            }
        },
        "max_output_tokens": 1400,
    }
    raw = json.dumps(request_body, ensure_ascii=False).encode("utf-8")
    request = Request(
        "https://api.openai.com/v1/responses",
        data=raw,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=15) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as error:
        return {"ok": False, "status": "http_error", "provider": "OpenAI", "message": f"OpenAI 응답 오류: {error.code}"}
    except (URLError, TimeoutError, OSError, json.JSONDecodeError) as error:
        return {"ok": False, "status": "network_error", "provider": "OpenAI", "message": f"OpenAI 호출 실패: {error}"}

    parsed = parse_advice_json(extract_response_text(response_payload), "OpenAI")
    if not parsed.get("ok"):
        return parsed
    return {
        "ok": True,
        "status": "called",
        "provider": "OpenAI",
        "model": response_payload.get("model", model),
        "responseId": response_payload.get("id", ""),
        "message": "OpenAI 제안을 생성했습니다.",
        "advice": parsed["advice"],
    }


def call_gemini_advisor(config: dict, task: str, context: dict) -> dict:
    api_key = as_text(config.get("apiKey"))
    model = as_text(config.get("model")) or AI_DEFAULT_MODELS["gemini"]
    if not api_key:
        return {"ok": False, "status": "missing", "provider": "Gemini", "message": "Gemini API 키가 없어 로컬 진단만 사용했습니다."}
    request_body = {
        "contents": [{
            "parts": [{
                "text": f"{AI_SYSTEM_INSTRUCTIONS}\n\nReturn only JSON matching the schema.\n\n{ai_prompt(task, context)}"
            }]
        }],
        "generationConfig": {
            "responseMimeType": "application/json",
            "responseJsonSchema": schema_without_unsupported_fields(AI_ADVICE_SCHEMA),
            "maxOutputTokens": 1400,
        },
    }
    request = Request(
        f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent",
        data=json.dumps(request_body, ensure_ascii=False).encode("utf-8"),
        headers={
            "x-goog-api-key": api_key,
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=15) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as error:
        return {"ok": False, "status": "http_error", "provider": "Gemini", "message": f"Gemini 응답 오류: {error.code}"}
    except (URLError, TimeoutError, OSError, json.JSONDecodeError) as error:
        return {"ok": False, "status": "network_error", "provider": "Gemini", "message": f"Gemini 호출 실패: {error}"}

    parsed = parse_advice_json(extract_gemini_text(response_payload), "Gemini")
    if not parsed.get("ok"):
        return parsed
    return {
        "ok": True,
        "status": "called",
        "provider": "Gemini",
        "model": model,
        "responseId": "",
        "message": "Gemini 제안을 생성했습니다.",
        "advice": parsed["advice"],
    }


def custom_base_url(config: dict) -> str:
    return as_text(config.get("baseUrl")).rstrip("/")


def call_custom_advisor(config: dict, task: str, context: dict) -> dict:
    api_key = as_text(config.get("apiKey"))
    model = as_text(config.get("model"))
    base_url = custom_base_url(config)
    if not base_url:
        return {"ok": False, "status": "missing_base_url", "provider": "Custom", "message": "Custom 제공자는 Base URL이 필요합니다."}
    if not api_key:
        return {"ok": False, "status": "missing", "provider": "Custom", "message": "Custom API 키가 없어 로컬 진단만 사용했습니다."}
    if not model:
        return {"ok": False, "status": "missing_model", "provider": "Custom", "message": "Custom 제공자는 모델명이 필요합니다."}
    request_body = {
        "model": model,
        "messages": [
            {"role": "system", "content": AI_SYSTEM_INSTRUCTIONS},
            {"role": "user", "content": ai_prompt(task, context)},
        ],
        "response_format": {"type": "json_object"},
        "max_tokens": 1400,
    }
    request = Request(
        f"{base_url}/chat/completions",
        data=json.dumps(request_body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=15) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as error:
        return {"ok": False, "status": "http_error", "provider": "Custom", "message": f"Custom 응답 오류: {error.code}"}
    except (URLError, TimeoutError, OSError, json.JSONDecodeError) as error:
        return {"ok": False, "status": "network_error", "provider": "Custom", "message": f"Custom 호출 실패: {error}"}

    parsed = parse_advice_json(extract_chat_completion_text(response_payload), "Custom")
    if not parsed.get("ok"):
        return parsed
    return {
        "ok": True,
        "status": "called",
        "provider": "Custom",
        "model": response_payload.get("model", model),
        "responseId": response_payload.get("id", ""),
        "message": "Custom AI 제안을 생성했습니다.",
        "advice": parsed["advice"],
    }


def call_ai_advisor(config: dict, task: str, context: dict) -> dict:
    normalized = normalize_ai_config(config)
    provider = normalized["provider"]
    if provider == "gemini":
        return call_gemini_advisor(normalized, task, context)
    if provider == "custom":
        return call_custom_advisor(normalized, task, context)
    return call_openai_advisor(normalized.get("apiKey", ""), task, context, normalized.get("model"))


def local_solve_advice(ai_summary: dict) -> dict:
    diagnostics = ai_summary.get("topDiagnostics", [])
    if ai_summary.get("unassignedCount") or ai_summary.get("errorCount"):
        steps = [
            item.get("suggestion") or item.get("reason", "")
            for item in diagnostics
            if item.get("suggestion") or item.get("reason")
        ]
        if not steps:
            steps = ["최대연강, 점심보호, 특별실, 배정금지 조건을 우선순위가 낮은 것부터 완화해 보세요."]
        return {
            "summary": "미배정 또는 hard 검증 오류가 있어 조건 완화 후보를 우선 검토해야 합니다.",
            "suggestions": [{
                "type": "local_diagnostic",
                "title": "조건 완화 후보",
                "explanation": "검증 엔진이 찾은 오류를 기준으로 한 로컬 진단입니다.",
                "steps": steps[:5],
            }],
        }
    return {
        "summary": "미배정과 hard 검증 오류가 없습니다.",
        "suggestions": [{
            "type": "local_review",
            "title": "균형 검토",
            "explanation": "교사별 보기에서 요일별/오전오후 쏠림을 확인하면 됩니다.",
            "steps": ["교사별 시간표 보기로 전환", "오전/오후 편차가 큰 교사의 수업을 수동 이동", "이동 후 검증 실행"],
        }],
    }


def build_ai_solve_advisor(records: dict, ai_summary: dict, ai_config=None) -> dict:
    config = normalize_ai_config(ai_config)
    context = {
        "maskedRecords": mask_records_for_ai(records),
        "candidateSummary": ai_summary,
        "settings": constraint_settings(records),
    }
    local_advice = local_solve_advice(ai_summary)
    provider_label = config.get("providerLabel", "AI")
    remote = call_ai_advisor(config, "solve_review", context) if as_text(config.get("apiKey")) else {
        "ok": False,
        "status": "not-configured",
        "provider": provider_label,
        "message": f"{provider_label} API 키가 없어 로컬 진단만 사용했습니다.",
    }
    return {
        "mode": f"{config.get('provider')}-advisor" if remote.get("ok") else "local-fallback",
        "privacy": "AI에는 실제 이름 대신 코드와 집계, 검증 결과만 전달합니다.",
        "aiConfig": public_ai_config(config),
        "remote": remote,
        "advice": remote.get("advice") if remote.get("ok") else local_advice,
    }


def solve_schedule(records: dict, api_key: str = "", ai_config=None, solve_options: dict | None = None, persist: bool = True, advisor: bool = True) -> dict:
    solve_options = solve_options or {}
    started_at = time.monotonic()
    run_id = uuid.uuid4().hex[:12]
    seed = solve_run_seed(solve_options)
    search_strength = normalize_search_strength(solve_options.get("searchStrength"))
    variation_mode = normalize_variation_mode(solve_options.get("variationMode"))
    records = apply_solve_options(records, solve_options)
    record_signature = records_signature(records)
    config = normalize_ai_config(ai_config, api_key=api_key)
    strict_candidates = [
        solve_greedy(records, "balanced", gene={"seed": 11, "randomness": 0.0, "strategy": "balanced"}),
        solve_greedy(records, "gap-light", gene={"seed": 17, "randomness": 0.0, "strategy": "gap-light"}),
        solve_greedy(records, "special-room-first", gene={"seed": 23, "randomness": 0.0, "strategy": "special-room-first"}),
    ]
    for candidate in strict_candidates:
        candidate.setdefault("relaxations", [])
        candidate.setdefault("effectiveConfig", records.get("config", {}))
        candidate.setdefault("aiGenerated", False)
    needs_relaxation = needs_repair_candidates(strict_candidates)
    repair_candidates = generate_ai_repair_candidates(records, strict_candidates)
    genetic_candidates, genetic_stats = solve_metaheuristic(records, include_relaxations=needs_relaxation, seed=seed, search_strength=search_strength, return_stats=True)
    annotate_candidate_signatures(genetic_candidates)
    previous_result = load_last_schedule() if variation_mode == "quality-first" else None
    best, best_changed, selection_source = select_quality_first_candidate(records, genetic_candidates, variation_mode, previous_result)
    best["signature"] = candidate_signature(best)
    candidates = selected_candidate_list(best, genetic_candidates, limit=4)
    ai_summary = summarize_candidate_for_ai(best)
    elapsed_ms = int((time.monotonic() - started_at) * 1000)
    ai_advisor = build_ai_solve_advisor(records, ai_summary, config) if advisor else {
        "mode": "local-progress",
        "privacy": "진행형 탐색 중에는 원격 AI 조언을 건너뛰고 배정 엔진 결과만 갱신합니다.",
        "aiConfig": public_ai_config(config),
        "remote": {"ok": False, "status": "not-run", "provider": config.get("providerLabel", "AI"), "message": "진행형 탐색 chunk에서는 원격 AI를 호출하지 않았습니다."},
        "advice": local_solve_advice(ai_summary),
    }
    result = {
        "runId": run_id,
        "seed": seed,
        "recordSignature": record_signature,
        "createdAt": now_iso(),
        "bestStrategy": best["strategy"],
        "candidates": candidates,
        "selected": best,
        "aiSummary": ai_summary,
        "aiAdvisor": ai_advisor,
        "attemptCount": genetic_stats.get("attemptCount", 0),
        "bestChanged": best_changed,
        "bestSignature": best["signature"],
        "timedOut": False,
        "elapsedMs": elapsed_ms,
        "progressMessage": "자동배정 탐색 chunk가 완료되었습니다.",
        "searchStats": {
            **genetic_stats,
            "variationMode": variation_mode,
            "selectionSource": selection_source,
            "previousSignature": candidate_signature((previous_result or {}).get("selected", {})) if previous_result and previous_result.get("recordSignature") == record_signature else "",
        },
        "repairSummary": {
            "strictCandidateCount": len(strict_candidates),
            "repairCandidateCount": len(repair_candidates),
            "geneticCandidateCount": len(genetic_candidates),
            "selectedRelaxations": best.get("relaxations", []),
        },
        "solver": {
            "algorithm": "metaheuristic-genetic",
            "objective": "미배정과 hard 검증 오류를 최우선으로 줄이고, 교사별 요일/오전오후 균등성을 최적화합니다.",
            "options": {key: value for key, value in solve_options.items() if key != "apiKey"},
        },
    }
    append_operation_log("solve", {
        "runId": run_id,
        "seed": seed,
        "elapsedMs": elapsed_ms,
        "timedOut": False,
        "attemptCount": result["attemptCount"],
        "unassigned": len(best.get("unassigned", [])),
        "errors": ai_summary.get("errorCount", 0),
        "provider": config.get("provider"),
        "remoteStatus": result.get("aiAdvisor", {}).get("remote", {}).get("status"),
    })
    if persist:
        save_last_schedule(result)
    return result


def solve_result_rank(result: dict | None):
    if not result:
        return (-10**9, -10**9, -10**9, -10**9)
    return candidate_rank((result or {}).get("selected", {}))


def solve_best_summary(result: dict | None) -> dict:
    selected = (result or {}).get("selected", {})
    validation = selected.get("validation", {})
    violations = validation.get("violations", [])
    teacher_issues = selected.get("teacherIssues", [])
    return {
        "unassigned": len(selected.get("unassigned", [])),
        "errors": len([item for item in violations if item.get("severity") == "error"]),
        "lunchShortage": len([item for item in violations if item.get("type") == "lunch_protection"]),
        "consecutive": len([item for item in violations if item.get("type") == "max_consecutive"]),
        "imbalance": len([item for item in teacher_issues if any("안배" in as_text(tag) for tag in item.get("issues", []))]),
        "score": selected.get("score", 0),
        "signature": candidate_signature(selected) if selected else "",
    }


def solve_session_key(session_id: str) -> str:
    return f"solve_session:{session_id}"


def save_solve_session(session: dict) -> None:
    session_id = session.get("id")
    if not session_id:
        return
    if postgres_url():
        save_state_postgres(solve_session_key(session_id), session)
        return
    redis_url, redis_token = redis_config()
    if redis_url and redis_token:
        save_state_redis(solve_session_key(session_id), session)
        return
    ensure_dirs()
    (SOLVE_SESSION_DIR / f"{session_id}.json").write_text(json.dumps(session, ensure_ascii=False, indent=2), encoding="utf-8")


def load_solve_session(session_id: str) -> dict | None:
    if not session_id:
        return None
    if postgres_url():
        return load_state_postgres(solve_session_key(session_id))
    redis_url, redis_token = redis_config()
    if redis_url and redis_token:
        return load_state_redis(solve_session_key(session_id))
    path = SOLVE_SESSION_DIR / f"{session_id}.json"
    if not path.exists():
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def solve_session_chunk_options(solve_options: dict | None, chunk_index: int) -> dict:
    options = deepcopy(solve_options or {})
    strength = normalize_search_strength(options.get("searchStrength"))
    chunk_iterations = {"fast": 24, "balanced": 36, "strong": 48}[strength]
    requested = parse_positive_int(options.get("iterations")) or chunk_iterations
    options["iterations"] = max(18, min(requested, chunk_iterations))
    options["searchStrength"] = strength
    options["variationMode"] = "random"
    options["seed"] = solve_run_seed(None) + chunk_index * 1009
    return options


def run_solve_session_chunk(records: dict, session: dict) -> tuple[dict, bool]:
    chunk_index = parse_positive_int(session.get("chunkCount")) or 0
    chunk_options = solve_session_chunk_options(session.get("solveOptions") or {}, chunk_index)
    result = solve_schedule(records, ai_config={}, solve_options=chunk_options, persist=False, advisor=False)
    previous_best = session.get("bestResult")
    previous_rank = solve_result_rank(previous_best)
    new_rank = solve_result_rank(result)
    new_signature = result.get("bestSignature") or candidate_signature(result.get("selected", {}))
    previous_signature = (previous_best or {}).get("bestSignature") or candidate_signature((previous_best or {}).get("selected", {}))
    best_changed = False
    if not previous_best or new_rank > previous_rank or (new_rank == previous_rank and new_signature != previous_signature):
        session["bestResult"] = result
        best_changed = True
    session["chunkCount"] = chunk_index + 1
    session["attemptCount"] = (parse_positive_int(session.get("attemptCount")) or 0) + (parse_positive_int(result.get("attemptCount")) or 0)
    session["lastResultSummary"] = solve_best_summary(result)
    session["bestSummary"] = solve_best_summary(session.get("bestResult"))
    session["lastUpdatedAt"] = now_iso()
    session["elapsedMs"] = int((time.time() - float(session.get("startedAtEpoch", time.time()))) * 1000)
    return result, best_changed


def solve_session_response(session: dict, best_changed: bool = False) -> dict:
    return {
        "ok": True,
        "sessionId": session.get("id"),
        "startedAt": session.get("startedAt"),
        "elapsedMs": session.get("elapsedMs", 0),
        "chunkCount": session.get("chunkCount", 0),
        "attemptCount": session.get("attemptCount", 0),
        "bestChanged": best_changed,
        "bestSummary": session.get("bestSummary") or solve_best_summary(session.get("bestResult")),
        "lastResultSummary": session.get("lastResultSummary") or {},
        "canAccept": bool(session.get("bestResult")),
        "progressMessage": "탐색을 계속 진행 중입니다. 20초 이후 현재 최선안을 사용할 수 있습니다.",
    }


def start_solve_session(records: dict, solve_options: dict | None = None) -> dict:
    session = {
        "id": uuid.uuid4().hex[:12],
        "startedAt": now_iso(),
        "startedAtEpoch": time.time(),
        "solveOptions": {key: value for key, value in (solve_options or {}).items() if key != "apiKey"},
        "chunkCount": 0,
        "attemptCount": 0,
        "bestResult": None,
    }
    _, best_changed = run_solve_session_chunk(records, session)
    save_solve_session(session)
    append_operation_log("solve_start", {
        "sessionId": session["id"],
        "attemptCount": session.get("attemptCount"),
        "bestSummary": session.get("bestSummary"),
    })
    return solve_session_response(session, best_changed=best_changed)


def continue_solve_session(records: dict, session_id: str) -> dict:
    session = load_solve_session(session_id)
    if not session:
        return {"ok": False, "error": "진행 중인 자동배정 탐색 세션을 찾을 수 없습니다."}
    _, best_changed = run_solve_session_chunk(records, session)
    save_solve_session(session)
    append_operation_log("solve_continue", {
        "sessionId": session_id,
        "chunkCount": session.get("chunkCount"),
        "attemptCount": session.get("attemptCount"),
        "bestSummary": session.get("bestSummary"),
    })
    return solve_session_response(session, best_changed=best_changed)


def accept_solve_session(session_id: str) -> dict:
    session = load_solve_session(session_id)
    if not session or not session.get("bestResult"):
        return {"ok": False, "error": "현재 최선안을 찾을 수 없습니다. 자동배정을 먼저 시작하세요."}
    result = deepcopy(session["bestResult"])
    result["solveSession"] = {
        "sessionId": session_id,
        "chunkCount": session.get("chunkCount", 0),
        "attemptCount": session.get("attemptCount", 0),
        "elapsedMs": session.get("elapsedMs", 0),
        "acceptedBestSummary": session.get("bestSummary") or solve_best_summary(result),
    }
    result["progressMessage"] = "현재까지의 최선안을 반영했습니다."
    save_last_schedule(result)
    append_operation_log("solve_accept", {
        "sessionId": session_id,
        "chunkCount": session.get("chunkCount"),
        "attemptCount": session.get("attemptCount"),
        "bestSummary": result["solveSession"]["acceptedBestSummary"],
    })
    return result


def validate_schedule(records: dict, schedule: dict, unassigned=None) -> dict:
    unassigned = unassigned or []
    violations = []
    teacher_slots = defaultdict(list)
    room_slots = defaultdict(list)
    sync_slots = defaultdict(list)
    counts = Counter()
    settings = constraint_settings(records)
    forbidden = build_forbidden_index(records)

    for class_code, class_data in schedule.get("classes", {}).items():
        for day, periods in class_data.get("grid", {}).items():
            for period_text, cell in periods.items():
                if not cell:
                    continue
                period = parse_int(period_text)
                if not class_period_available(schedule, class_code, day, period):
                    violations.append({
                        "severity": "error",
                        "type": "class_day_limit",
                        "message": f"{display_name(records, '학급', class_code)} {day} {period}교시는 학급 요일별시수 범위를 벗어났습니다.",
                    })
                    continue
                if cell.get("source") == "fixed":
                    if cell.get("teacherCode"):
                        teacher_slots[(cell["teacherCode"], day, period)].append(class_code)
                    continue
                teacher_slots[(cell["teacherCode"], day, period)].append(class_code)
                if cell.get("roomCode"):
                    room_slots[(cell["roomCode"], day, period)].append(class_code)
                if cell.get("syncOccurrenceId"):
                    sync_slots[cell["syncOccurrenceId"]].append({
                        "classCode": class_code,
                        "day": day,
                        "period": period,
                        "syncGroup": cell.get("syncGroup", ""),
                        "syncLaneKey": cell.get("syncLaneKey", ""),
                    })
                counts[(cell["teacherCode"], cell["subjectCode"], class_code)] += 1
                if hard_forbidden(forbidden, cell, day, period):
                    violations.append({
                        "severity": "error",
                        "type": "forbidden",
                        "message": f"{display_name(records, '학급', class_code)} {day} {period}교시에 배정금지 조건을 위반했습니다.",
                    })

    for (teacher, day, period), classes in teacher_slots.items():
        if len(classes) > 1:
            violations.append({
                "severity": "error",
                "type": "teacher_conflict",
                "message": f"교사 {display_name(records, '교사', teacher)}가 {day} {period}교시에 {', '.join(display_names(records, '학급', classes))}에 중복 배정되었습니다.",
            })
    teacher_day_periods = defaultdict(set)
    for teacher, day, period in teacher_slots:
        teacher_day_periods[(teacher, day)].add(period)
    for (teacher, day), periods in teacher_day_periods.items():
        max_consecutive = settings.get("maxConsecutive") or 0
        if max_consecutive and max_consecutive_count(periods) > max_consecutive:
            violations.append({
                "severity": "error",
                "type": "max_consecutive",
                "message": f"교사 {display_name(records, '교사', teacher)}가 {day}요일에 최대연강 {max_consecutive}교시를 초과했습니다.",
            })
        lunch_after = settings.get("lunchAfter") or 0
        if settings.get("protectLunch") and lunch_after and {lunch_after, lunch_after + 1}.issubset(periods):
            violations.append({
                "severity": "error",
                "type": "lunch_protection",
                "message": f"교사 {display_name(records, '교사', teacher)}가 {day}요일 점심 전후({lunch_after},{lunch_after + 1}교시)에 모두 배정되었습니다.",
            })
    for (room, day, period), classes in room_slots.items():
        if len(classes) > 1:
            violations.append({
                "severity": "error",
                "type": "room_conflict",
                "message": f"특별실 {display_name(records, '특별실', room)}이 {day} {period}교시에 {', '.join(display_names(records, '학급', classes))}에 중복 배정되었습니다.",
            })

    expected_sync_sizes = records.get("_syncOccurrenceSize") or {}
    for occurrence_id, slots in sync_slots.items():
        positions = {(item["day"], item["period"]) for item in slots}
        if len(positions) > 1:
            violations.append({
                "severity": "error",
                "type": "sync_group_split",
                "message": f"동시그룹 occurrence {occurrence_id}가 서로 다른 시간에 나뉘어 배정되었습니다.",
            })
        expected_size = parse_positive_int(expected_sync_sizes.get(occurrence_id))
        if expected_size and len(slots) != expected_size:
            violations.append({
                "severity": "error",
                "type": "sync_group_incomplete",
                "message": f"동시그룹 occurrence {occurrence_id}는 {expected_size}개 lane이 동시에 있어야 하지만 {len(slots)}개만 배정되었습니다.",
            })

    expected_counts = Counter()
    load_display = {}
    for load in records.get("loads", []):
        key = (load["teacherCode"], load["subjectCode"], load["classCode"])
        expected_counts[key] += load["weeklyHours"]
        load_display[key] = load
    for key, expected in expected_counts.items():
        teacher_code, subject_code, class_code = key
        actual = counts[key]
        if actual != expected:
            severity = "error" if actual < expected else "warning"
            load = load_display[key]
            violations.append({
                "severity": severity,
                "type": "load_mismatch",
                "message": f"{display_name(records, '교사', teacher_code)} / {display_name(records, '과목', subject_code)} / {display_name(records, '학급', class_code)} 시수 {expected} 중 {actual}시간 배정되었습니다.",
            })

    for item in unassigned:
        violations.append({
            "severity": "error",
            "type": "unassigned",
            "message": f"{describe_unassigned_item(records, item)} 미배정: {item['reason']}",
        })

    return {
        "ok": not any(item["severity"] == "error" for item in violations),
        "violations": violations,
    }


def schedule_cell(schedule: dict, class_code: str, day: str, period) -> dict | None:
    return (schedule.get("classes", {}).get(class_code, {}).get("grid", {}).get(day, {}) or {}).get(str(period))


def sync_occurrence_cells(schedule: dict, occurrence_id: str) -> list[dict]:
    cells = []
    if not occurrence_id:
        return cells
    for class_code, class_data in schedule.get("classes", {}).items():
        for day in schedule.get("days", []):
            for period in schedule.get("periods", []):
                cell = class_data.get("grid", {}).get(day, {}).get(str(period))
                if cell and cell.get("syncOccurrenceId") == occurrence_id:
                    cells.append({
                        "classCode": class_code,
                        "day": day,
                        "period": period,
                        "cell": cell,
                    })
    return cells


def cell_belongs_to_occurrence(cell: dict | None, occurrence_id: str) -> bool:
    return bool(cell and cell.get("syncOccurrenceId") == occurrence_id)


def can_place_sync_cells_at(schedule: dict, cells: list[dict], day: str, period: int, allowed_occurrences: set[str]) -> tuple[bool, str]:
    for item in cells:
        class_code = item["classCode"]
        if not class_period_available(schedule, class_code, day, period):
            return False, f"{class_code}의 {day} {period}교시가 요일별시수 범위를 벗어납니다."
        target = schedule_cell(schedule, class_code, day, period)
        if target and target.get("syncOccurrenceId") not in allowed_occurrences:
            return False, f"{class_code}의 {day} {period}교시에 다른 수업이 있어 동시그룹 전체 이동이 어렵습니다."
    return True, ""


def finish_move_result(records: dict, schedule: dict, action: str) -> dict:
    validation = validate_schedule(records, schedule)
    diagnostics = diagnose_schedule(records, schedule, validation)
    return {
        "ok": validation["ok"],
        "message": f"{action}을 적용했습니다." if validation["ok"] else f"{action} 후 검증 오류가 있습니다.",
        "schedule": schedule,
        "validation": validation,
        "diagnostics": diagnostics,
        "teacherIssues": teacher_issue_summary(records, schedule, validation),
    }


def move_sync_occurrence(records: dict, updated: dict, move: dict, source_cell: dict) -> dict:
    mode = as_text(move.get("mode")) or "auto"
    src = move.get("from", {})
    dst = move.get("to", {})
    src_day = src.get("day")
    src_period = parse_int(src.get("period"))
    dst_day = dst.get("day")
    dst_period = parse_int(dst.get("period"))
    if dst_period is None:
        return {"ok": False, "message": "이동할 교시가 올바르지 않습니다.", "schedule": updated}
    source_occurrence_id = source_cell.get("syncOccurrenceId")
    source_cells = sync_occurrence_cells(updated, source_occurrence_id)
    if not source_cells:
        return {"ok": False, "message": "동시그룹 묶음을 찾을 수 없습니다.", "schedule": updated}
    class_code = src.get("classCode")
    target_cell = schedule_cell(updated, class_code, dst_day, dst_period)
    target_occurrence_id = target_cell.get("syncOccurrenceId") if target_cell else ""
    if mode == "move" and target_cell and not cell_belongs_to_occurrence(target_cell, source_occurrence_id):
        return {"ok": False, "message": "동시그룹 이동은 비어 있는 같은 시간대 칸으로만 이동할 수 있습니다.", "schedule": updated}
    if mode == "swap" and not target_occurrence_id:
        return {"ok": False, "message": "동시그룹 맞교환은 다른 동시그룹 묶음과만 가능합니다.", "schedule": updated}

    if target_occurrence_id and target_occurrence_id != source_occurrence_id:
        target_cells = sync_occurrence_cells(updated, target_occurrence_id)
        source_ok, source_message = can_place_sync_cells_at(updated, source_cells, dst_day, dst_period, {source_occurrence_id, target_occurrence_id})
        target_ok, target_message = can_place_sync_cells_at(updated, target_cells, src_day, src_period, {source_occurrence_id, target_occurrence_id})
        if not source_ok or not target_ok:
            return {"ok": False, "message": source_message or target_message, "schedule": updated}
        for item in source_cells + target_cells:
            updated["classes"][item["classCode"]]["grid"][item["day"]][str(item["period"])] = None
        for item in source_cells:
            updated["classes"][item["classCode"]]["grid"][dst_day][str(dst_period)] = item["cell"]
        for item in target_cells:
            updated["classes"][item["classCode"]]["grid"][src_day][str(src_period)] = item["cell"]
        return finish_move_result(records, updated, "동시그룹 맞교환")

    ok, message = can_place_sync_cells_at(updated, source_cells, dst_day, dst_period, {source_occurrence_id})
    if not ok:
        return {"ok": False, "message": message, "schedule": updated}
    for item in source_cells:
        updated["classes"][item["classCode"]]["grid"][item["day"]][str(item["period"])] = None
    for item in source_cells:
        updated["classes"][item["classCode"]]["grid"][dst_day][str(dst_period)] = item["cell"]
    return finish_move_result(records, updated, "동시그룹 이동")


def move_schedule(records: dict, schedule: dict, move: dict) -> dict:
    updated = deepcopy(schedule)
    mode = as_text(move.get("mode")) or "auto"
    src = move.get("from", {})
    dst = move.get("to", {})
    class_code = src.get("classCode")
    src_day = src.get("day")
    src_period = str(src.get("period"))
    dst_day = dst.get("day")
    dst_period = str(dst.get("period"))
    if class_code not in updated.get("classes", {}):
        return {"ok": False, "message": "이동할 학급을 찾을 수 없습니다.", "schedule": schedule}
    grid = updated["classes"][class_code]["grid"]
    cell = grid.get(src_day, {}).get(src_period)
    if not cell:
        return {"ok": False, "message": "이동할 배정이 없습니다.", "schedule": schedule}
    if cell.get("source") == "fixed":
        return {"ok": False, "message": "고정 일과는 엑셀의 고정 일과 시트에서 수정하세요.", "schedule": schedule}
    if cell.get("syncOccurrenceId"):
        return move_sync_occurrence(records, updated, move, cell)
    dst_period_int = parse_int(dst_period)
    if not class_period_available(updated, class_code, dst_day, dst_period_int):
        return {"ok": False, "message": "대상 교시는 해당 학급의 요일별시수 범위를 벗어납니다.", "schedule": schedule}
    dst_cell = grid.get(dst_day, {}).get(dst_period)
    if dst_cell and dst_cell.get("source") == "fixed":
        return {"ok": False, "message": "고정 일과 칸으로는 이동할 수 없습니다.", "schedule": schedule}
    if mode == "move" and dst_cell:
        return {"ok": False, "message": "빈칸 이동 모드에서는 비어 있는 칸으로만 이동할 수 있습니다.", "schedule": schedule}
    if mode == "swap" and not dst_cell:
        return {"ok": False, "message": "맞교환 모드에서는 대상 칸에 교환할 배정이 있어야 합니다.", "schedule": schedule}
    if dst_cell:
        grid[src_day][src_period], grid[dst_day][dst_period] = dst_cell, grid[src_day][src_period]
        action = "맞교환"
    else:
        grid[dst_day][dst_period] = cell
        grid[src_day][src_period] = None
        action = "이동"
    return finish_move_result(records, updated, action)


def violation_signature(violation: dict) -> tuple:
    return (
        violation.get("severity", ""),
        violation.get("type", ""),
        violation.get("message", ""),
    )


def hard_error_signatures(validation: dict) -> set[tuple]:
    return {
        violation_signature(item)
        for item in validation.get("violations", [])
        if item.get("severity") == "error"
    }


def move_option_reason(result: dict, delta: float, mode: str, new_errors: list[dict] | None = None) -> list[str]:
    reasons = ["맞교환" if mode == "swap" else "빈칸"]
    errors = new_errors if new_errors is not None else [item for item in result.get("validation", {}).get("violations", []) if item.get("severity") == "error"]
    if errors:
        reason_map = {
            "teacher_conflict": "교사중복",
            "room_conflict": "특별실중복",
            "max_consecutive": "연강초과",
            "lunch_protection": "식사부족",
            "load_mismatch": "시수불일치",
            "forbidden": "금지위반",
        }
        reasons.extend(reason_map.get(item.get("type"), "검증오류") for item in errors[:2])
        return reasons
    if delta > 2:
        reasons.append("안배개선")
    elif delta < -2:
        reasons.append("안배주의")
    else:
        reasons.append("유지")
    teacher_issues = result.get("teacherIssues", [])
    if teacher_issues:
        reasons.append(f"불량{len(teacher_issues)}")
    return reasons


def quick_move_options(records: dict, schedule: dict, source: dict) -> dict:
    class_code = source.get("classCode")
    src_day = source.get("day")
    src_period = str(source.get("period"))
    class_data = schedule.get("classes", {}).get(class_code)
    if not class_data:
        return {"ok": False, "message": "학급을 찾을 수 없습니다.", "options": []}
    source_cell = class_data.get("grid", {}).get(src_day, {}).get(src_period)
    if not source_cell:
        return {"ok": False, "message": "이동할 수업이 없습니다.", "options": []}
    if source_cell.get("source") == "fixed":
        return {"ok": False, "message": "고정 일과는 간편수정 대상이 아닙니다.", "options": []}

    base_metrics = teacher_distribution_metrics(schedule)
    base_validation = validate_schedule(records, schedule)
    base_error_signatures = hard_error_signatures(base_validation)
    options = []
    for day in schedule.get("days", []):
        for period in schedule.get("periods", []):
            if day == src_day and str(period) == src_period:
                continue
            if not class_period_available(schedule, class_code, day, period):
                continue
            target_cell = class_data.get("grid", {}).get(day, {}).get(str(period))
            if target_cell and target_cell.get("source") == "fixed":
                continue
            mode = "swap" if target_cell else "move"
            result = move_schedule(records, schedule, {
                "mode": mode,
                "from": {"classCode": class_code, "day": src_day, "period": parse_int(src_period)},
                "to": {"day": day, "period": period},
            })
            if "validation" not in result:
                options.append({
                    "day": day,
                    "period": period,
                    "mode": mode,
                    "grade": "bad",
                    "score": 0,
                    "reasons": [result.get("message", "불가")],
                    "target": target_cell,
                })
                continue
            candidate_errors = [item for item in result["validation"].get("violations", []) if item.get("severity") == "error"]
            new_errors = [item for item in candidate_errors if violation_signature(item) not in base_error_signatures]
            errors = len(new_errors)
            after_metrics = teacher_distribution_metrics(result["schedule"])
            delta = base_metrics.get("imbalance", 0) - after_metrics.get("imbalance", 0)
            if errors:
                grade = "bad"
                score = max(0, 35 - errors * 8)
            elif delta >= 2:
                grade = "good"
                score = 90 + min(9, int(delta))
            elif delta >= -2:
                grade = "ok"
                score = 72
            else:
                grade = "warn"
                score = max(40, 68 + int(delta))
            options.append({
                "day": day,
                "period": period,
                "mode": mode,
                "grade": grade,
                "score": score,
                "reasons": move_option_reason(result, delta, mode, new_errors),
                "target": target_cell,
                "errorCount": len(candidate_errors),
                "newErrorCount": errors,
            })
    return {
        "ok": True,
        "source": {"classCode": class_code, "day": src_day, "period": parse_int(src_period), "cell": source_cell},
        "options": sorted(options, key=lambda item: (-item["score"], item["day"], item["period"])),
        "teacherIssues": teacher_issue_summary(records, schedule, base_validation),
    }


def affected_teacher_codes_for_move(schedule: dict, move: dict, result_schedule: dict | None = None) -> list[str]:
    if result_schedule:
        codes = set()
        for source_schedule in [schedule, result_schedule]:
            for class_data in source_schedule.get("classes", {}).values():
                for day in source_schedule.get("days", []):
                    for period in source_schedule.get("periods", []):
                        cell = class_data.get("grid", {}).get(day, {}).get(str(period))
                        if cell and cell.get("teacherCode"):
                            codes.add(cell["teacherCode"])
        changed = []
        for teacher_code in codes:
            before = json.dumps(teacher_schedule_cells(schedule, teacher_code), ensure_ascii=False, sort_keys=True, default=str)
            after = json.dumps(teacher_schedule_cells(result_schedule, teacher_code), ensure_ascii=False, sort_keys=True, default=str)
            if before != after:
                changed.append(teacher_code)
        src = move.get("from", {})
        dst = move.get("to", {})
        class_code = src.get("classCode")
        for source_schedule, day, period in [
            (schedule, src.get("day"), str(src.get("period"))),
            (schedule, dst.get("day"), str(dst.get("period"))),
            (result_schedule, src.get("day"), str(src.get("period"))),
            (result_schedule, dst.get("day"), str(dst.get("period"))),
        ]:
            cell = schedule_cell(source_schedule, class_code, day, period)
            teacher_code = cell.get("teacherCode") if isinstance(cell, dict) else ""
            if teacher_code and teacher_code not in changed:
                changed.append(teacher_code)
        return changed
    codes = []
    src = move.get("from", {})
    dst = move.get("to", {})
    class_code = src.get("classCode")
    for source_schedule, day, period in [
        (schedule, src.get("day"), str(src.get("period"))),
        (schedule, dst.get("day"), str(dst.get("period"))),
        (result_schedule or {}, src.get("day"), str(src.get("period"))),
        (result_schedule or {}, dst.get("day"), str(dst.get("period"))),
    ]:
        cell = (source_schedule.get("classes", {}).get(class_code, {}).get("grid", {}).get(day, {}) or {}).get(period)
        teacher_code = cell.get("teacherCode") if isinstance(cell, dict) else ""
        if teacher_code and teacher_code not in codes:
            codes.append(teacher_code)
    return codes


def teacher_schedule_cells(schedule: dict, teacher_code: str) -> list[dict]:
    cells = []
    for class_code, class_data in schedule.get("classes", {}).items():
        for day in schedule.get("days", []):
            for period in schedule.get("periods", []):
                cell = class_data.get("grid", {}).get(day, {}).get(str(period))
                if not cell or cell.get("teacherCode") != teacher_code:
                    continue
                cells.append({
                    "classCode": class_code,
                    "className": class_data.get("name", class_code),
                    "day": day,
                    "period": period,
                    "subjectCode": cell.get("subjectCode", ""),
                    "subjectName": cell.get("subjectName", ""),
                    "roomCode": cell.get("roomCode", ""),
                    "label": f"{cell.get('subjectName') or cell.get('subjectCode', '')} {class_data.get('name', class_code)}".strip(),
                })
    return cells


def schedule_diff_cells(before_schedule: dict, after_schedule: dict, move: dict) -> list[dict]:
    class_code = move.get("from", {}).get("classCode")
    changed = []
    if not class_code:
        return changed
    for day in before_schedule.get("days", []):
        for period in before_schedule.get("periods", []):
            before = before_schedule.get("classes", {}).get(class_code, {}).get("grid", {}).get(day, {}).get(str(period))
            after = after_schedule.get("classes", {}).get(class_code, {}).get("grid", {}).get(day, {}).get(str(period))
            if json.dumps(before, ensure_ascii=False, sort_keys=True, default=str) != json.dumps(after, ensure_ascii=False, sort_keys=True, default=str):
                changed.append({"classCode": class_code, "day": day, "period": period, "before": before, "after": after})
    return changed


def move_preview(records: dict, schedule: dict, move: dict) -> dict:
    result = move_schedule(records, schedule, move)
    after_schedule = result.get("schedule", schedule)
    teacher_codes = affected_teacher_codes_for_move(schedule, move, after_schedule)
    affected = [
        {
            "teacherCode": teacher_code,
            "teacherName": display_name(records, "교사", teacher_code),
            "beforeCells": teacher_schedule_cells(schedule, teacher_code),
            "afterCells": teacher_schedule_cells(after_schedule, teacher_code),
        }
        for teacher_code in teacher_codes
    ]
    affected.sort(key=lambda item: item.get("teacherName", ""))
    return {
        "ok": bool(result.get("validation")) and result.get("ok", False),
        "message": result.get("message", ""),
        "move": move,
        "before": {"schedule": schedule},
        "after": {"schedule": after_schedule},
        "affectedTeachers": affected,
        "validation": result.get("validation", {}),
        "diagnostics": result.get("diagnostics", []),
        "teacherIssues": result.get("teacherIssues", []),
        "diffPreview": schedule_diff_cells(schedule, after_schedule, move),
    }


def apply_schedule_proposal(records: dict, proposal: dict) -> dict:
    schedule_result = deepcopy(proposal.get("scheduleResult") or proposal)
    selected = schedule_result.get("selected") or schedule_result
    schedule = selected.get("schedule")
    if not schedule:
        return {"ok": False, "error": "적용할 시간표가 없습니다."}
    validation = validate_schedule(records, schedule, selected.get("unassigned") or [])
    selected["validation"] = validation
    selected["diagnostics"] = diagnose_schedule(records, schedule, validation, selected.get("unassigned") or [])
    selected["teacherIssues"] = teacher_issue_summary(records, schedule, validation)
    schedule_result["selected"] = selected
    schedule_result.setdefault("candidates", [selected])
    schedule_result["bestStrategy"] = selected.get("strategy", schedule_result.get("bestStrategy", "ai-proposal"))
    schedule_result["createdAt"] = schedule_result.get("createdAt") or now_iso()
    save_last_schedule(schedule_result)
    append_operation_log("proposal_apply", {
        "strategy": selected.get("strategy"),
        "ok": validation.get("ok"),
        "errors": len([item for item in validation.get("violations", []) if item.get("severity") == "error"]),
    })
    return {"ok": validation.get("ok", False), "scheduleResult": schedule_result}


def save_moved_schedule_result(move_result: dict, body: dict) -> None:
    if "validation" not in move_result:
        return
    last = load_last_schedule() or {
        "createdAt": now_iso(),
        "bestStrategy": body.get("strategy", "manual"),
        "candidates": [],
        "selected": {},
    }
    if body.get("recordSignature"):
        last["recordSignature"] = body.get("recordSignature")
    selected = last.setdefault("selected", {})
    selected["schedule"] = move_result["schedule"]
    selected["validation"] = move_result.get("validation", {})
    selected["diagnostics"] = move_result.get("diagnostics", [])
    selected["teacherIssues"] = move_result.get("teacherIssues", [])
    selected["strategy"] = body.get("strategy") or selected.get("strategy", "manual")
    selected["effectiveConfig"] = body.get("effectiveConfig") or selected.get("effectiveConfig", {})
    selected["relaxations"] = body.get("relaxations") or selected.get("relaxations", [])
    selected["manualEdited"] = True
    selected["lastManualMove"] = body.get("move", {})
    for candidate in last.get("candidates", []):
        if candidate.get("strategy") == selected.get("strategy"):
            candidate.update({
                "schedule": selected["schedule"],
                "validation": selected["validation"],
                "diagnostics": selected["diagnostics"],
                "teacherIssues": selected["teacherIssues"],
                "manualEdited": True,
            })
            break
    save_last_schedule(last)


def mask_records_for_ai(records: dict) -> dict:
    return {
        "counts": summarize_records(records),
        "classes": list(records.get("classes", {}).keys()),
        "classDayLimits": {
            code: item.get("_dayLimits", {})
            for code, item in records.get("classes", {}).items()
        },
        "teachers": list(records.get("teachers", {}).keys()),
        "subjects": list(records.get("subjects", {}).keys()),
        "rooms": list(records.get("rooms", {}).keys()),
        "loads": [
            {
                "teacherCode": item["teacherCode"],
                "subjectCode": item["subjectCode"],
                "classCode": item["classCode"],
                "weeklyHours": item["weeklyHours"],
                "roomCode": item.get("roomCode", ""),
            }
            for item in records.get("loads", [])
        ],
        "fixedPeriods": [
            {
                "classCode": item.get("classCode"),
                "day": item.get("day"),
                "period": item.get("period"),
                "label": item.get("label"),
                "kind": item.get("kind"),
                "supervisorCode": item.get("supervisorCode", ""),
            }
            for item in records.get("fixedPeriods", [])
        ],
    }


def validate_openai_key(config: dict) -> dict:
    api_key = as_text(config.get("apiKey"))
    if not api_key:
        return {"ok": False, "status": "missing", "provider": "OpenAI", "message": "OpenAI API 키를 입력하세요."}
    request = Request(
        "https://api.openai.com/v1/models",
        headers={"Authorization": f"Bearer {api_key}"},
        method="GET",
    )
    try:
        with urlopen(request, timeout=8) as response:
            if 200 <= response.status < 300:
                return {"ok": True, "status": "verified", "provider": "OpenAI", "message": "OpenAI API 키 검증에 성공했습니다.", "aiConfig": public_ai_config({**config, "validated": True})}
            return {"ok": False, "status": "unexpected", "provider": "OpenAI", "message": f"OpenAI API 응답 상태가 예상과 다릅니다: {response.status}"}
    except HTTPError as error:
        if error.code in {401, 403}:
            return {"ok": False, "status": "auth_failed", "provider": "OpenAI", "message": "OpenAI API 키가 유효하지 않거나 권한이 없습니다."}
        return {"ok": False, "status": "http_error", "provider": "OpenAI", "message": f"OpenAI API 검증 중 HTTP 오류가 발생했습니다: {error.code}"}
    except (URLError, TimeoutError) as error:
        return {"ok": False, "status": "network_error", "provider": "OpenAI", "message": f"OpenAI API 검증 네트워크 오류: {error}"}


def validate_gemini_key(config: dict) -> dict:
    api_key = as_text(config.get("apiKey"))
    model = as_text(config.get("model")) or AI_DEFAULT_MODELS["gemini"]
    if not api_key:
        return {"ok": False, "status": "missing", "provider": "Gemini", "message": "Gemini API 키를 입력하세요."}
    request_body = {
        "contents": [{"parts": [{"text": "Return OK."}]}],
        "generationConfig": {"maxOutputTokens": 8},
    }
    request = Request(
        f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent",
        data=json.dumps(request_body, ensure_ascii=False).encode("utf-8"),
        headers={
            "x-goog-api-key": api_key,
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=8) as response:
            if 200 <= response.status < 300:
                return {"ok": True, "status": "verified", "provider": "Gemini", "message": "Gemini API 키 검증에 성공했습니다.", "aiConfig": public_ai_config({**config, "validated": True, "model": model})}
            return {"ok": False, "status": "unexpected", "provider": "Gemini", "message": f"Gemini API 응답 상태가 예상과 다릅니다: {response.status}"}
    except HTTPError as error:
        if error.code in {400, 401, 403, 404}:
            return {"ok": False, "status": "auth_failed", "provider": "Gemini", "message": f"Gemini API 키, 모델명, 권한을 확인하세요. HTTP {error.code}"}
        return {"ok": False, "status": "http_error", "provider": "Gemini", "message": f"Gemini API 검증 중 HTTP 오류가 발생했습니다: {error.code}"}
    except (URLError, TimeoutError) as error:
        return {"ok": False, "status": "network_error", "provider": "Gemini", "message": f"Gemini API 검증 네트워크 오류: {error}"}


def validate_custom_key(config: dict) -> dict:
    api_key = as_text(config.get("apiKey"))
    base_url = custom_base_url(config)
    model = as_text(config.get("model"))
    if not base_url:
        return {"ok": False, "status": "missing_base_url", "provider": "Custom", "message": "Custom 제공자는 Base URL이 필요합니다."}
    if not api_key:
        return {"ok": False, "status": "missing", "provider": "Custom", "message": "Custom API 키를 입력하세요."}
    if not model:
        return {"ok": False, "status": "missing_model", "provider": "Custom", "message": "Custom 제공자는 모델명이 필요합니다."}
    request = Request(
        f"{base_url}/models",
        headers={"Authorization": f"Bearer {api_key}"},
        method="GET",
    )
    try:
        with urlopen(request, timeout=8) as response:
            if 200 <= response.status < 300:
                return {"ok": True, "status": "verified", "provider": "Custom", "message": "Custom API 키 검증에 성공했습니다.", "aiConfig": public_ai_config({**config, "validated": True})}
            return {"ok": False, "status": "unexpected", "provider": "Custom", "message": f"Custom API 응답 상태가 예상과 다릅니다: {response.status}"}
    except HTTPError as error:
        if error.code in {401, 403}:
            return {"ok": False, "status": "auth_failed", "provider": "Custom", "message": "Custom API 키가 유효하지 않거나 권한이 없습니다."}
        return {"ok": False, "status": "http_error", "provider": "Custom", "message": f"Custom API 검증 중 HTTP 오류가 발생했습니다: {error.code}"}
    except (URLError, TimeoutError) as error:
        return {"ok": False, "status": "network_error", "provider": "Custom", "message": f"Custom API 검증 네트워크 오류: {error}"}


def validate_ai_key(api_key_or_config) -> dict:
    config = normalize_ai_config(api_key_or_config if isinstance(api_key_or_config, dict) else None, api_key=api_key_or_config if isinstance(api_key_or_config, str) else "")
    if config["provider"] == "gemini":
        return validate_gemini_key(config)
    if config["provider"] == "custom":
        return validate_custom_key(config)
    return validate_openai_key(config)


def describe_unassigned_item(records: dict, item: dict) -> str:
    teacher = item.get("teacherName") or display_name(records, "교사", item.get("teacherCode", ""))
    subject = item.get("subjectName") or display_name(records, "과목", item.get("subjectCode", ""))
    class_name = item.get("className") or display_name(records, "학급", item.get("classCode", ""))
    return f"{teacher} / {subject} / {class_name} {item.get('hours', 1)}시간"


def extract_constraint_drafts(records: dict, text: str) -> list[dict]:
    entity = find_entity_in_text(records, text)
    if not entity:
        return []
    target_type, target_code, target_name = entity
    days = [day for day in DEFAULT_DAYS if day in text]
    periods = [int(item) for item in re.findall(r"(\d+)\s*교시", text)]
    if not days and any(token in text for token in ["매일", "전체요일", "모든 요일"]):
        days = DEFAULT_DAYS[:]
    soft_hint = any(token in text for token in ["가능하면", "되도록", "가급적", "선호", "비선호"])
    hope_hint = any(token in text for token in ["희망", "원해", "넣어", "배정해", "좋아"])
    avoid_hint = any(token in text for token in ["피", "금지", "빼", "넣지", "하지마", "하지 말", "말아", "안되", "안돼", "불가", "싫"])
    if hope_hint and not avoid_hint:
        condition_type = "희망"
        strength = "soft"
    elif soft_hint and avoid_hint:
        condition_type = "비선호"
        strength = "soft"
    else:
        condition_type = "배정금지"
        strength = "hard"
    if not periods and not days:
        return []
    period_text = ",".join(str(item) for item in periods)
    title = f"{target_name} {condition_type}"
    when = " ".join(days + ([f"{period_text}교시"] if period_text else []))
    return [{
        "id": uuid.uuid4().hex[:8],
        "title": f"{title}: {when}".strip(),
        "targetType": target_type,
        "targetCode": target_code,
        "targetName": target_name,
        "conditionType": condition_type,
        "days": days,
        "periods": periods,
        "periodsText": period_text,
        "strength": strength,
        "priority": 5 if strength == "soft" else 10,
        "description": text,
    }]


def extract_constraint_drafts_v2(records: dict, text: str) -> list[dict]:
    drafts = []
    sentence_candidates = [part.strip() for part in re.split(r"[\n.;。]+", as_text(text)) if part.strip()]
    for sentence in sentence_candidates or [as_text(text)]:
        entities = find_entities_in_text(records, sentence)
        days = [day for day in DEFAULT_DAYS if day in sentence]
        periods = [int(item) for item in re.findall(r"(\d+)\s*교시", sentence)]
        if not days and any(token in sentence for token in ["매일", "전체요일", "모든 요일"]):
            days = DEFAULT_DAYS[:]
        soft_hint = any(token in sentence for token in ["가능하면", "하도록", "가급적", "선호", "비선호"])
        hope_hint = any(token in sentence for token in ["희망", "원해", "넣어", "배정해", "좋아"])
        avoid_hint = any(token in sentence for token in ["안", "금지", "비", "넣지", "하지마", "하지 마", "말아", "안돼", "안되", "불가", "X"])
        if hope_hint and not avoid_hint:
            condition_type = "희망"
            strength = "soft"
        elif soft_hint and avoid_hint:
            condition_type = "비선호"
            strength = "soft"
        else:
            condition_type = "배정금지"
            strength = "hard"
        period_text = ",".join(str(item) for item in periods)
        if entities and (periods or days):
            for target_type, target_code, target_name in entities:
                when = " ".join(days + ([f"{period_text}교시"] if period_text else []))
                drafts.append({
                    "id": uuid.uuid4().hex[:8],
                    "title": f"{target_name} {condition_type}: {when}".strip(),
                    "targetType": target_type,
                    "targetCode": target_code,
                    "targetName": target_name,
                    "conditionType": condition_type,
                    "days": days,
                    "periods": periods,
                    "periodsText": period_text,
                    "strength": strength,
                    "priority": 5 if strength == "soft" else 10,
                    "description": sentence,
                    "engineSupported": True,
                })
        elif sentence:
            drafts.append({
                "id": uuid.uuid4().hex[:8],
                "title": "메모형 제약: 엔진 미반영",
                "targetType": "",
                "targetCode": "",
                "targetName": "",
                "conditionType": "메모",
                "days": days,
                "periods": periods,
                "periodsText": period_text,
                "strength": "soft",
                "priority": 1,
                "description": sentence,
                "engineSupported": False,
                "unsupportedReason": "대상 또는 요일/교시를 구조화하지 못해 자동배정 엔진에는 아직 반영되지 않습니다.",
            })
    unique = {}
    for draft in drafts:
        key = (
            draft.get("targetType"),
            draft.get("targetCode"),
            draft.get("conditionType"),
            tuple(draft.get("days") or []),
            draft.get("periodsText"),
            draft.get("description"),
            draft.get("engineSupported"),
        )
        unique.setdefault(key, draft)
    return list(unique.values())[:12]


def unassigned_advice_steps(records: dict, schedule_context: dict, unassigned: list[dict]) -> list[str]:
    steps = []
    for item in unassigned[:5]:
        steps.append(f"{describe_unassigned_item(records, item)}: 배정금지·연강·점심·특별실 조건을 낮은 우선순위부터 완화해 후보를 다시 비교하세요.")
    diagnostics = schedule_context.get("diagnostics") or []
    for item in diagnostics:
        if item.get("type") == "unassigned" and item.get("reason"):
            steps.append(item["reason"])
    if not steps:
        steps.append("현재 선택된 시간표의 미배정 정보가 없으면 자동배정을 먼저 실행한 뒤 다시 질문하세요.")
    steps.append("미배정 우선 탐색을 다시 실행하고, 남은 항목은 시간표 보기에서 해당 학급 수업 칸을 선택해 간편수정 후보를 확인하세요.")
    return steps[:7]


def chat_requests_repair(text: str) -> bool:
    value = as_text(text)
    if "미배정" not in value:
        return False
    return any(token in value for token in ["없애", "처리", "줄여", "해결", "배정", "수정", "모두"])


def repair_chat_solve_options(records: dict, solve_options: dict | None = None) -> dict:
    settings = constraint_settings(records)
    options = deepcopy(solve_options or {})
    options["assignmentMethod"] = "unassigned-only"
    options["allowRelaxForUnassigned"] = "Y"
    options["iterations"] = max(parse_positive_int(options.get("iterations")) or settings.get("metaIterations") or 60, 160)
    options.setdefault("balanceStrength", settings.get("balanceStrength") or "soft")
    options.setdefault("protectLunch", "Y" if settings.get("protectLunch") else "N")
    if settings.get("maxConsecutive"):
        options.setdefault("maxConsecutive", settings["maxConsecutive"])
    return options


def repair_schedule_by_chat(records: dict, text: str, solve_options: dict | None = None, unassigned=None, ai_config=None) -> tuple[dict | None, dict | None]:
    if not chat_requests_repair(text) or not records.get("loads"):
        return None, None
    before = len(unassigned or [])
    options = repair_chat_solve_options(records, solve_options)
    result = solve_schedule(records, ai_config=ai_config, solve_options=options, persist=False)
    after = len(result.get("selected", {}).get("unassigned", []))
    action = {
        "type": "repair-solve",
        "applied": False,
        "requiresApproval": True,
        "beforeUnassigned": before,
        "afterUnassigned": after,
        "message": f"미배정 우선 재탐색을 실행했습니다. 미배정 {before}건 → {after}건",
        "solveOptions": {key: value for key, value in options.items() if key != "apiKey"},
    }
    return action, result


def mask_constraint_drafts_for_ai(drafts: list[dict]) -> list[dict]:
    masked = []
    for item in drafts:
        copied = {key: value for key, value in item.items() if key not in {"targetName", "description", "title"}}
        copied["targetName"] = copied.get("targetCode", "")
        copied["description"] = "사용자 대화에서 추출한 코드화 제약조건"
        copied["title"] = "대화 제약조건 초안"
        masked.append(copied)
    return masked


def mask_text_for_ai(records: dict, text: str) -> str:
    masked = as_text(text)
    replacements = []
    for collection in ENTITY_PREFIXES:
        for code, item in records.get(collection, {}).items():
            name = item.get(ENTITY_NAME_FIELDS.get(collection, "")) or item.get("displayName")
            if name:
                replacements.append((len(name), name, code))
    for _, name, code in sorted(replacements, reverse=True):
        masked = masked.replace(name, code)
    return masked


def mask_object_for_ai(records: dict, value):
    if isinstance(value, str):
        return mask_text_for_ai(records, value)
    if isinstance(value, list):
        return [mask_object_for_ai(records, item) for item in value]
    if isinstance(value, dict):
        return {key: mask_object_for_ai(records, item) for key, item in value.items()}
    return value


def mask_suggestions_for_ai(suggestions: list[dict]) -> list[dict]:
    masked = []
    for item in suggestions:
        copied = deepcopy(item)
        if copied.get("type") == "constraint_draft":
            copied["title"] = "대화 제약조건 초안"
            copied["explanation"] = "사용자 대화에서 요일/교시/대상 코드 기반 제약조건 초안을 추출했습니다."
            draft = copied.get("draft")
            if isinstance(draft, dict):
                copied["draft"] = mask_constraint_drafts_for_ai([draft])[0]
        masked.append(copied)
    return masked


def extract_schedule_context(records: dict, schedule: dict | None, unassigned=None) -> dict:
    unassigned = unassigned or []
    if not schedule:
        return {"hasSchedule": False, "validation": None, "diagnostics": []}
    validation = validate_schedule(records, schedule, unassigned)
    diagnostics = diagnose_schedule(records, schedule, validation, unassigned)
    return {
        "hasSchedule": True,
        "validation": validation,
        "diagnostics": diagnostics,
        "unassigned": unassigned,
    }


def ai_chat(records: dict, message: str, api_key_present: bool = False, schedule: dict | None = None, api_key: str = "", ai_config=None, unassigned=None, solve_options: dict | None = None) -> dict:
    config = normalize_ai_config(ai_config, api_key=api_key)
    masked = mask_records_for_ai(records)
    unassigned = unassigned or []
    schedule_context = extract_schedule_context(records, schedule, unassigned)
    suggestions = []
    text = message.strip()
    day_hits = [day for day in DEFAULT_DAYS if day in text]
    period_hits = re.findall(r"(\d+)\s*교시", text)
    settings = constraint_settings(records)
    constraint_drafts = extract_constraint_drafts_v2(records, text)
    schedule_action, schedule_result = repair_schedule_by_chat(records, text, solve_options=solve_options, unassigned=unassigned, ai_config=config)
    if schedule_action:
        suggestions.append({
            "type": "schedule_action",
            "title": "시간표 수정안",
            "explanation": schedule_action["message"],
            "steps": [schedule_action["message"], "변경안은 승인 후 적용됩니다."],
        })

    if constraint_drafts:
        draft = constraint_drafts[0]
        suggestions.append({
            "type": "constraint_draft",
            "title": "대화 제약조건 초안",
            "draft": draft,
            "explanation": f"{draft['targetName']}에 대한 {draft['conditionType']} 조건을 만들었습니다. 적용하면 다음 자동배정부터 반영됩니다.",
        })
    elif "배정금지" in text or "넣지" in text or "피" in text:
        suggestions.append({
            "type": "constraint_draft",
            "title": "대화 제약조건 입력 도움",
            "draft": {
                "conditionType": "배정금지",
                "days": day_hits,
                "periods": [int(item) for item in period_hits],
                "strength": "hard",
            },
            "explanation": "대상 교사명·학급명·과목명·특별실명을 함께 말하면 바로 적용 가능한 제약조건 초안을 만들 수 있습니다.",
        })
    if "연강" in text:
        suggestions.append({
            "type": "setting",
            "title": "최대연강 조건",
            "draft": {"setting": "최대연강허용", "current": settings.get("maxConsecutive")},
            "explanation": f"현재 최대연강허용은 {settings.get('maxConsecutive') or '제한 없음'}입니다. 기본설정 시트에서 숫자를 조정하면 hard 조건으로 반영됩니다.",
        })
    if "점심" in text:
        suggestions.append({
            "type": "setting",
            "title": "점심시간 보호",
            "draft": {"setting": "점심시간보호", "current": "Y" if settings.get("protectLunch") else "N", "lunchAfter": settings.get("lunchAfter")},
            "explanation": "점심시간보호가 Y이면 점심 전후 교시에 같은 교사가 모두 배정되는 후보를 제외합니다.",
        })
    if "미배정" in text or "왜" in text:
        suggestions.append({
            "type": "diagnostic",
            "title": "미배정/오류 진단",
            "steps": unassigned_advice_steps(records, schedule_context, unassigned),
        })
    if schedule_context.get("hasSchedule"):
        validation = schedule_context["validation"]
        error_count = len([item for item in validation.get("violations", []) if item.get("severity") == "error"])
        suggestions.append({
            "type": "schedule_review",
            "title": "현재 시간표 검토",
            "steps": [
                f"hard 검증 오류 {error_count}건",
                f"최대연강 {settings.get('maxConsecutive') or '제한 없음'}",
                f"점심보호 {'적용' if settings.get('protectLunch') else '미적용'}",
            ],
        })
    if not suggestions:
        suggestions.append({
            "type": "advisor",
            "title": "시간표 개선 제안",
            "steps": [
                "조건 우선순위가 높은 hard 조건부터 유지합니다.",
                "미배정이 생기면 최대연강, 점심보호, 특별실 조건 순서로 완화 후보를 비교합니다.",
                "교사별 보기에서 오전/오후 쏠림을 확인한 뒤 수동 이동으로 후보를 다듬습니다.",
            ],
        })

    local_advice = {
        "summary": "보조 진단입니다." if api_key_present else "로컬 규칙 기반 제안입니다.",
        "suggestions": [
            {
                "type": item.get("type", "advisor"),
                "title": item.get("title", "제안"),
                "explanation": item.get("explanation") or " → ".join(item.get("steps", [])) or json.dumps(item.get("draft", {}), ensure_ascii=False),
                "steps": item.get("steps", []) or ([item.get("explanation")] if item.get("explanation") else []),
            }
            for item in suggestions
        ],
    }
    provider_label = config.get("providerLabel", "AI")
    remote = {"ok": False, "status": "not-configured", "provider": provider_label, "message": f"{provider_label} API 키가 없어 로컬 진단만 사용했습니다."}
    if as_text(config.get("apiKey")):
        remote = call_ai_advisor(
            config,
            "chat_advice",
            {
                "userMessage": mask_text_for_ai(records, text),
                "maskedRecords": masked,
                "scheduleContext": mask_object_for_ai(records, schedule_context),
                "localSuggestions": mask_object_for_ai(records, mask_suggestions_for_ai(suggestions)),
                "settings": settings,
                "constraintDrafts": mask_constraint_drafts_for_ai(constraint_drafts),
                "unassigned": [
                    {
                        "teacherCode": item.get("teacherCode"),
                        "subjectCode": item.get("subjectCode"),
                        "classCode": item.get("classCode"),
                        "hours": item.get("hours"),
                        "reason": item.get("reason"),
                    }
                    for item in unassigned
                ],
            },
        )
    remote_failure = remote if as_text(config.get("apiKey")) and not remote.get("ok") else None
    active_advice = remote.get("advice") if remote.get("ok") else local_advice

    return {
        "mode": f"{config.get('provider')}-advisor" if remote.get("ok") else "masked-local-advisor",
        "externalApi": "called" if remote.get("ok") else ("fallback-local-analysis" if api_key_present else "not-configured"),
        "privacy": "AI 제안에는 실제 교사명/학급명 대신 코드와 집계, 시간표 상태만 사용하도록 마스킹했습니다.",
        "aiConfig": public_ai_config(config),
        "maskedPayload": masked,
        "scheduleContext": schedule_context,
        "constraintDrafts": constraint_drafts,
        "scheduleAction": schedule_action,
        "scheduleProposal": {
            "requiresApproval": True,
            "scheduleResult": schedule_result,
            "message": schedule_action.get("message") if schedule_action else "",
        } if schedule_result else None,
        "remote": remote,
        "remoteFailure": remote_failure,
        "localAdvice": local_advice,
        "advice": active_advice,
        "suggestions": active_advice.get("suggestions", suggestions),
    }


def export_schedule_workbook(result: dict) -> Workbook:
    wb = Workbook()
    selected = result.get("selected", result)
    schedule = selected.get("schedule", {})
    summary = wb.active
    summary.title = "요약"
    summary.append(["항목", "값"])
    summary.append(["생성시각", result.get("createdAt", now_iso())])
    summary.append(["선택전략", result.get("bestStrategy", selected.get("strategy", ""))])
    summary.append(["점수", selected.get("score", "")])
    summary.append(["미배정", len(selected.get("unassigned", []))])
    summary.append(["AI개선완화", " / ".join(selected.get("relaxations", []))])
    for class_code, class_data in schedule.get("classes", {}).items():
        ws = wb.create_sheet(class_data.get("name", class_code)[:31])
        ws.append(["요일/교시"] + [str(p) for p in schedule.get("periods", [])])
        for day in schedule.get("days", []):
            row = [day]
            day_limit = parse_int(class_data.get("dayLimits", {}).get(day), len(schedule.get("periods", [])))
            for period in schedule.get("periods", []):
                if period > day_limit:
                    row.append("수업 없음")
                    continue
                cell = class_data["grid"][day][str(period)]
                if not cell:
                    row.append("")
                elif cell.get("source") == "fixed":
                    parts = [cell["subjectName"]]
                    if cell.get("teacherName"):
                        parts.append(cell["teacherName"])
                    row.append("\n".join(parts))
                else:
                    row.append(f"{cell['subjectName']}\n{cell['teacherName']}" + (f"\n{cell['roomName']}" if cell.get("roomName") else ""))
            ws.append(row)
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="173F3F")
            cell.font = Font(color="FFFFFF", bold=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions["A"].width = 12
        for index in range(2, len(schedule.get("periods", [])) + 2):
            ws.column_dimensions[get_column_letter(index)].width = 18
    return wb


def export_neis_csv(result: dict) -> bytes:
    selected = result.get("selected", result)
    schedule = selected.get("schedule", {})
    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(["학급코드", "학급명", "요일", "교시", "과목코드", "과목명", "교사코드", "교사명"])
    for class_code, class_data in schedule.get("classes", {}).items():
        for day in schedule.get("days", []):
            day_limit = parse_int(class_data.get("dayLimits", {}).get(day), len(schedule.get("periods", [])))
            for period in schedule.get("periods", []):
                if period > day_limit:
                    continue
                cell = class_data["grid"][day][str(period)]
                if cell:
                    subject_code = "" if cell.get("source") == "fixed" else cell.get("subjectCode", "")
                    writer.writerow([class_code, class_data.get("name", class_code), day, period, subject_code, cell.get("subjectName", ""), cell.get("teacherCode", ""), cell.get("teacherName", "")])
    return ("\ufeff" + stream.getvalue()).encode("utf-8")


def save_last_schedule(result: dict) -> None:
    if postgres_url():
        save_state_postgres("last_schedule", result)
        redis_url, redis_token = redis_config()
        if redis_url and redis_token:
            save_state_redis("last_schedule", result)
        return
    redis_url, redis_token = redis_config()
    if redis_url and redis_token:
        save_state_redis("last_schedule", result)
        return
    ensure_dirs()
    LAST_SCHEDULE_FILE.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")


def load_last_schedule() -> dict | None:
    if postgres_url():
        result = load_state_postgres("last_schedule")
        if result is not None:
            return result
        redis_url, redis_token = redis_config()
        if redis_url and redis_token:
            return load_state_redis("last_schedule")
        return None
    redis_url, redis_token = redis_config()
    if redis_url and redis_token:
        return load_state_redis("last_schedule")
    if LAST_SCHEDULE_FILE.exists():
        return json.loads(LAST_SCHEDULE_FILE.read_text(encoding="utf-8"))
    return None


def routed_request_path(raw_path: str) -> str:
    parsed = urlparse(raw_path)
    query = parse_qs(parsed.query, keep_blank_values=True)
    if parsed.path in {"/api", "/api/"} and "__path" in query:
        forwarded = (query.get("__path") or [""])[0].strip("/")
        return f"/{forwarded}" if forwarded else "/"
    return parsed.path


class AppHandler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        safe_log("[%s] %s" % (self.log_date_time_string(), fmt % args))

    def send_json(self, payload, status=200):
        raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(raw)))
        self.end_headers()
        self.wfile.write(raw)

    def send_bytes(self, payload: bytes, content_type: str, filename: str | None = None, status=200):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(payload)))
        if filename:
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.end_headers()
        self.wfile.write(payload)

    def do_GET(self):
        path = routed_request_path(self.path)
        if path == "/api/health":
            self.send_json({"ok": True, "time": now_iso(), "storage": storage_mode()})
            return
        if path == "/templates/timetable-input.xlsx":
            self.send_bytes(make_workbook_bytes(create_template_workbook()), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "timetable-input.xlsx")
            return
        if path == "/api/imports":
            self.send_json({"imports": list_imports()})
            return
        if path == "/logs/recent":
            self.send_json({"logs": read_operation_logs()})
            return
        if path == "/logs/download.txt":
            self.send_bytes(operation_logs_text(), "text/plain; charset=utf-8", "timetable-logs.txt")
            return
        if path == "/schedules/current":
            result = load_last_schedule()
            if not result:
                self.send_json({"ok": False, "error": "현재 시간표를 불러오지 못했습니다. 자동배정을 먼저 실행하세요."}, status=404)
                return
            self.send_json({"ok": True, "scheduleResult": result})
            return
        if path.startswith("/imports/") and path.endswith("/report.xlsx"):
            import_id = path.split("/")[2]
            report_bytes = load_report_bytes(import_id)
            if report_bytes is None:
                self.send_json({"error": "리포트를 찾을 수 없습니다."}, status=404)
                return
            self.send_bytes(report_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"import-{import_id}-report.xlsx")
            return
        if path == "/exports/excel":
            result = load_last_schedule()
            if not result:
                self.send_json({"error": "먼저 자동배정을 실행하세요."}, status=404)
                return
            self.send_bytes(make_workbook_bytes(export_schedule_workbook(result)), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "timetable-result.xlsx")
            return
        if path == "/exports/neis":
            result = load_last_schedule()
            if not result:
                self.send_json({"error": "먼저 자동배정을 실행하세요."}, status=404)
                return
            self.send_bytes(export_neis_csv(result), "text/csv; charset=utf-8", "neis-basic-timetable.csv")
            return
        self.serve_static(path)

    def do_POST(self):
        path = routed_request_path(self.path)
        try:
            if path == "/imports/timetable-input":
                file_name, payload = parse_multipart_file(self)
                wb = load_workbook(io.BytesIO(payload), data_only=True)
                validation = validate_workbook(wb)
                metadata = save_import(validation, file_name, payload)
                self.send_json({
                    "id": metadata["id"],
                    "createdAt": metadata["createdAt"],
                    "fileName": metadata["fileName"],
                    "ok": metadata["ok"],
                    "issues": metadata["issues"],
                    "stats": metadata["stats"],
                    "reportUrl": f"/imports/{metadata['id']}/report.xlsx",
                })
                return
            if path == "/schedules/solve":
                body = read_json_body(self)
                records = get_records_from_body(body)
                if records is None:
                    self.send_json({"error": "importId 또는 records가 필요합니다."}, status=400)
                    return
                self.send_json(solve_schedule(records, ai_config=ai_config_from_body(body, require_validated=True), solve_options=body.get("solveOptions")))
                return
            if path == "/schedules/solve/start":
                body = read_json_body(self)
                records = get_records_from_body(body)
                if records is None:
                    self.send_json({"error": "importId 또는 records가 필요합니다."}, status=400)
                    return
                self.send_json(start_solve_session(records, solve_options=body.get("solveOptions")))
                return
            if path == "/schedules/solve/continue":
                body = read_json_body(self)
                records = get_records_from_body(body)
                if records is None:
                    self.send_json({"error": "importId 또는 records가 필요합니다."}, status=400)
                    return
                result = continue_solve_session(records, as_text(body.get("sessionId")))
                self.send_json(result, status=200 if result.get("ok") else 404)
                return
            if path == "/schedules/solve/accept":
                body = read_json_body(self)
                result = accept_solve_session(as_text(body.get("sessionId")))
                self.send_json(result, status=200 if result.get("selected") else 404)
                return
            if path == "/schedules/validate":
                body = read_json_body(self)
                records = get_records_from_body(body)
                schedule = get_schedule_from_body(body)
                if records is None or schedule is None:
                    self.send_json({"error": "records/importId와 schedule이 필요합니다."}, status=400)
                    return
                self.send_json(validate_schedule(records, schedule))
                return
            if path == "/schedules/move-options":
                body = read_json_body(self)
                records = get_records_from_body(body)
                schedule = get_schedule_from_body(body)
                source = body.get("from", {})
                if records is None or schedule is None:
                    self.send_json({"error": "현재 시간표를 불러오지 못했습니다. 자동배정을 먼저 실행하세요."}, status=400)
                    return
                self.send_json(quick_move_options(records, schedule, source))
                return
            if path == "/schedules/move-preview":
                body = read_json_body(self)
                records = get_records_from_body(body)
                schedule = get_schedule_from_body(body)
                move = body.get("move", {})
                if records is None or schedule is None:
                    self.send_json({"error": "현재 시간표를 불러오지 못했습니다. 자동배정을 먼저 실행하세요."}, status=400)
                    return
                self.send_json(move_preview(records, schedule, move))
                return
            if path == "/schedules/move":
                body = read_json_body(self)
                records = get_records_from_body(body)
                schedule = get_schedule_from_body(body)
                move = body.get("move", {})
                if records is None or schedule is None:
                    self.send_json({"error": "현재 시간표를 불러오지 못했습니다. 자동배정을 먼저 실행하세요."}, status=400)
                    return
                result = move_schedule(records, schedule, move)
                save_moved_schedule_result(result, body)
                append_operation_log("manual_move", {"ok": result.get("ok"), "message": result.get("message"), "move": move})
                self.send_json(result)
                return
            if path == "/schedules/proposals/apply":
                body = read_json_body(self)
                records = get_records_from_body(body)
                if records is None:
                    self.send_json({"error": "importId 또는 records가 필요합니다."}, status=400)
                    return
                self.send_json(apply_schedule_proposal(records, body.get("proposal") or body.get("scheduleProposal") or body))
                return
            if path == "/ai/chat/local":
                body = read_json_body(self)
                records = get_records_from_body(body) or {"config": {}, "teachers": {}, "classes": {}, "subjects": {}, "rooms": {}, "loads": [], "constraints": []}
                quick_options = deepcopy(body.get("solveOptions") or {})
                quick_options["iterations"] = min(parse_positive_int(quick_options.get("iterations")) or 24, 24)
                response = ai_chat(records, body.get("message", ""), False, body.get("schedule"), ai_config={}, unassigned=body.get("unassigned") or [], solve_options=quick_options)
                response["localOnly"] = True
                append_operation_log("ai_chat_local", {
                    "proposal": bool(response.get("scheduleProposal")),
                    "constraintDrafts": len(response.get("constraintDrafts") or []),
                })
                self.send_json(response)
                return
            if path == "/ai/chat":
                body = read_json_body(self)
                records = get_records_from_body(body) or {"config": {}, "teachers": {}, "classes": {}, "subjects": {}, "rooms": {}, "loads": [], "constraints": []}
                ai_config = ai_config_from_body(body, require_validated=True)
                response = ai_chat(records, body.get("message", ""), bool(ai_config.get("apiKey") or body.get("apiKey")), body.get("schedule"), ai_config=ai_config, unassigned=body.get("unassigned") or [], solve_options=body.get("solveOptions"))
                append_operation_log("ai_chat", {
                    "provider": ai_config.get("provider"),
                    "remoteStatus": response.get("remote", {}).get("status"),
                    "proposal": bool(response.get("scheduleProposal")),
                    "constraintDrafts": len(response.get("constraintDrafts") or []),
                })
                self.send_json(response)
                return
            if path == "/ai/validate-key":
                body = read_json_body(self)
                self.send_json(validate_ai_key(body.get("aiConfig") if body.get("aiConfig") else body.get("apiKey", "")))
                return
        except Exception as exc:  # Keep the prototype honest and debuggable.
            append_operation_log("error", {"path": path, "message": str(exc)})
            self.send_json({"error": str(exc)}, status=500)
            return
        self.send_json({"error": "지원하지 않는 POST 경로입니다."}, status=404)

    def serve_static(self, path: str):
        if path == "/":
            path = "/index.html"
        target = (WEB_DIR / path.lstrip("/")).resolve()
        if not str(target).startswith(str(WEB_DIR.resolve())) or not target.exists() or target.is_dir():
            self.send_json({"error": "파일을 찾을 수 없습니다."}, status=404)
            return
        content_type, _ = mimetypes.guess_type(target.name)
        if content_type is None:
            content_type = "application/octet-stream"
        payload = target.read_bytes()
        self.send_bytes(payload, content_type)


# Vercel's Python runtime looks for a top-level HTTP handler.
handler = AppHandler


def main():
    ensure_dirs()
    port = int(os.environ.get("PORT", "8765"))
    server = ThreadingHTTPServer(("127.0.0.1", port), AppHandler)
    safe_log(f"AI timetable app running at http://127.0.0.1:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        safe_log("Stopping server.")


if __name__ == "__main__":
    main()
